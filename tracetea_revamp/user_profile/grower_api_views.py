from django.shortcuts import render
from django.shortcuts import render, redirect
from django.contrib import messages
from django.urls import reverse
from django.db.models import Q
from django.http import JsonResponse
from django.db.models.functions import Coalesce
from django.db.models import Sum, Count, Value
from django.template.loader import render_to_string
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect
from django.contrib.auth import get_user_model
User = get_user_model()
import time
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, DetailView, UpdateView, CreateView
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.paginator import Paginator , EmptyPage, PageNotAnInteger
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser, IsAuthenticatedOrReadOnly, AllowAny
from django.core.paginator import Paginator
from knox.auth import TokenAuthentication
from django.db import transaction
from rest_framework import status
from rest_framework.exceptions import APIException
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from django.contrib.auth.tokens import PasswordResetTokenGenerator
from django.utils.encoding import smart_str, force_str, smart_bytes, DjangoUnicodeDecodeError
from django.db import transaction
from django.db.models import Q
from master.common import CommonMixin
from .models import *
from master.forms import *
from accounts.forms import *
# Create your views here.
from accounts.models import *
from accounts.serializers import *
import collections
from django.contrib.auth import authenticate, login 
from knox.auth import TokenAuthentication
from rest_framework import permissions
from knox.models import AuthToken
#from AuthTokenSerializer import *
from knox.views import LoginView as KnoxLoginView
from user_profile.models import *
from user_profile.serializers import *
from .serializers import *
from chemical_data.models import *
from django.db.models import Max
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from master.utils import send_sms
import re
from .helpers import *
import pdfkit
from django.template.loader import get_template
from django.http import HttpResponse
from django.db.models import OuterRef, Subquery, Sum, IntegerField, Value, FloatField

class GrowerListGrowerAPIView(APIView):
    """ Grower list  View"""
    authentication_classes = (TokenAuthentication,)
    permission_classes=(IsAuthenticated,)
    queryset = GrowerProfile.objects.filter(is_deleted=False)
    serializer_class = GrowerProfileSerializer
    def get(self, request, *args, **kwargs):
        try:
            id = self.request.query_params.get('id', None)
    
            if id:
                queryset = self.queryset.filter(id=id).exclude(user=request.user).order_by('-id')
            else:
                queryset=self.queryset.exclude(user=request.user).order_by('-id')    
            page_size = int(request.query_params.get('page_size', 10))
            paginator = Paginator(queryset, page_size)
            page_number = self.request.query_params.get('page', 1)
            page = paginator.get_page(page_number)
            serializer = GrowerProfileSerializer(queryset, many=True)

            return Response({
                'count': paginator.count,
                'next': page.next_page_number() if page.has_next() else None,
                'previous': page.previous_page_number() if page.has_previous() else None,
                'results': queryset.values(),
            })
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')


class BlfMappedGrowerAPIView(APIView):
    """ List of Blf mapped with grower mapping View"""
    authentication_classes = (TokenAuthentication,)
    permission_classes=(IsAuthenticated,)

    queryset=GrowerProfile.objects.filter(is_deleted=False).exclude(associated_entity__isnull=True).order_by('-id')
    serializer_class = GrowerProfileSerializer
    def get(self, request, *args, **kwargs):
        try:
       
            id = self.request.query_params.get('id', None)
    
            if id:
                queryset = self.queryset.filter(user=request.user,id=id).order_by('-id')
            else:
                queryset=self.queryset.filter(user=request.user).order_by('-id')    
            page_size = int(request.query_params.get('page_size', 10))
            paginator = Paginator(queryset, page_size)
            page_number = self.request.query_params.get('page', 1)
            page = paginator.get_page(page_number)
            serializer = GrowerProfileSerializer(queryset, many=True)

            return Response({
                'count': paginator.count,
                'next': page.next_page_number() if page.has_next() else None,
                'previous': page.previous_page_number() if page.has_previous() else None,
                # 'results': serializer.data,
                # 'results':queryset.values('id','user_id','associated_entity__user__username','associated_entity','associated_entity__user__id',\
                #                            'associated_entity__user__username','associated_entity__email',\
                #                             'associated_entity__mobile_number','associated_entity__profile_type',
                #                         ),
                'result':serializer.data
            })
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')
	
# class AggregatorMappedGrowerAPIView(APIView):
#     """ List of Aggregator mapped with grower mapping View"""
#     authentication_classes = (TokenAuthentication,)
#     permission_classes=(IsAuthenticated,)

#     queryset=GrowerProfile.objects.filter(is_deleted=False).exclude(associated_aggregator__isnull=True).order_by('-id')
#     serializer_class = GrowerProfileSerializer
#     def get(self, request, *args, **kwargs):
       
       
#         id = self.request.query_params.get('id', None)
	
#         if id:
#             queryset = self.queryset.filter(user=request.user,id=id).order_by('-id')
#         else:
#             queryset=self.queryset.filter(user=request.user).order_by('-id')    
#         page_size = int(request.query_params.get('page_size', 10))
#         paginator = Paginator(queryset, page_size)
#         page_number = self.request.query_params.get('page', 1)
#         page = paginator.get_page(page_number)
#         serializer = GrowerProfileSerializer(queryset, many=True)

#         return Response({
#             'count': paginator.count,
#             'next': page.next_page_number() if page.has_next() else None,
#             'previous': page.previous_page_number() if page.has_previous() else None,
#             # 'results': serializer.data,
#             'results':queryset.values('id','user_id','associated_aggregator',\
#                                        'associated_aggregator__user__username','associated_aggregator__name','associated_aggregator__email',\
#                                         'associated_aggregator__mobile_number','associated_aggregator__profile_type','associated_aggregator__address',
#                                     ),
#         })
class SupplyToFactoryGrowerAPIView(APIView):
    """ Supply to factory from grower view"""
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset= SupplyManagement.objects.filter(is_deleted=False)
    # serializer_class = SupplyManagementSerializer
    def get(self, request, *args, **kwargs):
        try:
            id = self.request.query_params.get('id', None)   
            driver_name=  self.request.query_params.get('driver_name', None)   
            date_from= self.request.query_params.get('date_from', None)#for supply report
            date_to= self.request.query_params.get('date_to', None)#for supply report      
            if id:
                queryset = self.queryset.filter(created_by=request.user,id=id).order_by('-id')
            elif driver_name:
                queryset = self.queryset.filter(created_by=request.user,driver_name__icontains=driver_name).order_by('-id')    
            elif (date_from and date_to ) :
                 queryset=self.queryset.filter(created_by=request.user,date_of_supply__range=[date_from,date_to]).\
                    order_by('-id') 
            else:
                queryset=self.queryset.filter(created_by=request.user).order_by('-id')
                print(queryset)
            page_size = int(request.query_params.get('page_size', 10))
            paginator = Paginator(queryset, page_size)
            page_number = self.request.query_params.get('page', 1)
            page = paginator.get_page(page_number)
            serializer = SupplyManagementSerializer(queryset, many=True)
            return Response({
                'count': paginator.count,
                'next': page.next_page_number() if page.has_next() else None,
                'previous': page.previous_page_number() if page.has_previous() else None,
                
                'result':serializer.data
            })
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')
    def post(self, request, format=None):
        try:
			
            supply_to = request.data.get('supply_to', None)
            consumer_id= request.data.get('consumer_id', None)#user id of aggregator or factory    
            date_of_supply= request.data.get('date_of_supply', None)
            quantity= request.data.get('quantity', None)
            supply_bag_id= request.data.get('supply_bag_id', None)
            vehicle_option= request.data.get('vehicle_option', None)
            alloted_vehicle_id= request.data.get('alloted_vehicle_id', None)
            driver_name= request.data.get('driver_name', None)
            mobile_number= request.data.get('mobile_number', None)
  
            with transaction.atomic(): 
                created =  SupplyManagement.objects.create(supply_to=supply_to,consumer_id=consumer_id,\
                		quantity=quantity,supply_bag_id=supply_bag_id,vehicle_option=vehicle_option,\
                             date_of_supply=date_of_supply,gross_leaf=quantity,\
                               driver_name=driver_name, mobile_number=mobile_number, created_by=request.user)
                if alloted_vehicle_id:
                    if VehicleManagement.objects.filter(id=alloted_vehicle_id,is_available=True).exists():
                        created.alloted_vehicle_id=alloted_vehicle_id
                        created.save()
                        VehicleManagement.objects.filter(id=alloted_vehicle_id).update(is_available=False)#vehicle availablity status
                # current_time = int(time.time() * 1000) #convert into millisecond
                # random_number = random.randint(0, 9999)
                # created.supply_challan_id= "TR"+ f"{current_time:013d}{random_number:04d}"[:9]

                # created.supply_challan_id="Tran"+ str(random.randint(10,1999999999))
                agg_supplier_details = AggregatorProfile.cmobjects.filter(user_id=request.user.id).first()
                grower_supplier_details = GrowerProfile.cmobjects.filter(user_id=request.user.id).first()
                if agg_supplier_details:
                    if agg_supplier_details.region:
                        region_code=agg_supplier_details.region.region_id
                    else:
                        region_code= ""    
                elif grower_supplier_details:
                    if grower_supplier_details.region:
                        region_code=grower_supplier_details.region.region_id 
                    else:
                        region_code= ""    
                else :
                    region_code= ""              
            
                # last_challan = SupplyManagement.objects.filter(supply_challan_id__contains=str(region_code) + str(request.user).upper()).aggregate(Max('supply_challan_id'))
                # last_challan_id = last_challan.get('supply_challan_id__max')
                # print(last_challan_id)
                # # Calculate the next supply_challan_id
                # if last_challan_id and last_challan_id.startswith("CH"+str(region_code) + str(request.user).upper()):
                #     # Extract the numeric part and increment it
                #     print("check")

                #     # numeric_part = int(last_challan_id[-2:]) + 1
                #     numeric_part_index = len("CH" + str(region_code) + str(request.user).upper())
                #     numeric_part = int(last_challan_id[numeric_part_index:]) + 1  
                #     next_challan_id = f"CH{region_code}{str(request.user).upper()}{numeric_part:0>2d}"
                # else:
                #     # Handle the case when there are no existing supply_challan_id values
                #     next_challan_id = f"CH{region_code}{str(request.user).upper()}01"
                last_challan = SupplyManagement.objects.filter(supply_challan_id__contains="CH"+str(region_code) + str(request.user).upper(),\
                created_by_id=request.user.id).last()
                # Calculate the next supply_challan_id
                if last_challan and last_challan.supply_challan_id.startswith("CH"+str(region_code) + str(request.user).upper()):
                    # Extract the numeric part and increment it
                    numeric_part_index = len("CH" + str(region_code) + str(request.user).upper())
                 
                    numeric_part = int(last_challan.supply_challan_id[numeric_part_index:]) + 1    
                    next_challan_id = f"CH{region_code}{str(request.user).upper()}{numeric_part:0>2d}"
                else:
                    # Handle the case when there are no existing supply_challan_id values
                    next_challan_id = f"CH{region_code}{str(request.user).upper()}01"

                # Set the new supply_challan_id
                created.supply_challan_id = next_challan_id
                created.save()


         
                serializer = SupplyManagementSerializer(created)
         

                return Response({'results':{
                                            'supply_data':serializer.data,
                                        },
                                        'msg': 'Supply to factory Created Successfully',
                                        'status':status.HTTP_201_CREATED,
                                        "request_status": 1})
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},\
                                              code='authorization')
    def put(self, request, *args, **kwargs):
        """
        Below are the methods for Edit method and Delete method
        """
        try:
            method = self.request.query_params.get('method', None)
            id = self.request.query_params.get('id', None)


            with transaction.atomic():
                if method.lower() == 'edit':
                    supply_to = request.data.get('supply_to', None)
                    consumer_id= request.data.get('consumer_id', None)#user id of aggregator or factory
                    date_of_supply= request.data.get('date_of_supply', None)
                    quantity= request.data.get('quantity', None)
                    supply_bag_id= request.data.get('supply_bag_id', None)
                    vehicle_option= request.data.get('vehicle_option', None)
                    alloted_vehicle_id= request.data.get('alloted_vehicle_id', None)
                    driver_name= request.data.get('driver_name', None)
                    mobile_number= request.data.get('mobile_number', None)

                    supply_details = SupplyManagement.cmobjects.filter(pk=id).first()




                    if not supply_details:
                        return Response({'results': [],
                                        'msg': "Selected supply Does'nt exists!",
                                        "request_status": 0}, status=status.HTTP_400_BAD_REQUEST)




                    update=SupplyManagement.cmobjects.filter(id=id)\
                        .update(supply_to=supply_to,consumer_id=consumer_id,date_of_supply=date_of_supply,\
                		quantity=quantity,supply_bag_id=supply_bag_id,vehicle_option=vehicle_option,\
                            mobile_number=mobile_number ,updated_by=request.user)
                    if alloted_vehicle_id:
                        if VehicleManagement.objects.filter(id=alloted_vehicle_id,is_available=True).exists():
                            supply_details.alloted_vehicle_id=alloted_vehicle_id
                            supply_details.save()
                            VehicleManagement.objects.filter(id=alloted_vehicle_id).update(is_available=False)
                        
                        
                    if supply_details.alloted_vehicle_id and str(vehicle_option.lower()) == 'no'  :
                        vehicle_details=VehicleManagement.objects.filter(id=supply_details.alloted_vehicle_id)\
                            .update(is_available=True)    
                        vehicle_details.save() 
                    return Response({'results':{
                                            'Data':SupplyManagement.objects.filter(pk=id).values(),
                                            },
                                            'msg': 'Supply Management Data  Updated SuccessFully',
                                            'status':status.HTTP_200_OK,
                                            "request_status": 1})

                elif method.lower() == 'delete':
                    """
                    Delete 
                    """
                    if SupplyManagement.cmobjects.filter(id=id).exists():
                        supply_details= SupplyManagement.cmobjects.filter(id=id).first()
                        supply_details.is_deleted = True
                        supply_details.save()

                        return Response({'results':{
                                            'form_details':SupplyManagement.objects.filter(pk=id).values(),
                                            },
                                            'msg': 'Selected Supply Details Deleted Successfully',
                                            "request_status": 1})
                    else:
                        return Response({'results': [],
                                        'msg': "Selected Supply Details Does'nt exists!",
                                        'status': status.HTTP_404_NOT_FOUND,
                                        "request_status": 0})       
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},\
                                              code='authorization')    
        
class YearGrowerAPIView(APIView):
    """ Year API view"""
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset= Year.objects.filter(is_deleted=False)
    def get(self, request, *args, **kwargs):
        try:
            id = self.request.query_params.get('id', None)          
            if id:
                queryset = self.queryset.filter(id=id).order_by('-id')
            else:
                queryset=self.queryset.order_by('-id')    
                print(queryset)
            page_size = int(request.query_params.get('page_size', 10))
            paginator = Paginator(queryset, page_size)
            page_number = self.request.query_params.get('page', 1)
            page = paginator.get_page(page_number)
            return Response({
                'count': paginator.count,
                'next': page.next_page_number() if page.has_next() else None,
                'previous': page.previous_page_number() if page.has_previous() else None,
                'result': queryset.values( ),
            })
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')
    def post(self, request, format=None):
        try:
			
            year_name = request.data.get('year_name', None)
         
  
            with transaction.atomic(): 
                created =  Year.objects.create(year_name=year_name, created_by=request.user)
                
               
                created.save()
         
         

                return Response({       'result':created.year_name,
                                        'msg': 'Year created Created Successfully',
                                        'status':status.HTTP_201_CREATED,
                                        "request_status": 1})
            
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},\
                                              code='authorization')
    def put(self, request, *args, **kwargs):
        """
        Below are the methods for Edit method and Delete method
        """
        try:
            method = self.request.query_params.get('method', None)
            id = self.request.query_params.get('id', None)


            with transaction.atomic():
                if method.lower() == 'edit':
                    year_name = request.data.get('year_name', None)

                    year_details = Year.cmobjects.filter(pk=id).first()




                    if not year_details:
                        return Response({'results': [],
                                        'msg': "Selected YearDoes'nt exists!",
                                        "request_status": 0}, status=status.HTTP_400_BAD_REQUEST)




                    update=Year.cmobjects.filter(id=id)\
                        .update(year_name=year_name,updated_by=request.user)
                    return Response({'results':{
                                            'Data':Year.objects.filter(pk=id).values(),
                                            },
                                            'msg': 'Year  Updated SuccessFully',
                                            'status':status.HTTP_200_OK,
                                            "request_status": 1})

                elif method.lower() == 'delete':
                    """
                    Delete 
                    """
                    if Year.cmobjects.filter(id=id).exists():
                        year_details= Year.cmobjects.filter(id=id).first()
                        year_details.is_deleted = True
                        year_details.save()

                        return Response({'results':{
                                            'form_details':Year.objects.filter(pk=id).values(),
                                            },
                                            'msg': 'Selected Year Deleted Successfully',
                                            "request_status": 1})
                    else:
                        return Response({'results': [],
                                        'msg': "Selected year Details Does'nt exists!",
                                        'status': status.HTTP_404_NOT_FOUND,
                                        "request_status": 0})   
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},\
    
    
                                              code='authorization')                   
class MonthlyScheduleGrowerAPIView(APIView):
    """ Monthly Schedule of Work API view created by Aggregator and Grower """
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    def get(self, request):
        id = self.request.query_params.get('id', None)
        all = self.request.query_params.get('all', None)
        order_by = self.request.query_params.get('order_by', '-id')
        search = {}
        search['created_by'] = request.user
        search = custom_filters(self.request, search, [])
        if id:
            list_data = MonthlySchedule.cmobjects.filter(id=id).first()
            serializer = MonthlyScheduleSerializer(list_data)
            return Response(serializer.data)
        list_data = MonthlySchedule.cmobjects.filter(*search).order_by(*str(order_by).split(","))
        if all == 'true':
            serializer = MonthlyScheduleSerializer(list_data, many=True)
            return Response({'results': serializer.data})
        page_size = int(request.query_params.get('page_size', settings.MIN_PAGE_SIZE))
        paginator = Paginator(list_data, page_size)
        page_number = self.request.query_params.get('page', 1)
        page = paginator.get_page(page_number)
        serializer = MonthlyScheduleSerializer(page, many=True)
        return Response({
            'count': paginator.count,
            'next': page.next_page_number() if page.has_next() else None,
            'previous': page.previous_page_number() if page.has_previous() else None,
            'results': serializer.data,
        })

    def post(self, request):
        request.data['created_by'] = request.user.id
        serializer = MonthlyScheduleSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            try:
                serializer.save()
            except Exception as e:
                error_message = str(e.args[0]) if e.args else str(e)
                raise APIException({'request_status': 0, 'msg': error_message})
            return Response(
                {'results': {
                    'Data': serializer.data,
                },
                    'msg': 'Successfully created',
                    'status': status.HTTP_201_CREATED,
                    "request_status": 1})
        raise APIException({'request_status': 0, 'msg': serializer.errors})
    
    def put(self, request):
        method = self.request.query_params.get('method', None)
        id = self.request.query_params.get('id', None)
        details = MonthlySchedule.cmobjects.filter(pk=id).first()
        request.data['updated_by'] = request.user.id
        with transaction.atomic():
            if method.lower() == 'edit':
                if MonthlySchedule.cmobjects.filter(pk=id).exists():
                    details = MonthlySchedule.cmobjects.filter(pk=id).first()
                    serializer = MonthlyScheduleSerializer(details, data=request.data,
                                                            context={'request': request})
                    if serializer.is_valid():
                        try:
                            serializer.save()
                        except Exception as e:
                            error_message = str(e.args[0]) if e.args else str(e)
                            raise APIException({'request_status': 0, 'msg': error_message})
                        return Response({'results': {'Data': serializer.data,},
                            'msg': "Successfully updated",
                            'status': status.HTTP_202_ACCEPTED,
                            "request_status": 1})
                    raise APIException({'request_status': 0, 'msg': serializer.errors})
                else:
                    raise APIException({'request_status': 1, 'msg': "Something went wrong"})    
            elif method.lower() == 'delete':
                if MonthlySchedule.cmobjects.filter(pk=id).exists():
                    details = MonthlySchedule.cmobjects.get(pk=id)
                    details.is_deleted = True
                    details.save()
                    return Response({'results': {
                        'details': MonthlySchedule.cmobjects.filter(pk=id).values(),
                    },
                        'msg': 'Successfully deleted',
                        "request_status": 1})
                else:
                    raise APIException({'request_status': 1, 'msg': "Something went wrong"})

class BlfListGrowerAPIView(APIView):
    """ Blf list to be map with grower View"""
    authentication_classes = (TokenAuthentication,)
    permission_classes=(IsAuthenticated,)

    queryset=BlfProfile.objects.filter(is_deleted=False)    
    serializer_class = BlfProfileSerializer
    def get(self, request, *args, **kwargs):
        
        # queryset=AggregatorProfile.objects.filter(is_deleted=False).exclude(user=request.user).order_by('-id')
        try:
            id = self.request.query_params.get('id', None)
            region_id= self.request.query_params.get('region_id', None)
            state_id=self.request.query_params.get('state_id', None)
            # if id:
            #     queryset = self.queryset.filter(id=id).order_by('-id')
            if region_id and state_id:
                queryset=self.queryset.filter(region_id=region_id,state_id=state_id).order_by('-id') 
            else:
                queryset = ""     
            page_size = int(request.query_params.get('page_size', 10))
            paginator = Paginator(queryset, page_size)
            page_number = self.request.query_params.get('page', 1)
            page = paginator.get_page(page_number)
            serializer = BlfProfileSerializer(queryset, many=True)

            return Response({
                'count': paginator.count,
                'next': page.next_page_number() if page.has_next() else None,
                'previous': page.previous_page_number() if page.has_previous() else None,
                'results': serializer.data,
            })
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')
class BlfManagementGrowerAPIView(APIView):
    """ Grower  with BLF mapping View"""
    authentication_classes = (TokenAuthentication,)
    permission_classes=(IsAuthenticated,)

    queryset=GrowerProfile.objects.filter(is_deleted=False).exclude(associated_entity__isnull=True).order_by('-id')
    serializer_class = GrowerProfileSerializer
    def get(self, request, *args, **kwargs):
        try:
       
            id = self.request.query_params.get('id', None)
    
            if id:
                queryset = self.queryset.filter(user=request.user,id=id).order_by('-id')
            else:
                queryset=self.queryset.filter(user=request.user).order_by('-id')    
            page_size = int(request.query_params.get('page_size', 10))
            paginator = Paginator(queryset, page_size)
            page_number = self.request.query_params.get('page', 1)
            page = paginator.get_page(page_number)
            serializer = GrowerProfileSerializer(queryset, many=True)

            return Response({
                'count': paginator.count,
                'next': page.next_page_number() if page.has_next() else None,
                'previous': page.previous_page_number() if page.has_previous() else None,
                'results': serializer.data,
                # 'results':queryset.values('id','user_id','associated_entity',\
                #                            'associated_entity__user__username','associated_entity__email',\
                #                             'associated_entity__mobile_number','associated_entity__profile_type',\
                #                             'associated_entity__region','associated_entity__state',
                #                         ),
            })
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')
    def put(self, request, *args, **kwargs):
        try:
            method = self.request.query_params.get('method', None)
            # id = self.request.query_params.get('id', None)
            id=request.user# user id get from token verification
            with transaction.atomic():
                if method.lower() == 'edit':
                    # region_id= request.data.get('region_id', None)
                    # state_id= request.data.get('state_id', None)
                    blf_id= request.data.get('blf_id', None)
                    grower_details= GrowerProfile.cmobjects.filter(user=id).first()
                    # blf_details= BlfProfile.cmobjects.filter(id=blf_id).first()
                    print(grower_details)
                    if not grower_details:
                        return Response({'results': [],
                                        'msg': "Selected Grower Does'nt exists!",
                                        "request_status": 0}, status=status.HTTP_400_BAD_REQUEST)
                    # blf_details.region_id=region_id
                    # blf_details.state_id=state_id
                    grower_details.associated_entity.add(blf_id)
                    # agent_details.associated_entity_id=blf_id
                    grower_details.updated_by=id
                    grower_details.save()
                    # blf_details.save()
                    return Response({'results':{
                                            'Data':GrowerProfile.objects.filter(user=id).values('id','associated_entity','associated_entity__user__username'),
                                            },
                                            'msg': 'Blf or Factory  Mapped with Grower SuccessFully',
                                            'status':status.HTTP_200_OK,
                                            "request_status": 1})  
                if method.lower() == 'delete':
                    # region_id= request.data.get('region_id', None)
                    # state_id= request.data.get('state_id', None)
                    blf_id= self.request.query_params.get('blf_id', None)
                    print(blf_id)
                    grower_details= GrowerProfile.cmobjects.filter(user=id).first()
                    print(grower_details)
                    if not grower_details:
                        return Response({'results': [],
                                        'msg': "Selected Grower Does'nt exists!",
                                        "request_status": 0}, status=status.HTTP_400_BAD_REQUEST)
                    # grower_details.region_id=region_id
                    # grower_details.state_id=state_id
                    grower_details.associated_entity.remove(blf_id)
                    # agent_details.associated_entity_id=blf_id
                    grower_details.updated_by=id
                    grower_details.save()
                
                    return Response({
                                            'msg': 'Blf or Factory  Mapped with Grower Removed SuccessFully',
                                            'status':status.HTTP_200_OK,
                                            "request_status": 1})           
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')        
    

class AggregatorListGrowerAPIView(APIView):
    """ Aggregator list for map with grower """
    authentication_classes = (TokenAuthentication,)
    permission_classes=(IsAuthenticated,)
    queryset=AggregatorProfile.objects.filter(is_deleted=False)    
    serializer_class = AggregatorProfileSerializer
    def get(self, request, *args, **kwargs):
    
        # queryset=AggregatorProfile.objects.filter(is_deleted=False).exclude(user=request.user).order_by('-id')
        id = self.request.query_params.get('id', None)
        region_id = self.request.query_params.get('region_id', None)
        state_id = self.request.query_params.get('state_id', None)    
        # if id:
        #     queryset = self.queryset.filter(id=id).order_by('-id')
        if region_id and state_id:
            queryset:self.queryset.filter(region_id=region_id,state_id=state_id).order_by('-id')  # filter according to state and region   
        # else:
        #     queryset=self.queryset.order_by('-id')    
        page_size = int(request.query_params.get('page_size', 10))
        paginator = Paginator(queryset, page_size)
        page_number = self.request.query_params.get('page', 1)
        page = paginator.get_page(page_number)
        serializer = AggregatorProfileSerializer(queryset, many=True)
        return Response({
            'count': paginator.count,
            'next': page.next_page_number() if page.has_next() else None,
            'previous': page.previous_page_number() if page.has_previous() else None,
            'results': serializer.data,
        })    
    

class AggregatorManagementGrowerAPIView(APIView):
    """ Grower  with Aggregator mapping View"""
    authentication_classes = (TokenAuthentication,)
    permission_classes=(IsAuthenticated,)

    queryset=GrowerProfile.objects.filter(is_deleted=False).exclude(associated_aggregator__isnull=True).order_by('-id')
    serializer_class = GrowerProfileSerializer
    def get(self, request, *args, **kwargs):
        try:
       
            id = self.request.query_params.get('id', None)
	
            if id:
                queryset = self.queryset.filter(user=request.user,id=id).order_by('-id')
            else:
                queryset=self.queryset.filter(user=request.user).order_by('-id')    
            page_size = int(request.query_params.get('page_size', 10))
            paginator = Paginator(queryset, page_size)
            page_number = self.request.query_params.get('page', 1)
            page = paginator.get_page(page_number)
            serializer = GrowerProfileSerializer(queryset, many=True)

            return Response({
                'count': paginator.count,
                'next': page.next_page_number() if page.has_next() else None,
                'previous': page.previous_page_number() if page.has_previous() else None,
                'results': serializer.data,
                # 'results':queryset.values('id','user_id','associated_aggregator',\
                #                            'associated_aggregator__user__username','associated_aggregator__name','associated_aggregator__email',\
                #                             'associated_aggregator__mobile_number','associated_aggregator__profile_type','associated_aggregator__address',
                #                         ),
            })
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')
    def put(self, request, *args, **kwargs):
        try:
            method = self.request.query_params.get('method', None)
            # id = self.request.query_params.get('id', None)
            id=request.user# user id get from token verification
            with transaction.atomic():
                if method.lower() == 'edit':
                    mobile_number= request.data.get('mobile_number', None)
                    address= request.data.get('address', None)
                    aggregator_id= request.data.get('aggregator_id', None)


                    grower_details= GrowerProfile.cmobjects.filter(user=id).first()
                    aggregator_details=AggregatorProfile.cmobjects.filter(id=aggregator_id).first()
                    print(aggregator_details)
                    print(grower_details)
                    if not grower_details:
                        return Response({'results': [],
                                        'msg': "Selected Grower Does'nt exists!",
                                        "request_status": 0}, status=status.HTTP_400_BAD_REQUEST)
                    aggregator_details.mobile_number=mobile_number
                    aggregator_details.address=address
                    grower_details.associated_aggregator.add(aggregator_id)
                    # agent_details.associated_entity_id=blf_id
                    grower_details.updated_by=id
                    grower_details.save()
                    aggregator_details.save()
                    return Response({'results':{
                                            'Data':GrowerProfile.objects.filter(user=id).values('id','associated_aggregator','associated_aggregator__user__username',\
                                                                                                'associated_aggregator__name'),
                                            },
                                            'msg': 'Aggregator  Mapped with Grower SuccessFully',
                                            'status':status.HTTP_200_OK,
                                            "request_status": 1})      
                elif method.lower() == 'delete':
                    # mobile_number= request.data.get('mobile_number', None)
                    # address= request.data.get('address', None)
                    aggregator_id= self.request.query_params.get('aggregator_id', None)


                    grower_details= GrowerProfile.cmobjects.filter(user=id).first()
                    print(grower_details)
                    if not grower_details:
                        return Response({'results': [],
                                        'msg': "Selected Grower Does'nt exists!",
                                        "request_status": 0}, status=status.HTTP_400_BAD_REQUEST)
                    # grower_details.mobile_number=mobile_number
                    # grower_details.address=address
                    grower_details.associated_aggregator.remove(aggregator_id)
                    # agent_details.associated_entity_id=blf_id
                    grower_details.updated_by=id
                    grower_details.save()
                
                    return Response({
                                            'msg': 'Aggregator  Mapped with Grower Remove SuccessFully',
                                            'status':status.HTTP_200_OK,
                                            "request_status": 1})              
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')
        

class ChemicalRegisterReportAPIView(APIView):
    """ Chemical Registration Report API View"""
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = UseOfChemical.objects.filter(is_deleted=False)
    serializer_class = UseOfChemicalSerializer
    def get(self, request, *args, **kwargs):
        try:
            id = self.request.query_params.get('id', None) 
            date_from= self.request.query_params.get('date_from', None)
            date_to= self.request.query_params.get('date_to', None)   
            chemical_type= self.request.query_params.get('chemical_type', None)     
            if id:
                queryset = self.queryset.filter(created_by=request.user,id=id).order_by('-id')
            elif (date_from and date_to ) :
                 queryset=self.queryset.filter(created_by=request.user,date__range=[date_from,date_to],\
                                              chemical__chemical_type__name__iexact=chemical_type ).\
                    order_by('-id')   
            else:
                queryset=self.queryset.filter(created_by=request.user,chemical__chemical_type__name__iexact=chemical_type).order_by('-id')    
                print(queryset)
            page_size = int(request.query_params.get('page_size', 10))
            paginator = Paginator(queryset, page_size)
            page_number = self.request.query_params.get('page', 1)
            page = paginator.get_page(page_number)
            serializer = UseOfChemicalSerializer(queryset, many=True)

            return Response({
                'count': paginator.count,
                'next': page.next_page_number() if page.has_next() else None,
                'previous': page.previous_page_number() if page.has_previous() else None,
                


                                           
                 'result':serializer.data                          
            })
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')

class DivisionListGrowerAPIView(APIView):
    """ Division API for grower app View """
    authentication_classes = (TokenAuthentication,)
    permission_classes=(IsAuthenticated,)
    queryset = Division.objects.all()
    print(queryset)
    serializer_class = DivisionListSerializer
    def get(self, request, *args, **kwargs):
        try:
            id = self.request.query_params.get('id', None)      
            if id:
                queryset = self.queryset.filter(id=id,garden__grower__user=request.user).order_by('-id')
            else:
                queryset=self.queryset.filter(garden__grower__user=request.user).order_by('-id')  
                print(queryset)  
            page_size = int(request.query_params.get('page_size', 10))
            paginator = Paginator(queryset, page_size)
            page_number = self.request.query_params.get('page', 1)
            page = paginator.get_page(page_number)
            serializer = DivisionListSerializer(queryset, many=True)

            return Response({
                'count': paginator.count,
                'next': page.next_page_number() if page.has_next() else None,
                'previous': page.previous_page_number() if page.has_previous() else None,
                'results': serializer.data,
            })
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')
class TotalSupplyGrowerAPIView(APIView):
    """ Total  Supply by Grower"""
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    queryset= SupplyManagement.objects.filter(is_deleted=False)
    def get(self, request, *args, **kwargs):
        try:
            id = self.request.query_params.get('id', None)    
            date_from=self.request.query_params.get('date_from', None)
            date_to=self.request.query_params.get('date_to', None)
            
            if (date_from and date_to):
               
                queryset=self.queryset.filter(created_by=request.user,date_of_supply__range=[date_from,date_to]).\
                    order_by('-id')  
                total_supply=0
         
                if queryset:
                    for data in queryset:
                        if data.quantity is not None:
                            quantity=(data.quantity)
                            total_supply=total_supply+float(quantity)
                total_supply_from_grower=total_supply
              
            else:
    
                queryset=self.queryset.filter(created_by=request.user)
                total_supply=0
                if queryset:
                    for data in queryset:
                        if data.quantity is not None:
                            quantity=(data.quantity)
                            total_supply=total_supply+float(quantity)
                total_supply_from_grower=total_supply
                     
            return Response({
                          
                'total_supply_quantity_by_grower':total_supply_from_grower
            })
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')   

class LabourGrowerAPIView(APIView):
    """ Labour API list for grower"""
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = Labour.objects.filter(is_deleted=False,)
    serializer_class = Labour
    def get(self, request, *args, **kwargs):
        try:
            id = self.request.query_params.get('id', None)
            name = self.request.query_params.get('name', None) 
            type = self.request.query_params.get('type', None) 
            date_from=self.request.query_params.get('date_from', None) 
            date_to=self.request.query_params.get('date_to', None)
            grower_id= self.request.query_params.get('grower_id', None)#grower profile id
            if id:
                queryset = self.queryset.filter(grower_id=grower_id,id=id).order_by('-id')
            elif type:
                 queryset = self.queryset.filter(grower_id=grower_id,type=type).order_by('-id')    
            elif name:   
                 queryset = self.queryset.filter(grower_id=grower_id,name__icontains=name).order_by('-id') 
            elif (date_from and date_to):
                 queryset = self.queryset.filter(grower_id=grower_id,created_at__range=[date_from,date_to]).order_by('-id')        
            else:
                queryset=self.queryset.filter(grower_id=grower_id,).order_by('-id')    
                print(queryset)
            page_size = int(request.query_params.get('page_size', 10))
            paginator = Paginator(queryset, page_size)
            page_number = self.request.query_params.get('page', 1)
            page = paginator.get_page(page_number)
            serializer = LabourSerializer(queryset, many=True)

            return Response({
                'count': paginator.count,
                'next': page.next_page_number() if page.has_next() else None,
                'previous': page.previous_page_number() if page.has_previous() else None,
                'results': serializer.data,
            })
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')
        
class UseOfChemicalGrowerAPIView(APIView):
    """ Use of chemical list for grower app"""
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    queryset = UseOfChemical.objects.filter(is_deleted=False)
    serializer_class = UseOfChemicalSerializer
    def get(self, request, *args, **kwargs):
        try:
            id = self.request.query_params.get('id', None) 
            grower_id = self.request.query_params.get('grower_id', None) #for grower specific
            chemical_type=self.request.query_params.get('chemical_type', None) #for list according to chemical type
            date_from= self.request.query_params.get('date_from', None)
            date_to= self.request.query_params.get('date_to', None)         
            if id:
                queryset = self.queryset.filter(grower_id=grower_id,id=id,\
                                                chemical__chemical_type__name__iexact=chemical_type).order_by('-id')
            elif (date_from and date_to ) :
                 queryset=self.queryset.filter(grower_id=grower_id,\
                                               chemical__chemical_type__name__iexact=chemical_type,date__range=[date_from,date_to]).\
                    order_by('-id')   
            else:
                queryset=self.queryset.filter(grower_id=grower_id,\
                                              chemical__chemical_type__name__iexact=chemical_type).order_by('-id')    
                print(queryset)
            page_size = int(request.query_params.get('page_size', 10))
            paginator = Paginator(queryset, page_size)
            page_number = self.request.query_params.get('page', 1)
            page = paginator.get_page(page_number)
            serializer = UseOfChemicalSerializer(queryset, many=True)

            return Response({
                'count': paginator.count,
                'next': page.next_page_number() if page.has_next() else None,
                'previous': page.previous_page_number() if page.has_previous() else None,
                'results': serializer.data,
            })
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')        
        

class ProfileUpdateAPIView(APIView):
    # def get(self, request):
    #     id = self.request.query_params.get('id', None )
    #     if id :
    #         grower_details = GrowerProfile.objects.filter(id=id).order_by('-id')
    #     else :
    #         grower_details = GrowerProfile.objects.filter(is_deleted=False).order_by('-id')
    #     serializer = GrowerProfileSerializer(grower_details, many=True, context={'request': request})
    #     return Response({'results':{'Data':serializer.data ,
    #                                     },
    #                                     'msg': 'SuccessFull',
    #                                     'status':status.HTTP_200_OK,
    #                                     "request_status": 1},status=status.HTTP_200_OK)

    def put(self, request,*args, **kwargs):
        try:
            profile_id = request.query_params.get('profile_id', None)
            user_type = request.query_params.get('user_type', None)
            if user_type.lower() == 'grower':
                try:
                    details = GrowerProfile.objects.get(id=profile_id)
                except GrowerProfile.DoesNotExist:
                        return Response({'results':{'Data':[] ,
                                                            },
                                                                'msg':"Instance not found",
                                                                'status':status.HTTP_404_NOT_FOUND,
                                                                "request_status": 0},status=status.HTTP_404_NOT_FOUND)
                data = request.data.copy()
                data['updated_by'] = request.user.id
                serializer = GrowerProfileUpdateSerializer(details,data=data, partial=True, context={'request': request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({'results':{'Data':serializer.data ,
                                                },
                                                'msg': 'SuccessFull',
                                                'status':status.HTTP_200_OK,
                                                "request_status": 1},status=status.HTTP_200_OK)
                return Response({'results': {'Data': [],
                                    },
                                    'msg': serializer.errors,
                                    'status': status.HTTP_400_BAD_REQUEST,
                                    'request_status': 0
                                }, status=status.HTTP_400_BAD_REQUEST)
            elif user_type.lower() == 'aggregator':
                try:
                    details = AggregatorProfile.objects.get(id=profile_id)
                except AggregatorProfile.DoesNotExist:
                        return Response({'results':{'Data':[] ,
                                                            },
                                                                'msg':"Instance not found",
                                                                'status':status.HTTP_404_NOT_FOUND,
                                                                "request_status": 0},status=status.HTTP_404_NOT_FOUND)
                data = request.data.copy()
                data['updated_by'] = request.user.id
                serializer = AggregatorProfileUpdateSerializer(details,data=data, partial=True, context={'request': request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({'results':{'Data':serializer.data ,
                                                },
                                                'msg': 'SuccessFull',
                                                'status':status.HTTP_200_OK,
                                                "request_status": 1},status=status.HTTP_200_OK)
                return Response({'results': {'Data': [],
                                    },
                                    'msg': serializer.errors,
                                    'status': status.HTTP_400_BAD_REQUEST,
                                    'request_status': 0
                                }, status=status.HTTP_400_BAD_REQUEST)
            elif user_type.lower() == 'blf':
                try:
                    details = BlfProfile.objects.get(id=profile_id)
                except BlfProfile.DoesNotExist:
                        return Response({'results':{'Data':[] ,
                                                            },
                                                                'msg':"Instance not found",
                                                                'status':status.HTTP_404_NOT_FOUND,
                                                                "request_status": 0},status=status.HTTP_404_NOT_FOUND)
                data = request.data.copy()
                data['updated_by'] = request.user.id
                serializer = BlfProfileUpdateSerializer(details,data=data, partial=True, context={'request': request})
                if serializer.is_valid():
                    serializer.save()
                    return Response({'results':{'Data':serializer.data ,
                                                },
                                                'msg': 'SuccessFull',
                                                'status':status.HTTP_200_OK,
                                                "request_status": 1},status=status.HTTP_200_OK)
                return Response({'results': {'Data': [],
                                    },
                                    'msg': serializer.errors,
                                    'status': status.HTTP_400_BAD_REQUEST,
                                    'request_status': 0
                                }, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({'results': {'Data': [],
                                    },
                                    'msg': "user type required",
                                    'status': status.HTTP_400_BAD_REQUEST,
                                    'request_status': 0
                                }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,\
                                               'request_status': 0, 'msg': e},code='authorization') 
from django.db.models import Q
from accounts.views import send_otp
from django.utils import timezone  # Import Django's timezone

class MobileNumberUpdateAPIView(APIView):
    permission_classes = [AllowAny]

    # def is_mobile_number_associated(self, user_type, username, mobile_number):
    #     associated_user = None

    #     if user_type.lower() == 'grower':
    #         associated_user = GrowerProfile.objects.filter(mobile_number=mobile_number).exclude(user__username=username).first()
    #     elif user_type.lower() == 'aggregator':
    #         associated_user = AggregatorProfile.objects.filter(mobile_number=mobile_number).exclude(user__username=username).first()
    #     elif user_type.lower() == 'blf':
    #         associated_user = BlfProfile.objects.filter(mobile_number=mobile_number).exclude(user__username=username).first()

    #     return associated_user
    def is_mobile_number_associated(self, username, mobile_number):
        associated_user = None

        # grower_profile = GrowerProfile.objects.filter(Q(mobile_number=mobile_number) & ~Q(user__username=username)).first()
        grower_profile=GrowerProfile.objects.filter(mobile_number=mobile_number).exclude(user__username=username).first()
        # aggregator_profile = AggregatorProfile.objects.filter(Q(mobile_number=mobile_number) & ~Q(user__username=username)).first()
        aggregator_profile = AggregatorProfile.objects.filter(mobile_number=mobile_number).exclude(user__username=username).first()
        # blf_profile = BlfProfile.objects.filter(Q(mobile_number=mobile_number) & ~Q(user__username=username)).first()
        blf_profile = BlfProfile.objects.filter(mobile_number=mobile_number).exclude(user__username=username).first()
        if grower_profile:
            associated_user = grower_profile
        elif aggregator_profile:
            associated_user = aggregator_profile
        elif blf_profile:
            associated_user = blf_profile

        return associated_user

        return associated_user
    def put(self, request, *args, **kwargs):
        try:
            username = request.query_params.get('username')
            user_type = request.query_params.get('user_type', None)

            # Get the user instance based on the username
            user_instance = User.objects.get(username=username)

            # Get the new mobile number from request data
            new_mobile_number = request.data.get('mobile_number')

            # Check if the new mobile number is associated with any user of the specified type
            associated_user = self.is_mobile_number_associated(username, new_mobile_number)

            if associated_user:
                return Response({'msg': f"Mobile number {new_mobile_number} is already associated with another user"},
                                status=status.HTTP_400_BAD_REQUEST)

            # Update the mobile number for the specified user type
            if user_type.lower() == 'grower':
                profile = GrowerProfile.objects.get(user=user_instance)
                serializer = GrowerProfileUpdateSerializer(profile, data=request.data, partial=True)
            elif user_type.lower() == 'aggregator':
                profile = AggregatorProfile.objects.get(user=user_instance)
                serializer = AggregatorProfileUpdateSerializer(profile, data=request.data, partial=True)
            elif user_type.lower() == 'blf':
                profile = BlfProfile.objects.get(user=user_instance)
                serializer = BlfProfileUpdateSerializer(profile, data=request.data, partial=True)
            else:
                return Response({'msg': "Invalid user type"},
                                status=status.HTTP_400_BAD_REQUEST)

            if serializer.is_valid():
                serializer.save()
                otp = send_otp(new_mobile_number)
                if otp:
                    # Save the OTP to the user's profile in the database
                    profile.otp = otp
                    profile.otp_created_at = timezone.now() # Save OTP creation time
                    # ist_time = timezone.now().astimezone(timezone.pytz.timezone('Asia/Kolkata'))  # Convert to IST
                    # user_details.otp_created_at = ist_time
                    profile.save()		
                    # send sms #######
                    if new_mobile_number:
                        template_id='1007983856203558873'
                        message=f"Hi, Your one time password is: {otp}. Please don't share this with anyone - Trustea."
                        print(message)
                        send_sms(new_mobile_number, message,template_id)	
                return Response({'msg': f"Mobile number updated successfully for {user_type} and otp sent",
                                 'mobile_number': new_mobile_number,
                                 'otp':otp,
								'username':profile.user.username if profile.user.username else '',
								'user_id':profile.user.id if profile.user else None,
								'user_type':profile.profile_type.name if profile.profile_type else "",
                                 
                                 },
                                status=status.HTTP_200_OK)
            else:
                return Response({'msg': serializer.errors},
                                status=status.HTTP_400_BAD_REQUEST)

        except User.DoesNotExist:
            return Response({'msg': "User not found"},
                            status=status.HTTP_404_NOT_FOUND)

        except (GrowerProfile.DoesNotExist, AggregatorProfile.DoesNotExist, BlfProfile.DoesNotExist) as e:
            return Response({'msg': "Instance not found"},
                            status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return Response({'msg': str(e)},
                            status=status.HTTP_400_BAD_REQUEST)    
class SupplyReportListGrowerAPIView(APIView):
    """Aggregator supply report list """
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    def get(self, request):
        try:
            id = self.request.query_params.get('id', None) 
            date_from= self.request.query_params.get('date_from', None)
            date_to= self.request.query_params.get('date_to', None) 
            user_type=self.request.query_params.get('consumer_type', "") 
            total_supply_quantity_sum = 0
            if user_type.lower() == 'factory' :      
                if id:
                    supply_data = SupplyManagement.objects.filter(is_deleted=False,created_by=request.user, id=id,supply_to__iexact='Factory').values('id', 'consumer', 'alloted_vehicle', 'quantity', 'date_of_supply', 'supply_to', 'supply_challan_id',\
                                                                                                                                                'gross_leaf','supply_bag_id','driver_name').order_by('-id')
                elif (date_from and date_to ) :  
                    supply_data = SupplyManagement.objects.filter(is_deleted=False,created_by=request.user, date_of_supply__range=[date_from,date_to],\
                                                   supply_to__iexact='Factory').values('id', 'consumer', 'alloted_vehicle', 'quantity', 'date_of_supply', 'supply_to', 'supply_challan_id',\
                                                                                                                                                'gross_leaf','supply_bag_id','driver_name').order_by('-id')

                else:
                    supply_data = SupplyManagement.objects.filter(is_deleted=False,created_by=request.user,supply_to__iexact='Factory').values('id', 'consumer', 'alloted_vehicle', 'quantity', 'date_of_supply', 'supply_to', 'supply_challan_id',\
                                                                                                                                                'gross_leaf','supply_bag_id','driver_name').order_by('-id')
                grower_details=GrowerProfile.objects.filter(user=request.user).first()
                if grower_details:
                    grower_data = {
                        'id':grower_details.id  ,
                        'grower_name': grower_details.name if grower_details else '', 
                        'grower_user_name': grower_details.user.username if grower_details.user else '', 
                        'grower_type': grower_details.grower_type if grower_details else '' ,
                        'grower_address': grower_details.address if grower_details else '',
                        # Add more fields as needed
                        }
                else:    
                    grower_data = {
                            'id':"" ,
                            'grower_name':"", 
                            'grower_user_name':"",
                            'grower_type': "",
                            'grower_address': "",
                            # Add more fields as needed
                            }
                response_data = []
                for data in supply_data:
                
                    consumer=BlfProfile.objects.filter(user_id=data['consumer']).first()
                    alloted_vehicle=VehicleManagement.objects.filter(id= data['alloted_vehicle']).first()
                    if data['quantity'] is not None and str(data['quantity']).strip():
                        total_supply_quantity_sum += float(data['quantity'])
                        
                    response_data.append({
                    'consumer_name': consumer.entity_unit if consumer else "",
                    'consumer_username': consumer.user.username if consumer.user else "",
                    'consumer_id': consumer.id if consumer else "",
                    'vehicle_number': alloted_vehicle.vehicle_number if alloted_vehicle else "",
                   
                    'date_of_supply': data['date_of_supply'],
                    'supply_to': data['supply_to'],
                    'supply_challan_id': data['supply_challan_id'],
                    'gross_leaf': data['gross_leaf'],
                    'quantity':  data['quantity'],
                    'supply_bag_id': data['supply_bag_id'],
                    'driver_name':data['driver_name']
                    })

               
                # serializer = SupplyDetailsSerializer(response_data, many=True)
                return Response({
                            'result': response_data,
                            'grower_details':grower_data,
                            "date_from":date_from  if date_from else "",
                            'date_to':date_to if date_to else "",
                            'total_supply_quantity_sum':total_supply_quantity_sum
                        },status=status.HTTP_200_OK)      
            elif user_type.lower() == 'aggregator' :      
                if id:
                    supply_data = SupplyManagement.objects.filter(is_deleted=False,created_by=request.user, id=id,supply_to__iexact='Aggregator').values('id', 'consumer', 'alloted_vehicle', 'quantity', 'date_of_supply', 'supply_to', 'supply_challan_id',\
                                                                                                                                                'gross_leaf','supply_bag_id','driver_name').order_by('-id')
                elif (date_from and date_to ) :

                    supply_data = SupplyManagement.objects.filter(is_deleted=False,created_by=request.user, date_of_supply__range=[date_from,date_to],\
                                                   supply_to__iexact='Aggregator').values('id', 'consumer', 'alloted_vehicle', 'quantity', 'date_of_supply', 'supply_to', 'supply_challan_id',\
                                                                                                                                                'gross_leaf','supply_bag_id','driver_name').order_by('-id')

                else:
                    supply_data = SupplyManagement.objects.filter(is_deleted=False,created_by=request.user,supply_to__iexact='Aggregator').values('id', 'consumer', 'alloted_vehicle', 'quantity', 'date_of_supply', 'supply_to', 'supply_challan_id',\
                                                                                                                                                'gross_leaf','supply_bag_id','driver_name').order_by('-id')
                grower_details=GrowerProfile.objects.filter(user=request.user).first()
                if grower_details:
                    grower_data = {
                        'id':grower_details.id  ,
                        'grower_name': grower_details.name if grower_details else '', 
                        'grower_user_name': grower_details.user.username if grower_details.user else '',
                        'grower_type': grower_details.grower_type if grower_details else '' ,
                        'grower_address': grower_details.address if grower_details else '',
                        # Add more fields as needed
                        }
                else :
                    grower_data = {
                        'id':"" ,
                        'grower_name':"", 
                        'grower_user_name': "",
                        'grower_type': "",
                        'grower_address': "",
                        # Add more fields as needed
                        }

                response_data = []
                for data in supply_data:
                    
                    consumer=AggregatorProfile.objects.filter(user_id=data['consumer']).first()
                    alloted_vehicle=VehicleManagement.objects.filter(id= data['alloted_vehicle']).first()
                    if data['quantity'] is not None and str(data['quantity']).strip():
                        total_supply_quantity_sum += float(data['quantity'])
                    response_data.append({
                    'consumer_name': consumer.name if consumer else None,
                    'consumer_id': consumer.id if consumer else None,
                    'consumer_user_name': consumer.user.username if consumer.user else None,
                    'vehicle_number': alloted_vehicle.vehicle_number if alloted_vehicle else None,
           
                    'date_of_supply': data['date_of_supply'],
                    'supply_to': data['supply_to'],
                    'supply_challan_id': data['supply_challan_id'],
                    'gross_leaf': data['gross_leaf'],
                    'quantity':  data['quantity'],
                    'supply_bag_id': data['supply_bag_id'],
                    'driver_name':data['driver_name'],
                    })

              
                # serializer = SupplyDetailsSerializer(response_data, many=True)
                return Response({
                
                            'result': response_data,
                            'grower_details':grower_data,
                            "date_from":date_from  if date_from else "",
                            'date_to':date_to if date_to else "",
                            'total_supply_quantity_sum':total_supply_quantity_sum
                        },status=status.HTTP_200_OK)        
            else:
                return Response({'msg': 'Consumer type required either factory or aggregator.'},\
                                          status=status.HTTP_400_BAD_REQUEST)
            # total_supply_quantity = SupplyManagement.objects.filter(is_deleted=False).aggregate(total=Sum('quantity'))['total']
        
        except Exception as e:
            raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')


############ Update phone number ###################
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from django.conf import settings
from django.db.models.functions import Lower
from django.contrib.auth.hashers import make_password
from typing import Union

def generate_grower_password(full_name: str, production_area: Union[int, str]) -> str:
    full_name = full_name.strip().upper()
    production_area = str(production_area).strip()

    production_digits = ''.join(filter(str.isdigit, production_area))
    if not production_digits:
        raise ValueError("Production area must contain at least one digit.")
    
    suffix = production_digits[0]  # take first digit
    name_parts = [part for part in full_name.split() if part]
    if not name_parts:
        raise ValueError("Full name cannot be empty.")
    if len(name_parts) == 1:
        first_part = name_parts[0][:4]
        password = first_part + suffix
    else:
        first_part = name_parts[0][:2]
        last_part = name_parts[-1][:2]
        password = first_part + last_part + suffix
    password = password.replace('.', "").replace(" ", "")
    return password

def generate_grower_username(full_name: str, village: str, serial_number: int) -> str:
    full_name = full_name.strip().upper()
    village = village.strip().upper()
    
    name_parts = [part for part in full_name.split() if part]
    if not name_parts:
        raise ValueError("Full name cannot be empty.")
    
    first_name = name_parts[0]
    if len(name_parts) >= 2:
        last_name = name_parts[-1]
        last_part = last_name[:2]
    else:
        last_part = ""
    
    if len(first_name) == 1:
        first_part = first_name 
    else:
        first_part = first_name[:2]
    
    village_part = village[:2]
    prefix = (first_part + last_part + village_part) if last_part else (first_part + village_part)
    username = prefix + f"{serial_number:02d}"
    
    check_username = User.objects.annotate(
        username_lower=Lower('username')
    ).filter(
        username_lower=username.lower()
    ).first()
    
    if check_username:
        for i in range(2, len(village)):
            prefix = (first_part + last_part + village[:i+1]) if last_part else (first_part + village[:i+1])
            username = prefix + f"{serial_number:02d}"
            check_username = User.objects.annotate(
                username_lower=Lower('username')
            ).filter(
                username_lower=username.lower()
            ).first()
            if not check_username:
                break
    username = username.replace('.', "").replace(" ", "")
    return username

class GenerateGrowerUsernameAPIView(APIView):
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    def post(self, request, format=None):
        try:
            # Validate serial_number
            serial_param = request.query_params.get('serial_number')
            if serial_param is None:
                return Response({
                    'request_status': 0,
                    'msg': "Missing 'serial_number' query parameter (e.g., ?serial_number=270)"
                }, status=400)
            try:
                start_serial = int(serial_param)
                if start_serial < 1:
                    raise ValueError("Serial number must be positive")
            except (ValueError, TypeError):
                return Response({
                    'request_status': 0,
                    'msg': "'serial_number' must be a valid positive integer"
                }, status=400)

            # Validate file_name
            file_name = request.data.get('file_name')
            if not file_name:
                return Response({'request_status': 0, 'msg': "Missing 'file_name'"}, status=400)

            file_path = os.path.join('media', 'excel', file_name)
            if not os.path.exists(file_path):
                return Response({
                    'request_status': 0,
                    'msg': f"File '{file_name}' not found on server."
                }, status=400)
            # Load Excel
            try:
                excel_file = pd.ExcelFile(file_path)
            except Exception as e:
                return Response({
                    'request_status': 0,
                    'msg': f"Failed to read Excel file: {str(e)}"
                }, status=400)

            # Initialize result
            result = {
                'total_count': 0,
                'sheets_processed': [],
                'usernames': [],
                'passwords': [],
                'errors': []
            }

            current_serial = start_serial

            # Process sheets
            for sheet_name in excel_file.sheet_names:
                result['sheets_processed'].append(sheet_name)
                df = pd.read_excel(excel_file, sheet_name=sheet_name)

                # Clean data
                df = df.fillna("")
                df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                df.replace([None, "", "nan", np.nan], None, inplace=True)

                # Iterate rows
                for idx, row in df.iterrows():
                    row_num = idx + 2  # Excel row number (header = row 1)
                    try:
                        name = row.get('name')
                        village = row.get('village_or_town')
                        production_area = row.get('production_area')

                        if not name or pd.isna(name):
                            raise ValueError("Missing 'name'")
                        if not village or pd.isna(village):
                            raise ValueError("Missing 'village_or_town'")

                        name = str(name).strip()
                        village = str(village).strip()

                        if not name or not village:
                            raise ValueError("'name' and 'village_or_town' cannot be empty")

                        username = generate_grower_username(name, village, current_serial)
                        password = generate_grower_password(name, production_area)



                        
                        result['usernames'].append(username)
                        result['passwords'].append(password)
                        result['total_count'] += 1

                        current_serial += 1 

                    except Exception as e:
                        result['errors'].append(f"Row {row_num}: {str(e)}")
                        continue

            return Response({
                'request_status': 1,
                'msg': 'Grower usernames generated successfully.',
                'data': result
            }, status=200)

        except Exception as e:
            return Response({
                'request_status': 0,
                'msg': f"Unexpected error: {str(e)}"
            }, status=500)



IMPORT_BATCH_SIZE = 1000
MAX_IMPORT_ERRORS_RETURNED = 200


def _normalise_key(value):
    if value is None or pd.isna(value):
        return ""
    return str(value).strip().lower()


def _clean_value(value):
    if value is None or pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _clean_str(value):
    value = _clean_value(value)
    if value is None:
        return None
    return str(value).strip()


def _whole_str(value):
    value = _clean_value(value)
    if value is None:
        return None
    try:
        return str(int(round(float(value))))
    except (ValueError, TypeError):
        return str(value).strip()


def _int_or_none(value):
    value = _clean_value(value)
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def _date_or_none(value):
    value = _clean_value(value)
    if value is None:
        return None
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return None


def _choice_or_none(value, choices, title=False):
    value = _clean_str(value)
    if not value:
        return None
    value = value.title() if title else value
    return value if value in dict(choices).keys() else None


def _get_import_file_path(file_name):
    if not file_name:
        raise ValueError("Missing 'file_name' in request data.")
    file_path = os.path.join('media', 'excel', file_name)
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File '{file_name}' not found on server.")
    return file_path


def _iter_excel_records(file_path, required_columns):
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                yield worksheet.title, None, None, "Sheet is empty."
                continue

            columns = [str(header).strip() if header is not None else "" for header in headers]
            missing_columns = set(required_columns) - set(columns)
            if missing_columns:
                yield worksheet.title, None, None, f"Missing required columns: {sorted(missing_columns)}"
                continue

            for row_number, values in enumerate(rows, start=2):
                row = dict(zip(columns, values))
                if not any(_clean_value(value) is not None for value in row.values()):
                    continue
                yield worksheet.title, row_number, row, None
    finally:
        workbook.close()


def _first_by_key(queryset, field_name):
    cached = {}
    for obj in queryset:
        key = _normalise_key(getattr(obj, field_name, None))
        if key and key not in cached:
            cached[key] = obj
    return cached


def _append_import_error(errors, count_data, message):
    count_data['errors'] += 1
    if len(errors) < MAX_IMPORT_ERRORS_RETURNED:
        errors.append(message)


def _bulk_replace_m2m(field, source_ids, pairs):
    if not source_ids:
        return
    through = field.remote_field.through
    source_id_field = f"{field.m2m_field_name()}_id"
    target_id_field = f"{field.m2m_reverse_field_name()}_id"

    through.objects.filter(**{f"{source_id_field}__in": source_ids}).delete()
    through.objects.bulk_create(
        [
            through(**{source_id_field: source_id, target_id_field: target_id})
            for source_id, target_id in pairs
            if source_id and target_id
        ],
        batch_size=IMPORT_BATCH_SIZE,
        ignore_conflicts=True,
    )


def _sync_users(user_records):
    usernames = [record['username'] for record in user_records]
    existing_users = User.objects.in_bulk(usernames, field_name='username')
    users_to_create = []
    users_to_update = []

    for record in user_records:
        user = existing_users.get(record['username'])
        password_hash = make_password(record['password'])
        if user:
            user.first_name = (record.get('name') or '')[:30]
            user.password = password_hash
            user.is_active = True
            users_to_update.append(user)
        else:
            users_to_create.append(User(
                username=record['username'],
                first_name=(record.get('name') or '')[:30],
                password=password_hash,
                is_active=True,
            ))

    if users_to_create:
        User.objects.bulk_create(users_to_create, batch_size=IMPORT_BATCH_SIZE, ignore_conflicts=True)
    if users_to_update:
        User.objects.bulk_update(users_to_update, ['first_name', 'password', 'is_active'], batch_size=IMPORT_BATCH_SIZE)

    return User.objects.in_bulk(usernames, field_name='username')


def _sync_common_profiles(profile_records, user_type_obj, is_grower=True):
    user_ids = [record['user'].id for record in profile_records if record.get('user')]
    existing_profiles = Profile.objects.in_bulk(user_ids, field_name='user_id')
    profiles_to_create = []
    profiles_to_update = []

    for record in profile_records:
        user = record['user']
        profile = existing_profiles.get(user.id)
        defaults = {
            'full_name': record.get('name'),
            'user_type': user_type_obj,
            'phone_no': record.get('mobile_number'),
        }
        if is_grower:
            defaults.update({
                'region': record.get('region'),
                'state': record.get('state'),
                'district': record.get('district'),
                'address': record.get('address'),
                'trustea_id': record.get('tea_board_id'),
            })

        if profile:
            for field_name, value in defaults.items():
                setattr(profile, field_name, value)
            profiles_to_update.append(profile)
        else:
            profiles_to_create.append(Profile(user=user, **defaults))

    if profiles_to_create:
        Profile.objects.bulk_create(profiles_to_create, batch_size=IMPORT_BATCH_SIZE, ignore_conflicts=True)
    if profiles_to_update:
        fields = ['full_name', 'user_type', 'phone_no']
        if is_grower:
            fields += ['region', 'state', 'district', 'address', 'trustea_id']
        Profile.objects.bulk_update(profiles_to_update, fields, batch_size=IMPORT_BATCH_SIZE)


def _sync_gardens_and_plots(grower_records, profile_by_user_id):
    grower_ids = [profile_by_user_id[record['user'].id].id for record in grower_records if record['user'].id in profile_by_user_id]
    existing_gardens = {}
    for garden in Gardens.objects.filter(grower_id__in=grower_ids, is_deleted=False).order_by('id'):
        existing_gardens.setdefault(garden.grower_id, garden)

    gardens_to_create = []
    gardens_to_update = []
    for record in grower_records:
        profile = profile_by_user_id.get(record['user'].id)
        if not profile:
            continue
        garden = existing_gardens.get(profile.id)
        garden_name = record.get('garden_name') or profile.name
        if garden:
            garden.user = record['user']
            garden.name = garden_name
            garden.is_division = False
            garden.is_plot = True
            garden.production_area = record.get('production_area')
            gardens_to_update.append(garden)
        else:
            gardens_to_create.append(Gardens(
                grower=profile,
                user=record['user'],
                name=garden_name,
                is_division=False,
                is_plot=True,
                production_area=record.get('production_area'),
            ))

    if gardens_to_create:
        Gardens.objects.bulk_create(gardens_to_create, batch_size=IMPORT_BATCH_SIZE)
    if gardens_to_update:
        Gardens.objects.bulk_update(gardens_to_update, ['user', 'name', 'is_division', 'is_plot', 'production_area'], batch_size=IMPORT_BATCH_SIZE)

    gardens = {}
    for garden in Gardens.objects.filter(grower_id__in=grower_ids, is_deleted=False).order_by('id'):
        gardens.setdefault(garden.grower_id, garden)

    existing_plots = {}
    garden_ids = [garden.id for garden in gardens.values()]
    for plot in Plot.objects.filter(garden_id__in=garden_ids).order_by('id'):
        existing_plots.setdefault(plot.garden_id, plot)

    plots_to_create = []
    plots_to_update = []
    for record in grower_records:
        profile = profile_by_user_id.get(record['user'].id)
        garden = gardens.get(profile.id) if profile else None
        if not garden:
            continue
        plot = existing_plots.get(garden.id)
        plot_name = record.get('name') or garden.name
        if plot:
            plot.name = plot_name
            plot.plot_status = True
            plots_to_update.append(plot)
        else:
            plots_to_create.append(Plot(garden=garden, name=plot_name, plot_status=True))

    if plots_to_create:
        Plot.objects.bulk_create(plots_to_create, batch_size=IMPORT_BATCH_SIZE)
    if plots_to_update:
        Plot.objects.bulk_update(plots_to_update, ['name', 'plot_status'], batch_size=IMPORT_BATCH_SIZE)


class ImportGrowerExcelView(APIView):
    """Optimized grower import API view for large Excel files."""
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    def post(self, request, format=None):
        try:
            file_path = _get_import_file_path(request.data.get('file_name'))
            user_type_obj = ProfileType.objects.filter(name='grower').first()
            if not user_type_obj:
                raise ValueError("ProfileType 'grower' not found in database.")

            count_data = {
                'items_created': 0,
                'items_updated': 0,
                'errors': 0,
                'sheets_processed': [],
            }
            errors = []
            batch = []
            seen_sheets = set()

            regions = _first_by_key(Region.cmobjects.all(), 'region_name')
            states = _first_by_key(State.objects.filter(is_deleted=False), 'name')
            districts = _first_by_key(District.objects.filter(is_deleted=False), 'name')
            entities = _first_by_key(BlfProfile.cmobjects.all(), 'entity_unit')
            aggregators = _first_by_key(AggregatorProfile.cmobjects.select_related('user'), 'name')

            def flush_batch():
                if not batch:
                    return
                with transaction.atomic():
                    users = _sync_users(batch)
                    for record in batch:
                        record['user'] = users.get(record['username'])

                    user_ids = [record['user'].id for record in batch if record.get('user')]
                    existing_profiles = GrowerProfile.objects.in_bulk(user_ids, field_name='user_id')
                    profiles_to_create = []
                    profiles_to_update = []

                    for record in batch:
                        user = record.get('user')
                        if not user:
                            continue
                        profile = existing_profiles.get(user.id)
                        defaults = {
                            'profile_type': user_type_obj,
                            'name': record.get('name'),
                            'age': record.get('age'),
                            'gender': record.get('gender'),
                            'date_of_birth': record.get('date_of_birth'),
                            'grower_type': record.get('grower_type'),
                            'region': record.get('region'),
                            'state': record.get('state'),
                            'district': record.get('district'),
                            'village_or_town': record.get('village_or_town'),
                            'address': record.get('address'),
                            'mobile_number': record.get('mobile_number'),
                            'total_male_worker': record.get('total_male_worker'),
                            'total_female_worker': record.get('total_female_worker'),
                            'estimated_production_of_green_tea': record.get('estimated_production_of_green_tea'),
                            'estimated_production_of_made_tea': record.get('estimated_production_of_made_tea'),
                            'garden_name': record.get('garden_name'),
                            'production_area': record.get('production_area'),
                            'is_active': True,
                        }
                        if profile:
                            for field_name, value in defaults.items():
                                setattr(profile, field_name, value)
                            profiles_to_update.append(profile)
                        else:
                            profiles_to_create.append(GrowerProfile(user=user, **defaults))

                    if profiles_to_create:
                        GrowerProfile.objects.bulk_create(profiles_to_create, batch_size=IMPORT_BATCH_SIZE)
                        count_data['items_created'] += len(profiles_to_create)
                    if profiles_to_update:
                        GrowerProfile.objects.bulk_update(
                            profiles_to_update,
                            [
                                'profile_type', 'name', 'age', 'gender', 'date_of_birth', 'grower_type',
                                'region', 'state', 'district', 'village_or_town', 'address', 'mobile_number',
                                'total_male_worker', 'total_female_worker', 'estimated_production_of_green_tea',
                                'estimated_production_of_made_tea', 'garden_name', 'production_area', 'is_active'
                            ],
                            batch_size=IMPORT_BATCH_SIZE,
                        )
                        count_data['items_updated'] += len(profiles_to_update)

                    profile_by_user_id = GrowerProfile.objects.in_bulk(user_ids, field_name='user_id')
                    _sync_common_profiles(batch, user_type_obj, is_grower=True)
                    _sync_gardens_and_plots(batch, profile_by_user_id)

                    source_ids = [profile.id for profile in profile_by_user_id.values()]
                    agg_pairs = []
                    entity_pairs = []
                    for record in batch:
                        profile = profile_by_user_id.get(record['user'].id)
                        if not profile:
                            continue
                        agg_pairs.extend((profile.id, aggregator.id) for aggregator in record['aggregators'])
                        entity_pairs.append((profile.id, record['entity'].id))

                    _bulk_replace_m2m(GrowerProfile._meta.get_field('associated_aggregator'), source_ids, agg_pairs)
                    _bulk_replace_m2m(GrowerProfile._meta.get_field('associated_entity'), source_ids, entity_pairs)
                batch.clear()

            required_columns = {
                'username', 'name', 'region', 'state', 'district', 'associated_entity',
            }
            for sheet_name, row_num, row, sheet_error in _iter_excel_records(file_path, required_columns):
                if sheet_name not in seen_sheets:
                    seen_sheets.add(sheet_name)
                    count_data['sheets_processed'].append(sheet_name)
                if sheet_error:
                    _append_import_error(errors, count_data, f"Sheet '{sheet_name}': {sheet_error}")
                    continue

                username = _clean_str(row.get('username'))
                if not username:
                    _append_import_error(errors, count_data, f"Sheet '{sheet_name}' Row {row_num}: Username is required.")
                    continue

                region = regions.get(_normalise_key(row.get('region')))
                state = states.get(_normalise_key(row.get('state')))
                district = districts.get(_normalise_key(row.get('district')))
                entity = entities.get(_normalise_key(row.get('associated_entity')))

                if not region:
                    _append_import_error(errors, count_data, f"Sheet '{sheet_name}' Row {row_num}: Region '{_clean_str(row.get('region'))}' not found.")
                    continue
                if not state:
                    _append_import_error(errors, count_data, f"Sheet '{sheet_name}' Row {row_num}: State '{_clean_str(row.get('state'))}' not found.")
                    continue
                if not district:
                    _append_import_error(errors, count_data, f"Sheet '{sheet_name}' Row {row_num}: District '{_clean_str(row.get('district'))}' not found.")
                    continue
                if not entity:
                    _append_import_error(errors, count_data, f"Sheet '{sheet_name}' Row {row_num}: BLF Entity '{_clean_str(row.get('associated_entity'))}' not found.")
                    continue

                aggregator_names = [
                    name.strip() for name in (_clean_str(row.get('associated_aggregator')) or '').split('/') if name.strip()
                ]
                found_aggregators = []
                missing_aggregators = []
                for aggregator_name in aggregator_names:
                    aggregator = aggregators.get(_normalise_key(aggregator_name))
                    if aggregator:
                        found_aggregators.append(aggregator)
                    else:
                        missing_aggregators.append(aggregator_name)
                if missing_aggregators:
                    _append_import_error(errors, count_data, f"Sheet '{sheet_name}' Row {row_num}: Aggregator(s) not found: {', '.join(missing_aggregators)}.")
                    continue

                if any(item['username'] == username for item in batch):
                    flush_batch()

                batch.append({
                    'username': username,
                    'password': _clean_str(row.get('password')) or '123',
                    'name': _clean_str(row.get('name')) or '',
                    'age': _int_or_none(row.get('age')),
                    'gender': _choice_or_none(row.get('gender'), GrowerProfile.GENDER_TYPE, title=True),
                    'date_of_birth': _date_or_none(row.get('date_of_birth')),
                    'grower_type': _choice_or_none(row.get('grower_type'), GrowerProfile.grower_option),
                    'region': region,
                    'state': state,
                    'district': district,
                    'village_or_town': _clean_str(row.get('village_or_town')),
                    'address': _clean_str(row.get('address')),
                    'mobile_number': _whole_str(row.get('mobile_number')),
                    'total_male_worker': _int_or_none(row.get('total_male_worker')),
                    'total_female_worker': _int_or_none(row.get('total_female_worker')),
                    'estimated_production_of_green_tea': _whole_str(row.get('estimated_production_of_green_tea')),
                    'estimated_production_of_made_tea': _whole_str(row.get('estimated_production_of_made_tea')),
                    'garden_name': _clean_str(row.get('garden_name')),
                    'production_area': _clean_str(row.get('production_area')),
                    'entity': entity,
                    'aggregators': found_aggregators,
                })

                if len(batch) >= IMPORT_BATCH_SIZE:
                    flush_batch()

            flush_batch()
            count_data['total_processed'] = count_data['items_created'] + count_data['items_updated']
            count_data['errors_returned'] = len(errors)

            return Response({
                'request_status': 1,
                'msg': 'Grower data imported successfully.',
                'data': count_data,
                'errors': errors,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            print(f"Import error: {e}")
            raise APIException({'request_status': 0, 'msg': str(e)})


class ImportGrowerExcelViewBack(APIView):
    """Grower Import API View"""
    # authentication_classes = (TokenAuthentication,)
    # permission_classes = (IsAuthenticated,)

    def post(self, request, format=None):
        try:
            with transaction.atomic():
                count_data = {
                    'items_created': 0,
                    'items_updated': 0,
                    'errors': 0,
                    'sheets_processed': [],
                    'created_users' : [],
                    'updated_users' : [],
                }
                errors = []
                
                blf_username = request.data.get('blf_username')
                blf_details = BlfProfile.cmobjects.filter(user__username=blf_username).first()
                file_name = request.data.get('file_name')
                if not file_name:
                    raise APIException({'request_status': 0, 'msg': "Missing file_name"})

                file_path = os.path.join('media', 'excel', file_name)
                if not os.path.exists(file_path):
                    raise APIException({
                        'request_status': 0,
                        'msg': f"File '{file_name}' not found on server."
                    })
                
                try:
                    excel_file = pd.ExcelFile(file_path)
                except Exception as e:
                    raise APIException({
                        'request_status': 0,
                        'msg': f"Error opening Excel file: {str(e)}"
                    })
                
                def to_whole_str(value):
                    if pd.isna(value):
                        return None
                    try:
                        num = float(value)
                        return str(int(round(num))) 
                    except (ValueError, TypeError):
                        return str(value).strip()

                for sheet_name in excel_file.sheet_names:
                    count_data['sheets_processed'].append(sheet_name)
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)

                    df = df.fillna("")
                    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                    df.replace(["", "nan", np.nan], None, inplace=True)

                    for index, row in df.iterrows():
                        row_num = index + 2  
                        try:
                            username = str(row['username']).strip() if row['username'] else None
                            if not username:
                                raise ValueError("Username is required")

                            password = str(row['password']).strip() if row['password'] else '123'
                            name = str(row['name']).strip() if row['name'] else ''
                            age = int(row['age']) if pd.notna(row['age']) and str(row['age']).isdigit() else None

                            gender = str(row['gender']).strip().title() if pd.notna(row['gender']) else None
                            if gender not in dict(GrowerProfile.GENDER_TYPE).keys():
                                gender = None

                            dob = None
                            if pd.notna(row['date_of_birth']):
                                try:
                                    dob = pd.to_datetime(row['date_of_birth']).date()
                                except:
                                    pass

                            grower_type = str(row['grower_type']).strip() if pd.notna(row['grower_type']) else None
                            if grower_type not in dict(GrowerProfile.grower_option).keys():
                                grower_type = None

                            region_name_input = str(row['region']).strip().lower()
                            try:
                                region = Region.objects.annotate(
                                    region_name_lower=Lower('region_name')
                                ).filter(
                                    is_deleted=False,
                                    region_name_lower=region_name_input
                                ).first()
                            except (ValueError, Region.DoesNotExist):
                                errors.append(f"Row {row_num}: Region ID {region_name_input} not found.")
                                count_data['errors'] += 1
                                # continue

                            state_name = str(row['state']).strip()
                            try:
                                state = State.objects.get(name__iexact=state_name)
                            except State.DoesNotExist:
                                errors.append(f"Row {row_num}: State '{state_name}' not found.")
                                count_data['errors'] += 1
                                # continue

                            district_name = str(row['district']).strip()
                            try:
                                district = District.objects.get(name__iexact=district_name, state=state)
                            except District.DoesNotExist:
                                errors.append(f"Row {row_num}: District '{district_name}' not found for state '{state_name}'.")
                                count_data['errors'] += 1
                                # continue

                            village_or_town = str(row['village_or_town']) if pd.notna(row['village_or_town']) else None
                            address = str(row['address']) if pd.notna(row['address']) else None
                            mobile_number = str(row['mobile_number']) if pd.notna(row['mobile_number']) else None

                            # aggregator_name = str(row['associated_aggregator']).strip()
                            
                            # aggregator = AggregatorProfile.cmobjects.annotate(
                            #     name_lower=Lower('name')
                            # ).filter(
                            #     name_lower=aggregator_name.lower(),
                            #     # associated_entity__in=blf_details
                            # ).first()
                            # if aggregator is None:
                            #     errors.append(f"Row {row_num}: Aggregator '{aggregator_name}' not found or not linked to this BLF.")
                            #     count_data['errors'] += 1
                            #     continue

                            aggregator_str = str(row['associated_aggregator']).strip()
                            aggregator_names = [name.strip() for name in aggregator_str.split('/') if name.strip()]
                            aggregators_to_add = []
                            for aggregator_name in aggregator_names:
                                aggregator = AggregatorProfile.cmobjects.annotate(name_lower=Lower('name')).filter(name_lower=aggregator_name.lower(),associated_entity=blf_details).first()
                                if aggregator:
                                    aggregators_to_add.append(aggregator)
                                else:
                                    errors.append(f"Row {row_num}: Aggregator '{aggregator_name}' not found or not linked to this BLF.")
                                    count_data['errors'] += 1

                            # BLF (Entity)
                            entity_name = str(row['associated_entity']).strip()
                            try:
                                entity = BlfProfile.objects.get(entity_unit__iexact=entity_name)
                            except BlfProfile.DoesNotExist:
                                errors.append(f"Row {row_num}: BLF Entity '{entity_name}' not found.")
                                count_data['errors'] += 1
                                continue

                            # Optional fields
                            total_male_worker = int(row['total_male_worker']) if pd.notna(row['total_male_worker']) and str(row['total_male_worker']).isdigit() else None
                            total_female_worker = int(row['total_female_worker']) if pd.notna(row['total_female_worker']) and str(row['total_female_worker']).isdigit() else None
                            estimated_production_of_green_tea = to_whole_str(row['estimated_production_of_green_tea']) if pd.notna(row['estimated_production_of_green_tea']) else None
                            estimated_production_of_made_tea = to_whole_str(row['estimated_production_of_made_tea']) if pd.notna(row['estimated_production_of_made_tea']) else None
                            # garden_name = str(row['garden_name']) if pd.notna(row['garden_name']) else None
                            production_area = str(row['production_area']) if pd.notna(row['production_area']) else None
                            user_type_obj = ProfileType.objects.filter(name='grower').first()

                            # # Create or get User
                            # user, created = User.objects.get_or_create(
                            #     username=username,
                            #     defaults={
                            #         'first_name': name[:30],
                            #         'is_active': True
                            #     }
                            # )
                            # user.set_password(password)
                            # user.save()
                            # Profile.objects.update_or_create(user=user, defaults={'user_type' : user_type_obj})

                            # GrowerProfile.objects.filter(user=user)

                            # # Update or create GrowerProfile
                            # profile, created = GrowerProfile.objects.update_or_create(
                            #     user=user,
                            #     defaults={
                            #         'profile_type' : user_type_obj,
                            #         'username': username,
                            #         'password': password,
                            #         'name': name,
                            #         'age': age,
                            #         'gender': gender,
                            #         'date_of_birth': dob,
                            #         'grower_type': grower_type,
                            #         'region': region,
                            #         'state': state,
                            #         'district': district,
                            #         'village_or_town': village_or_town,
                            #         'address': address,
                            #         'mobile_number': mobile_number,
                            #         'total_male_worker': total_male_worker,
                            #         'total_female_worker': total_female_worker,
                            #         'estimated_production_of_green_tea': estimated_production_of_green_tea,
                            #         'estimated_production_of_made_tea': estimated_production_of_made_tea,
                            #         # 'garden_name': garden_name,
                            #         'production_area': production_area,
                            #         'is_active': True,
                            #     }
                            # )
                            # garden, garden_created = Gardens.objects.update_or_create(
                            #     grower=profile, is_deleted = False,
                            #     defaults={
                            #         'user' : user,
                            #         'name': f"{profile.name}",
                            #         'is_division': False,
                            #         'production_area' : production_area,
                            #     }
                            # )
                            # plot, plot_created = Plot.objects.update_or_create(
                            #     garden=garden,
                            #     defaults={
                            #         'name': name or garden.name,
                            #     }
                            # )
                            # profile.associated_aggregator.clear()
                            # if aggregator:
                            #     profile.associated_aggregator.add(aggregator)


                            # profile.associated_entity.clear()
                            # if entity:
                            #     profile.associated_entity.add(entity)

                            # if created:
                            #     count_data['created_users'].append(username)
                            #     count_data['items_created'] += 1
                            # else:
                            #     count_data['updated_users'].append(username)
                            #     count_data['items_updated'] += 1

                        except Exception as e:
                            errors.append(f"Row {row_num} (username={username}): {str(e)}")
                            count_data['errors'] += 1
                            continue

                count_data['total_processed'] = count_data['items_created'] + count_data['items_updated']
                return Response({
                    'request_status': 1,
                    'msg': 'Grower data imported successfully.',
                    'data': count_data,
                    'errors': errors,
                }, status=200)
        except Exception as e:
            print(f"Import error: {e}")
            raise APIException({'request_status': 0, 'msg': str(e)})


# Agg Username
def generate_agg_username(full_name: str, mobile_number: str) -> str:
    full_name = full_name.strip().upper()
    mobile_number = str(mobile_number).strip()

    mobile_digits = ''.join(filter(str.isdigit, mobile_number))
    if len(mobile_digits) < 2:
        raise ValueError("Mobile number must contain at least 2 digits.")

    name_parts = [part for part in full_name.split() if part]
    if not name_parts:
        raise ValueError("Full name cannot be empty.")

    first_name = name_parts[0]
    first_part = first_name[:2].ljust(2, 'X')[:2]  

    if len(name_parts) >= 2:
        last_name = name_parts[-1]
        last_part = last_name[:2].ljust(2, 'X')[:2]
    else:
        last_part = "XX" 
    mobile_part = mobile_digits[:2]  # first two digit
    username = first_part + last_part + mobile_part
    return username

def generate_agg_passwords(full_name: str, mobile_number: str) -> str:
    full_name = full_name.strip().upper()
    mobile_number = str(mobile_number).strip()
    mobile_digits = ''.join(filter(str.isdigit, mobile_number))
    if len(mobile_digits) < 2:
        raise ValueError("Mobile number must contain at least 2 digits.")

    name_parts = [part for part in full_name.split() if part]
    if not name_parts:
        raise ValueError("Full name cannot be empty.")

    first_name = name_parts[0]
    first_part = first_name[:2].ljust(2, 'X')[:2]  

    if len(name_parts) >= 2:
        last_name = name_parts[-1]
        last_part = last_name[:2].ljust(2, 'X')[:2]
    else:
        last_part = "XX" 
    mobile_part = mobile_number[-2:]  # last two digit
    password = first_part + last_part + mobile_part
    return password

class GenerateAggUsernameAPIView(APIView):
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    def post(self, request, format=None):
        try:
            file_name = request.data.get('file_name')
            if not file_name:
                return Response({'request_status': 0, 'msg': "Missing 'file_name'"}, status=400)

            file_path = os.path.join('media', 'excel', file_name)
            if not os.path.exists(file_path):
                return Response({
                    'request_status': 0,
                    'msg': f"File '{file_name}' not found on server."
                }, status=400)

            try:
                excel_file = pd.ExcelFile(file_path)
            except Exception as e:
                return Response({
                    'request_status': 0,
                    'msg': f"Failed to read Excel file: {str(e)}"
                }, status=400)

            result = {
                'total_count': 0,
                'sheets_processed': [],
                'usernames': [],
                'passwords': [],
                'errors': []
            }
            for sheet_name in excel_file.sheet_names:
                result['sheets_processed'].append(sheet_name)
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                # Clean DataFrame
                df = df.fillna("")
                df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
                for idx, row in df.iterrows():
                    row_num = idx + 2  
                    try:
                        name = row.get('name')
                        mobile_number = row.get('mobile_number')

                        if pd.isna(name) or not str(name).strip():
                            raise ValueError("Missing or empty 'name'")
                        if pd.isna(mobile_number) or not str(mobile_number).strip():
                            raise ValueError("Missing or empty 'mobile_number'")
                        name = str(name).strip()
                        mobile_number = str(mobile_number).strip()
                        username = generate_agg_username(name, mobile_number)
                        passwords = generate_agg_passwords(name, mobile_number)
                        result['usernames'].append(username)
                        result['passwords'].append(passwords)
                        result['total_count'] += 1
                    except Exception as e:
                        result['errors'].append(f"Row {row_num}: {str(e)}")
                        continue
            return Response({
                'request_status': 1,
                'msg': 'AGG usernames generated successfully.',
                'data': result
            }, status=200)
        except Exception as e:
            return Response({
                'request_status': 0,
                'msg': f"Unexpected error: {str(e)}"
            }, status=500)
        
class ImportAggregatorView(APIView):
    """Optimized aggregator import API view for large Excel files."""
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    def post(self, request, format=None):
        try:
            file_path = _get_import_file_path(request.data.get('file_name'))
        except Exception as e:
            return Response({
                'request_status': 0,
                'msg': str(e),
            }, status=status.HTTP_400_BAD_REQUEST)

        count_data = {
            'items_created': 0,
            'items_updated': 0,
            'errors': 0,
            'sheets_processed': [],
        }
        errors = []
        try:
            user_type_obj = ProfileType.objects.get(name='aggregator')
        except ProfileType.DoesNotExist:
            return Response({
                'request_status': 0,
                'msg': "ProfileType 'aggregator' not found in database."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        regions = _first_by_key(Region.cmobjects.all(), 'region_name')
        states = _first_by_key(State.objects.filter(is_deleted=False), 'name')
        districts = _first_by_key(District.objects.filter(is_deleted=False), 'name')
        entities = _first_by_key(BlfProfile.cmobjects.all(), 'entity_unit')
        batch = []
        seen_sheets = set()

        def flush_batch():
            if not batch:
                return
            with transaction.atomic():
                users = _sync_users(batch)
                for record in batch:
                    record['user'] = users.get(record['username'])

                user_ids = [record['user'].id for record in batch if record.get('user')]
                existing_profiles = AggregatorProfile.objects.in_bulk(user_ids, field_name='user_id')
                profiles_to_create = []
                profiles_to_update = []

                for record in batch:
                    user = record.get('user')
                    if not user:
                        continue
                    profile = existing_profiles.get(user.id)
                    defaults = {
                        'profile_type': user_type_obj,
                        'username': record.get('username'),
                        'password': record.get('password'),
                        'name': record.get('name'),
                        'region': record.get('region'),
                        'state': record.get('state'),
                        'district': record.get('district'),
                        'mobile_number': record.get('mobile_number'),
                        'is_active': True,
                        'is_import_users': True,
                    }
                    if profile:
                        for field_name, value in defaults.items():
                            setattr(profile, field_name, value)
                        profiles_to_update.append(profile)
                    else:
                        profiles_to_create.append(AggregatorProfile(user=user, **defaults))

                if profiles_to_create:
                    AggregatorProfile.objects.bulk_create(profiles_to_create, batch_size=IMPORT_BATCH_SIZE)
                    count_data['items_created'] += len(profiles_to_create)
                if profiles_to_update:
                    AggregatorProfile.objects.bulk_update(
                        profiles_to_update,
                        [
                            'profile_type', 'username', 'password', 'name', 'region',
                            'state', 'district', 'mobile_number', 'is_active', 'is_import_users'
                        ],
                        batch_size=IMPORT_BATCH_SIZE,
                    )
                    count_data['items_updated'] += len(profiles_to_update)

                profile_by_user_id = AggregatorProfile.objects.in_bulk(user_ids, field_name='user_id')
                _sync_common_profiles(batch, user_type_obj, is_grower=False)

                source_ids = [profile.id for profile in profile_by_user_id.values()]
                entity_pairs = []
                for record in batch:
                    profile = profile_by_user_id.get(record['user'].id)
                    if profile:
                        entity_pairs.append((profile.id, record['entity'].id))
                _bulk_replace_m2m(AggregatorProfile._meta.get_field('associated_entity'), source_ids, entity_pairs)
            batch.clear()

        required_columns = {'username', 'name', 'region', 'state', 'district', 'mobile_number', 'associated_entity'}
        for sheet_name, row_num, row, sheet_error in _iter_excel_records(file_path, required_columns):
            if sheet_name not in seen_sheets:
                seen_sheets.add(sheet_name)
                count_data['sheets_processed'].append(sheet_name)
            if sheet_error:
                _append_import_error(errors, count_data, f"Sheet '{sheet_name}': {sheet_error}")
                continue

            username = _clean_str(row.get('username'))
            if not username:
                _append_import_error(errors, count_data, f"Sheet '{sheet_name}' Row {row_num}: Username is required.")
                continue

            region = regions.get(_normalise_key(row.get('region')))
            state = states.get(_normalise_key(row.get('state')))
            district = districts.get(_normalise_key(row.get('district')))
            entity = entities.get(_normalise_key(row.get('associated_entity')))

            if not region:
                _append_import_error(errors, count_data, f"Sheet '{sheet_name}' Row {row_num}: Region '{_clean_str(row.get('region'))}' not found.")
                continue
            if not state:
                _append_import_error(errors, count_data, f"Sheet '{sheet_name}' Row {row_num}: State '{_clean_str(row.get('state'))}' not found.")
                continue
            if not district:
                _append_import_error(errors, count_data, f"Sheet '{sheet_name}' Row {row_num}: District '{_clean_str(row.get('district'))}' not found.")
                continue
            if not entity:
                _append_import_error(errors, count_data, f"Sheet '{sheet_name}' Row {row_num}: BLF Entity '{_clean_str(row.get('associated_entity'))}' not found.")
                continue

            if any(item['username'] == username for item in batch):
                flush_batch()

            batch.append({
                'username': username,
                'password': _clean_str(row.get('password')) or '123',
                'name': _clean_str(row.get('name')) or '',
                'region': region,
                'state': state,
                'district': district,
                'mobile_number': _whole_str(row.get('mobile_number')),
                'entity': entity,
            })

            if len(batch) >= IMPORT_BATCH_SIZE:
                flush_batch()

        flush_batch()

        count_data['total_processed'] = count_data['items_created'] + count_data['items_updated']
        count_data['errors_returned'] = len(errors)

        return Response({
            'request_status': 1,
            'msg': 'Data imported successfully.',
            'data': count_data,
            'errors': errors,
        }, status=status.HTTP_200_OK)


class PlotListGrowerAPIView(APIView):
    """ Plot API for grower app View """
    queryset = Plot.objects.filter(plot_status=True)
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    def get(self, request):
        id = request.query_params.get('id')
        all_results = request.query_params.get('all')
        order_by = request.query_params.get('order_by', '-id')
        search = {}
        search['plot_status'] = True
        search['garden__grower__user'] = request.user
        search = custom_filters(request, {}, [])
        if id:
            list_data = Plot.objects.filter(id=id).first()
            serializer = PlotListSerializer(list_data, many=False)
            return Response(serializer.data)
        list_data = Plot.objects.filter(*search).order_by(*str(order_by).split(","))
        if all_results == 'true':
            serializer = PlotListSerializer(list_data, many=True)
            return Response({'results': serializer.data})

        page_size = int(request.query_params.get('page_size', 10))
        paginator = Paginator(list_data, page_size)
        page_number = request.query_params.get('page', 1)
        page = paginator.get_page(page_number)

        serializer = PlotListSerializer(page.object_list, many=True)
        return Response({
            'count': paginator.count,
            'next': page.next_page_number() if page.has_next() else None,
            'previous': page.previous_page_number() if page.has_previous() else None,
            'results': serializer.data,
        })
    

class GenerateSupplyReportGrowerPdfAPIView(APIView):
    """ Supply Report List pdf Generate """
    # permission_classes = [AllowAny]
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    def get(self, request, *args, **kwargs):
        # try:
            date_from= self.request.query_params.get('date_from', None)
            date_to= self.request.query_params.get('date_to', None) 
            user_type=self.request.query_params.get('consumer_type', "") 
            grower_id = self.request.query_params.get('grower_id', "") 
            template_path = 'supply_report_grower.html' 

            if user_type.lower() == 'blf' :    
                if (date_from and date_to ) :
                    supply_data = SupplyManagement.objects.filter(is_deleted=False,created_by=request.user, date_of_supply__range=[date_from,date_to],\
                                                       supply_to__iexact='Factory')
                else:
                    supply_data = SupplyManagement.objects.filter(is_deleted=False,created_by=request.user,supply_to__iexact='Factory')
                supply_details = []
                grower_details=GrowerProfile.objects.filter(user=request.user).first()
                for data in supply_data:
                    grower_details_supply = GrowerDetailsSupply.objects.filter(supply_id=data.id).first()
                    consumer = BlfProfile.objects.filter(user=data.consumer).first()
                    alloted_vehicle = VehicleManagement.objects.filter(id=data.alloted_vehicle_id).first()
                    total_collected_quantity = GrowerDetailsSupply.objects.filter(supply_id=data.id).aggregate(total=Sum('collected_quantity'))['total'] or 0
                    total_supply_quantity = GrowerDetailsSupply.objects.filter(supply_id=data.id).aggregate(total=Sum('supply_quantity'))['total'] or 0
                    consumer_name = consumer.entity_unit if consumer else None
                    consumer_user_name = consumer.user.username if consumer else None
                    
                    # Add spaces after every 8 characters
                    if len(consumer_name) > 11 and ' ' not in consumer_name:
                        consumer_name = ' '.join([consumer_name[i:i+11] for i in range(0, len(consumer_name), 11)])

                    if len(consumer_user_name) > 11 and ' ' not in consumer_user_name:
                        consumer_user_name = ' '.join([consumer_user_name[i:i+11] for i in range(0, len(consumer_user_name), 11)])
                    challan_id = data.supply_challan_id if data.supply_challan_id else None
                    if challan_id and len(challan_id) > 11 and ' ' not in challan_id:
                        challan_id = ' '.join([challan_id[i:i+11] for i in range(0, len(challan_id), 11)])
                    supply_details.append({
                        'consumer_name': consumer_name,
                        'consumer_user_name': consumer_user_name,
                        'vehicle_number': alloted_vehicle.vehicle_number if alloted_vehicle else None,
                        'total_collected_quantity': total_collected_quantity,
                        'total_supply_quantity': total_supply_quantity,
                        'date_of_supply': data.date_of_supply,
                        'challan_id':challan_id,
                        'gross_leaf':data.gross_leaf,
                        'quantity': data.quantity,
                        'supply_bag_id':data.supply_bag_id,
                        'driver_name':data.driver_name
                    })
            elif user_type.lower() == 'aggregator':
                if (date_from and date_to ) :
                    supply_data = SupplyManagement.objects.filter(is_deleted=False,date_of_supply__range=[date_from,date_to],\
                                                       supply_to__iexact='Aggregator')
                else:
                    supply_data = SupplyManagement.objects.filter(is_deleted=False, supply_to__iexact='Aggregator')
                supply_details = []
                grower_details=GrowerProfile.objects.filter(id=grower_id).first()
                for data in supply_data:
                    grower_details_supply = GrowerDetailsSupply.objects.filter(supply_id=data.id).first()
                    consumer = AggregatorProfile.objects.filter(user=data.consumer).first()
                    alloted_vehicle = VehicleManagement.objects.filter(id=data.alloted_vehicle_id).first()
                    total_collected_quantity = GrowerDetailsSupply.objects.filter(supply_id=data.id).aggregate(total=Sum('collected_quantity'))['total'] or 0
                    total_supply_quantity = GrowerDetailsSupply.objects.filter(supply_id=data.id).aggregate(total=Sum('supply_quantity'))['total'] or 0
                    # total_collected_quantity_sum += total_collected_quantity
                    # total_supply_quantity_sum += float(total_supply_quantity)
                    # Add spaces after every 8 characters
                    consumer_name = consumer.entity_unit if consumer else None
                    consumer_user_name = consumer.user.username if consumer else None
                    if len(consumer_name) > 11 and ' ' not in consumer_name:
                        consumer_name = ' '.join([consumer_name[i:i+11] for i in range(0, len(consumer_name), 11)])

                    if len(consumer_user_name) > 11 and ' ' not in consumer_user_name:
                        consumer_user_name = ' '.join([consumer_user_name[i:i+11] for i in range(0, len(consumer_user_name), 11)])
                    challan_id = data.supply_challan_id if data.supply_challan_id else None
                    if challan_id and len(challan_id) > 11 and ' ' not in challan_id:
                        challan_id = ' '.join([challan_id[i:i+11] for i in range(0, len(challan_id), 11)])
                    supply_details.append({
                    'consumer_name': consumer_name,
                    'consumer_user_name':consumer_user_name,
                    'vehicle_number': alloted_vehicle.vehicle_number if alloted_vehicle else None,
                    'total_collected_quantity': total_collected_quantity,
                    'total_supply_quantity': total_supply_quantity,
                    'date_of_supply': data.date_of_supply,
                    'challan_id':challan_id,
                    'gross_leaf':data.gross_leaf,
                    'quantity': data.quantity,
                    'supply_bag_id':data.supply_bag_id,
                    'driver_name':data.driver_name
                    })        
            else:
                return Response({'msg': 'consumer type required either factory or aggregator.'},\
                                          status=status.HTTP_400_BAD_REQUEST)
            total_supply_quantity_sum = sum(
            float(item['quantity']) if item['quantity'] else 0 for item in supply_details)
            if date_from and date_to :    
                datetime_object_from = datetime.strptime(date_from, "%Y-%m-%d")
                date_from = datetime_object_from.strftime("%d-%m-%Y")    
                datetime_object_to = datetime.strptime(date_to, "%Y-%m-%d")
                date_to= datetime_object_to.strftime("%d-%m-%Y")   
            else:
                date_from =""
                date_to=""
            context = {
                'grower_details':grower_details.name if grower_details else "",
                'grower_address':grower_details.address if grower_details else "",
                'grower_user_name':grower_details.user if grower_details else "",
                "date_from":date_from  if date_from else "",
                'date_to':date_to if date_to else "",
                'supply_details' : supply_details,
                'consumer_type' : user_type.lower(),
                # 'total_collected_quantity_sum': total_collected_quantity_sum,
                'total_supply_quantity_sum': total_supply_quantity_sum,
            
	        }# Provide context data if needed
            template = get_template(template_path)
            html = template.render(context)
            response = HttpResponse(content_type='application/pdf')
            # response['Content-Disposition'] = 'attachment; filename="delivery_challan.pdf"'
            response['Content-Disposition'] = 'filename="supply_report.pdf"'
            pisa_status = pisa.CreatePDF(html, dest=response)
            if pisa_status.err:
                return HttpResponse('Error generating PDF', status=500)
            return response
        
        # except Exception as e:
        #     raise serializers.ValidationError({'status': status.HTTP_400_BAD_REQUEST,'request_status': 0, 'msg': e},code='authorization')   

from django.core.files.storage import FileSystemStorage
class SupplyReportPdfGenerateAPIView(APIView):
    """ Supply Report List PDF Generate """
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)
    
    def get(self, request, *args, **kwargs):
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        grower_id = request.query_params.get('grower_id')
        user_type = request.query_params.get('user_type', 'grower')
        if not date_from or not date_to:
            raise APIException({'request_status': 0, 'msg': "date_from and date_to are mandatory"})

        grower_details=GrowerProfile.objects.filter(Q(user=request.user)|Q(id=grower_id)).first()
        user_type = Profile.cmobjects.filter(user=request.user).values('user_type__name').first()
        print("user_type ####", user_type)

        supply_data = SupplyManagement.cmobjects.select_related(
            'alloted_vehicle', 'consumer'
        ).filter(
            created_by=request.user,
            date_of_supply__range=[date_from, date_to],
        )
        total_gross_leaf = supply_data.aggregate(
            total=Sum(Cast('gross_leaf', FloatField()))
        )['total']
        if grower_id:
            supply_data = supply_data.filter(created_by_id=grower_id)
        supply_data = supply_data.annotate(
            consumer_name=F('consumer__profile__full_name')
        ).values(
            'id',
            'supply_to',
            'date_of_supply',
            'alloted_vehicle__vehicle_number',
            'supply_challan_id',
            'consumer',
            'gross_leaf',
            'consumer__profile__user__username',
            'consumer__profile__phone_no',
            'consumer__profile__full_name',
        )
        template = get_template('supply_report_grower.html')
        html = template.render({
            'supply_details': supply_data,
            'date_from': date_from,
            'date_to': date_to,
            'total_gross_leaf' : total_gross_leaf,
            "grower_details" : grower_details,
        })
        options = {
            'page-size': 'A4',
            'encoding': "UTF-8",
            'dpi': 300,             
            'zoom': 1.3,             
            'no-outline': None,
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
            'enable-smart-shrinking': True,  # Better layout scaling
        }
        config = pdfkit.configuration(wkhtmltopdf=str(os.getenv('WKHTMLTOPDF_PATH', '')))
        pdf = pdfkit.from_string(html, False, options=options, configuration=config)

        # response = HttpResponse(pdf, content_type='application/pdf')
        # response['Content-Disposition'] = f'attachment; filename="supply_report_{date_from}_to_{date_to}.pdf"'
        # return response

        # Generate unique filename
        filename = f"supply_report_{date_from}_to_{date_to}_{uuid.uuid4().hex}.pdf"
        fs = FileSystemStorage(location=settings.MEDIA_ROOT)
        file_path = fs.save(filename, ContentFile(pdf))
        file_url = fs.url(file_path)
        return Response({
            'request_status': 1,
            'msg': 'PDF generated successfully',
            'file_url': request.build_absolute_uri(file_url)
        })



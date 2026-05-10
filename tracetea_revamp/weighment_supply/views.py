from django.shortcuts import render, redirect
from django.contrib import messages
from django.urls import reverse
from django.db.models import Q
from django.http import JsonResponse
from django.db.models.functions import Coalesce
from django.db.models import Sum, Count, Value
from django.template.loader import render_to_string
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect
from django.contrib.auth import get_user_model
User = get_user_model()
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, DetailView, UpdateView, CreateView
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.paginator import Paginator , EmptyPage, PageNotAnInteger
import pdfkit
from django.http import HttpResponse
from django.template.loader import render_to_string
from datetime import datetime
from itertools import chain
from django.db.models import Min, Max

import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter

from django.db.models.functions import Coalesce
from django.db.models import FloatField
from django.db.models import Count, F, Sum, Avg

from django.db.models import Max

from django.db import transaction
from master.common import CommonMixin
from .models import *
from .forms import *

from master.decorators import *
from user_profile.aggregator_api_models import SupplyManagement
from user_profile.blf_api_models import SupplierExit

from leaf_receipt.models import *

from datetime import datetime


class WeighmentSupplyListView(LoginRequiredMixin, CommonMixin, ListView):
    """
    Weighment Supply List View
    """
    model = WeighmentSupply
    context_object_name = 'weighment_list'
    template_name = 'weighment/weighmentsupply_list.html'
    paginate_by = 10

    def get(self, request, *args, **kwargs):
    # checking if the user is customer


        return super().get(request, *args, **kwargs)

    def get_queryset(self, *args, **kwargs):

        qs = super().get_queryset(*args, **kwargs)
        weighment_txn_id=self.request.GET.get('weighment_txn_id', '')
        supply_date = self.request.GET.get('supply_date', '')
        supply_challan = self.request.GET.get('supply_challan', '')

        weighment_txn_id_without_spaces = weighment_txn_id.replace(" ", "")
        supply_date_without_spaces = supply_date.replace(" ", "")
        supply_challan_without_spaces = supply_challan.replace(" ", "")

        if weighment_txn_id:
            return qs.filter(weighment_txn_id__icontains=weighment_txn_id_without_spaces, created_by_id=self.request.user.id)
        if supply_date:
            return qs.filter(supply_date__icontains=supply_date_without_spaces, created_by_id=self.request.user.id)
        if supply_challan:
            return qs.filter(supply_challan__supply_challan_id__icontains=supply_challan_without_spaces, created_by_id=self.request.user.id)
        
        return WeighmentSupply.cmobjects.filter(created_by_id=self.request.user.id).order_by('-id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # context['tea_type_list'] = TeaType.objects.all().order_by('id')
        return context



class WeighmentSupplyCreateProfile(LoginRequiredMixin, CreateView, CommonMixin):
    """
    Weightment and supply Create View
    """
    form_class = WeighmentSupplyForm
    template_name = 'weighment/add_profile.html'

    def get(self, request, *args, **kwargs):
    # checking if the user is customer
        # try:
        #     if not self.request.user.is_superuser:
        #         messages.error(self.request, 'You have no permission to access the requested resource!')
        #         return redirect(reverse('index'))
        # except AttributeError as error:
        #     messages.error(self.request, 'You have no permission to access the requested resource!')
        #     return redirect(reverse('index'))
        
        return super().get(request, *args, **kwargs)

    def get_success_url(self, **kwargs):
        messages.success(self.request, 'Weighment Supply Created Successfully')
        return reverse('weighment_supply:weighment_create')

    # def form_vails(self, ):


    def form_invalid(self, form):
        error_message = 'Error saving the Form, check fields below.'
        messages.error(self.request, error_message)
        return super().form_invalid(form)



def grower_supplier(request):

    type = request.GET.get('type', None)
    suppliers_list = GrowerProfile.objects.filter(associated_entity__user=request.user).order_by('-id')
    context ={
        'suppliers_list' : suppliers_list,
    }
    return render(request, 'weighment/suppliers_dropdown_list_options.html', context)

def load_suppliers(request):
    type = request.GET.get('type', None)
    suppliers_list = AggregatorProfile.objects.filter(associated_entity__user=request.user).order_by('-id')
    context ={
        'suppliers_list' : suppliers_list,
    }
    return render(request, 'weighment/suppliers_dropdown_list_options.html', context)



@login_required
def weighment_supply_add(request, supplier_type, supplier_id, challan_id):
    """ Weighment supply Create View """
    # try:
    #     if SupplyManagement.objects.get(id=challan_id, is_weighment_proceed=True).exists():
    #         messages.error(request, 'Weightment already proceeded for the given challan ID')
    #         return HttpResponseRedirect(request.path_info)
    # except AttributeError as error:
    #     messages.error(request, 'Weightment already proceeded for the given challan ID')
    #     return HttpResponseRedirect(request.path_info)
    

    # if SupplyManagement.objects.get(id=challan_id, is_weighment_proceed=True).exists():
    #     messages.error(request, 'Weightment already proceeded for the given challan ID')
        
    if str(supplier_type) == "aggregator":
        profile_model = AggregatorProfile
    elif str(supplier_type) == "grower":
        profile_model = GrowerProfile

    profile_details = profile_model.cmobjects.filter(user_id=supplier_id).first()
    supplier_details = SupplyManagement.cmobjects.filter(created_by_id=supplier_id, supply_challan_id=challan_id).first()

    if str(supplier_details.vehicle_option) == "Yes":
        vehicle_details = supplier_details.alloted_vehicle.vehicle_number
        vehicle_mobile_no = supplier_details.alloted_vehicle.mobile_number
    else:
        vehicle_details = ""
        vehicle_mobile_no = ""

    date_of_supply = supplier_details.date_of_supply
    
    date_time_str = date_of_supply.strftime("%d-%m-%Y")

    print("SUPPLER DETAILS====", supplier_details.id)

    print(supplier_details.vehicle_option)

    if request.method == 'POST':
      
        form = WeighmentSupplyForm(request.POST)
        if form.is_valid() and supplier_details.is_weighment_proceed == False:
            supplier_type = request.POST.get('supplier_type', None)
            supplier = request.POST.get('supplier', None)
            supply_challan = supplier_details.id
            mode_of_supply = request.POST.get('mode_of_supply', None)
            supply_date = request.POST.get('supply_date', None)
            vehicle_no = supplier_details.alloted_vehicle_id
            total_gross_weight_kg = request.POST.get('total_gross_weight_kg', None)
            mobile_number = request.POST.get('mobile_number', None)

            create = WeighmentSupply.objects.create(created_by=request.user, supplier_id=supplier,
                supplier_type=supplier_type, supply_date=supply_date, vehicle_no_id=vehicle_no, \
                    total_gross_weight_kg=total_gross_weight_kg, mobile_number=mobile_number, mode_of_supply=mode_of_supply, supply_challan_id=supply_challan)
            


            SupplyManagement.objects.filter(pk=supplier_details.pk).update(is_weighment_proceed=True)



            last_challan = WeighmentSupply.objects.filter(weighment_txn_id__contains="TXN").aggregate(Max('weighment_txn_id'))
            last_challan_id = last_challan.get('weighment_txn_id__max')
            print(last_challan_id)
            # Calculate the next supply_challan_id
            if last_challan_id and last_challan_id.startswith("TXN"):
                # Extract the numeric part and increment it
                print("check")
                numeric_part = int(last_challan_id[-1]) + 1
                next_challan_id = f"TXN{numeric_part:09d}"
                
                # Check if the generated ID already exists
                while WeighmentSupply.objects.filter(weighment_txn_id=next_challan_id).exists():
                    numeric_part += 2
                    next_challan_id = f"TXN{numeric_part:09d}"
            else:
                # Handle the case when there are no existing supply_challan_id values
                next_challan_id = "TXN000000001"
            create.weighment_txn_id=next_challan_id
            create.save()

            
            messages.success(request, 'Created Successfully')
            # return HttpResponseRedirect(request.META.get("HTTP_REFERER"))
            return redirect(reverse('weighment_supply:weighment_list'))
        else:
            messages.error(request, 'Challan ID allready Proceed')
    else:
        form = WeighmentSupplyForm()

    context ={
        'date_of_supply' : date_time_str, 
        'supplier_details' : supplier_details,
        'profile_details' : profile_details,
        'supplier_type' : supplier_type,
        'supplier_id' : supplier_id,
        'challan_id' : challan_id,
        'form' : form,
        'vehicle_details' : vehicle_details,
        'vehicle_mobile_no' : vehicle_mobile_no,
        'allow_vehicle' : str(supplier_details.vehicle_option),
    }

    return CommonMixin.render(request, 'weighment/create_form.html', context)



# LOAD DATE IN CREATE FORM
def weighment_load_supply_date(request):

    id = request.GET.get('id', None)

    supply_details= SupplyManagement.cmobjects.filter(pk=id).first()

    print(supply_details.date_of_supply)

    supply_date = supply_details.date_of_supply
    gross_leaf_weight = supply_details.gross_leaf

    data = {
        "value" : supply_date,
    }

    return JsonResponse(data)



def weighment_load_supply_gross_leaf(request):

    id = request.GET.get('id', None)

    supply_details= SupplyManagement.cmobjects.filter(pk=id).first()

    # print("supply gross", supply_details.gross_leaf)

    gross_leaf_weight = supply_details.gross_leaf


    data = {
        "value" : gross_leaf_weight,
    }

    return JsonResponse(data)






# EDIT 

def weighment_supply_edit(request, weighment_pk, supplier_type, supplier_id, challan_id):
    
    if str(supplier_type) == "aggregator":
        profile_model = AggregatorProfile
    elif str(supplier_type) == "grower":
        profile_model = GrowerProfile

    print(supplier_id)
    print(challan_id)

    profile_details = profile_model.cmobjects.filter(user_id=supplier_id).first()
    supplier_details = SupplyManagement.cmobjects.filter(created_by_id=supplier_id, supply_challan_id=challan_id).first()
    weighment_details = WeighmentSupply.cmobjects.filter(pk=weighment_pk).first()

    print(supplier_details)

    vehicle_details = supplier_details.alloted_vehicle.vehicle_number

    date_of_supply = supplier_details.date_of_supply
    
    date_time_str = date_of_supply.strftime("%Y-%m-%d")

    weight_supply_details = WeighmentSupply.cmobjects.filter(id=weighment_pk).first()

    if request.method == 'POST':
      
        form = WeighmentSupplyForm(request.POST, instance=weight_supply_details)
        if form.is_valid():
            supplier_type = request.POST.get('supplier_type', None)
            supplier = request.POST.get('supplier', None)
            supply_challan = supplier_details.id
            mode_of_supply = request.POST.get('mode_of_supply', None)
            supply_date = request.POST.get('supply_date', None)
            vehicle_no = supplier_details.alloted_vehicle_id
            total_gross_weight_kg = request.POST.get('total_gross_weight_kg', None)
            mobile_number = supplier_details.mobile_number

            print("POST INPUT VALUE")
            print("Supplier Type - ", supplier_type)
            print("Suplier ID - ", supplier)
            print("Challan - ", supply_challan)
            print("Mode Of Supply - ", mode_of_supply)
            print("Supply Date - ", supply_date)
            print("Vechile No - ", vehicle_no)
            print("KG - ", total_gross_weight_kg)
            print("Phone No - ", mobile_number)

            WeighmentSupply.objects.filter(id=weighment_pk).update(created_by=request.user, supplier_id=supplier,
                supplier_type=supplier_type, supply_date=supply_date, vehicle_no_id=vehicle_no, \
                    total_gross_weight_kg=total_gross_weight_kg, mobile_number=mobile_number, mode_of_supply=mode_of_supply, supply_challan_id=supply_challan)
            
            SupplyManagement.objects.filter(pk=supplier_details.pk).update(is_weighment_proceed=True)
            
            messages.success(request, 'Updated Successfully')
            
            return redirect(reverse('weighment_supply:weighment_list'))
        else:
            messages.error(request, 'Please Correct The Error Below.')
    else:
        form = WeighmentSupplyForm(instance=weight_supply_details)

    context ={
        'date_of_supply' : date_time_str, 
        'supplier_details' : supplier_details,
        'profile_details' : profile_details,
        'supplier_type' : supplier_type,
        'supplier_id' : supplier_id,
        'challan_id' : challan_id,
        'form' : form,
        'vehicle_details' : vehicle_details,
        'allow_vehicle' : str(supplier_details.vehicle_option),
    }

    return CommonMixin.render(request, 'weighment/create_form.html', context)





class WeighmentSupplyUpdateView(LoginRequiredMixin, CommonMixin, UpdateView):
    """
    Weighment Supply Update View
    """
    model = WeighmentSupply
    form_class = WeighmentSupplyForm
    template_name = 'weighment/edit_form.html'

    def get(self, request, *args, **kwargs):
        # checking if the user is customer

        return super().get(request, *args, **kwargs)

    def get_success_url(self, **kwargs):
        messages.success(self.request, 'Weighment Supply Updated Successfully')
        return reverse('weighment_supply:weighment_list')

    def get_object(self):
        weighment_details = WeighmentSupply.objects.filter(pk=self.kwargs['weighment_pk']).first()
        return weighment_details

    def form_valid(self, form):
        self.pk = self.kwargs['weighment_pk']
        context = self.get_context_data()
        with transaction.atomic():
            self.object = form.save()
            if not form.is_valid():
                return self.render_to_response(context)
        return super(WeighmentSupplyUpdateView, self).form_valid(form)
    

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['weighment_details'] = self.get_object()
        weighment_details = WeighmentSupply.objects.filter(pk=self.kwargs['weighment_pk']).first()
        supplier_details=SupplyManagement.cmobjects.filter(pk=weighment_details.supply_challan.pk).first()
        context['supplier_details'] = SupplyManagement.cmobjects.filter(pk=weighment_details.supply_challan.pk).first()
        if str(weighment_details.supplier_type) == "aggregator":
            profile_model = AggregatorProfile
        elif str(weighment_details.supplier_type) == "grower":
            profile_model = GrowerProfile
        profile_details = profile_model.cmobjects.filter(user_id=weighment_details.supplier_id).first()

        vehicle_details = supplier_details.alloted_vehicle.id


        print("Vechile No", supplier_details.alloted_vehicle.id)

        context['profile_details'] = profile_details

        return context

    def form_invalid(self, form):
        error_message = 'Error saving the Form, check fields below.'
        messages.error(self.request, error_message)
        return super().form_invalid(form)


# @permission_required_admin
# @login_required
# def grade_details(request, weighment_pk):
#     grade_details = WeighmentSupply.objects.filter(pk=weighment_pk).first()
#     context = {
#         'grade_details' : grade_details,
#     }
#     return CommonMixin.render(request, 'grade_details.html', context)


@login_required
def weighmentsupply_delete(request, weighment_pk):
    """
    Delete Weighment Supply View
    """
    attribute_details = WeighmentSupply.objects.filter(pk=weighment_pk).first()
    attribute_details.is_deleted = True
    attribute_details.save()
    messages.success(
        request, 'Weighment Supply Deleted Successfully')
    return redirect(reverse('weighment_supply:weighment_list'))

############################################
###########    Challan Views ##############

def load_challan(request):
    # profile_details = 
    supplier_id = request.GET.get('type', None)
    supplier_type = Profile.objects.filter(user_id=supplier_id).first().user_type
    non_proceed_challan_list = SupplyManagement.objects.filter(consumer_id=request.user.id, created_by_id=supplier_id, is_weighment_proceed=False).order_by('-id')
    proceed_challan = SupplyManagement.objects.filter(consumer_id=request.user.id, created_by_id=supplier_id, is_weighment_proceed=True).order_by('-id')
    
    print("Supplier ID", supplier_id)

    context ={
        'supplier_type' : supplier_type,
        'supplier_id' : supplier_id,
        'challan_list' : non_proceed_challan_list,
        'proceed_challan_list' : proceed_challan,
    }
    return render(request, 'weighment/challan_list.html', context)




@login_required
def challan_details(request, challn_pk, item_pk):
    """ Challan Details View """

    challan_details = SupplyManagement.cmobjects.filter(pk=challn_pk).first()
    blf_details = BlfProfile.cmobjects.filter(user__id=challan_details.consumer_id).first()
    agg_supplier_details = AggregatorProfile.cmobjects.filter(user_id=challan_details.created_by_id).first()
    grower_supplier_details = GrowerProfile.cmobjects.filter(user_id=challan_details.created_by_id).first()
    profile_type = str(Profile.objects.filter(user_id=challan_details.created_by_id).first().user_type)

    print("challan PK ####", challn_pk)
    print("Profile Type", profile_type)

    print("item_pk ", item_pk)


    weighment_details = WeighmentSupply.cmobjects.filter(supply_challan=challan_details).first()


    if profile_type == "grower":
        ProfileModel = AggregatorProfile
    else:
        ProfileModel = AggregatorProfile
        
    supplier_details = ProfileModel.cmobjects.filter(user_id=challan_details.created_by_id).first()
    print("SUPPLIER DETAILS",supplier_details)
    

    alloted_vehicle=VehicleManagement.objects.filter(id= challan_details.alloted_vehicle_id).first()
    grower_details_supply_list=GrowerDetailsSupply.objects.filter(supply_id=challan_details.id).order_by('-id')

    total_qty = grower_details_supply_list.aggregate(total_supply_quantity=Sum('supply_quantity'))['total_supply_quantity'] or 0
    total_collected_qty = grower_details_supply_list.aggregate(total_collected_qty=Sum('collected_quantity'))['total_collected_qty'] or 0


    context={   
        'challan_details' : challan_details,
        'blf_details' : blf_details,
        'agg_supplier_details' : agg_supplier_details,
        'grower_supplier_details' : grower_supplier_details,
        'profile_type' : profile_type,
        "alloted_vehicle" : alloted_vehicle,
        "grower_details_supply_list" : grower_details_supply_list,
        "total_supply_quantity" : total_qty,
        "total_collected_qty" : total_collected_qty,
        "weighment_details" : weighment_details,
    }

    return CommonMixin.render(request, 'weighment/challan_invoice.html', context)



################ START SUPPLIER EXIT VIEWS #######################

class SupplierExitListView(LoginRequiredMixin, CommonMixin, ListView):
    """
    Weighment Supply List View
    """
    model = SupplierExit
    context_object_name = 'supplier_exit_list'
    template_name = 'supplier_exit/supplier_exit_list.html'
    paginate_by = 10

    def get(self, request, *args, **kwargs):
    # checking if the user is customer
        # try:
        #     if not self.request.user.is_superuser:
        #         messages.error(self.request, 'You have no permission to access the requested resource!')
        #         return redirect(reverse('index'))
        # except AttributeError as error:
        #     messages.error(self.request, 'You have no permission to access the requested resource!')
        #     return redirect(reverse('index'))

        return super().get(request, *args, **kwargs)

    def get_queryset(self, *args, **kwargs):

        qs = super().get_queryset(*args, **kwargs)
        start_date= self.request.GET.get('start_date')
        end_date= self.request.GET.get('end_date')

        if start_date and end_date:
            qs = SupplierExit.cmobjects.filter(weighment_txn__supply_date__range=(
                start_date,
                end_date
            ), created_by_id=self.request.user.id)	
            return qs
        return SupplierExit.cmobjects.filter(created_by_id=self.request.user.id).order_by('-id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context



class SupplierExitCreateView(LoginRequiredMixin, CreateView, CommonMixin):
    """
    Supplier Exit Create View
    """
    form_class = SupplierExitForm
    template_name = 'supplier_exit/create_form.html'
    
    def get(self, request, *args, **kwargs):
    # checking if the user is customer
        # try:
        #     if not self.request.user.is_superuser:
        #         messages.error(self.request, 'You have no permission to access the requested resource!')
        #         return redirect(reverse('index'))
        # except AttributeError as error:
        #     messages.error(self.request, 'You have no permission to access the requested resource!')
        #     return redirect(reverse('index'))
        
        return super().get(request, *args, **kwargs)

    def get_success_url(self, **kwargs):
        messages.success(self.request, 'Weighment Supply Created Successfully')
        return reverse('weighment_supply:supplier_exit_list')


    def get_form_kwargs(self, **kwargs):
        """ Provides keyword arguemnt """
        kwargs = super(SupplierExitCreateView, self).get_form_kwargs()
        kwargs['user_id'] = self.request.user.id
        return kwargs

    def form_valid(self, form):
        context = self.get_context_data()
        response = super().form_valid(form)
        object_id = self.object.pk

        weighment_details = WeighmentSupply.cmobjects.filter(created_by_id=self.request.user.id,\
                                     pk=form.instance.weighment_txn.id).first()
        
        if weighment_details.vehicle_no:
            vehicle_number = weighment_details.vehicle_no
        else:
            vehicle_number = ""

        if str(weighment_details.supplier_type) == "aggregator":
            profile_model = AggregatorProfile
        elif str(weighment_details.supplier_type) == "grower":
            profile_model = GrowerProfile
        else:
            profile_model = ""

        # PROFILE DETAILS
        supplier_details = profile_model.cmobjects.filter(user_id=weighment_details.supplier).first()

        # SUPPLY DETAILS
        supply_chalan_id = weighment_details.supply_challan.id
        supply_details = SupplyManagement.cmobjects.filter(id=supply_chalan_id).first()


        supplier_details = profile_model.cmobjects.filter(user_id=weighment_details.supplier).first()
        if str(weighment_details.supplier_type) == "grower":
            if PluckingData.cmobjects.filter(grower_id=supplier_details.id).exists():
                plucking_details = PluckingData.cmobjects.filter(grower_id=supplier_details.id).latest('id')
                plucking_details_id = plucking_details.id
            else:
                plucking_details = None
                plucking_details_id = None
        else:
            plucking_details = None


        with transaction.atomic():

            form.instance.created_by_id = self.request.user.id

            # if str(form.instance.is_released) == "Yes":
            if form.instance.unloaded_vehicle_weight and str(form.instance.is_released) == "Yes" and form.instance.weighment_txn:

                leaf_weight_details = LeafManagement.cmobjects.filter(weighment_txn_id=form.instance.weighment_txn.id).first()

                net_leaf_weight = (weighment_details.total_gross_weight_kg - form.instance.unloaded_vehicle_weight)
                actual_net_leaf_weight = round(net_leaf_weight - (net_leaf_weight * (leaf_weight_details.deduction/100)), 2)
                form.instance.net_supplied_qty = actual_net_leaf_weight

                weighment_details.is_supplier_exit_proceed = True
                weighment_details.save()

                vehicle_details = weighment_details.vehicle_no
                vehicle_details.is_available = True
                vehicle_details.save()

                CollectionFromGrower.objects.filter(vehicle_number=weighment_details.vehicle_no).update(is_complete=True)
                
                if str(weighment_details.supplier_type) == "aggregator":
                    leaf_collection_create=LeafCollection.objects.create(
                        leaf_receipt_id_id = leaf_weight_details.id,
                        suppler_exit_id_id = object_id,
                        weighment_supply_id_id=weighment_details.id,
                        supply_id_id = supply_details.id,
                        supplier_id = weighment_details.supplier_id,    
                        nt_wght=actual_net_leaf_weight,
                        collection_date= form.instance.date_of_supply,
                        supplier_type = weighment_details.supplier_type,
                        aggregator_id = supplier_details.id,
                        supply_date = supply_details.date_of_supply,
                        created_by_id = self.request.user.id
                    )	
                    leaf_collection_create.save()
                    
                elif str(weighment_details.supplier_type) == "grower":
                    leaf_collection_create=LeafCollection.objects.create(
                        leaf_receipt_id_id = leaf_weight_details.id,
                        suppler_exit_id_id = object_id,
                        weighment_supply_id_id=weighment_details.id,
                        supply_id_id = supply_details.id,
                        supplier_id = weighment_details.supplier_id,
                        nt_wght=actual_net_leaf_weight,
                        collection_date= form.instance.date_of_supply,
                        supplier_type = weighment_details.supplier_type,
                        grower_id = supplier_details.id,
                        supply_date = supply_details.date_of_supply,
                        created_by_id = self.request.user.id
                    )
                    leaf_collection_create.save()

                    if plucking_details_id:
                        leaf_collection_create.plucking_data_id=plucking_details_id
                        leaf_collection_create.save()	
					

            # elif form.instance.unloaded_vehicle_weight and str(form.instance.is_released) == "Yes" and form.instance.weighment_txn:

            #     form.instance.unloaded_vehicle_weight and str(form.instance.is_released) == "Yes" and form.instance.weighment_txn:

            #     weighment_details = WeighmentSupply.cmobjects.filter(created_by_id=self.request.user.id,\
            #                          pk=form.instance.weighment_txn.id).first()
            #     vehicle_details = weighment_details.vehicle_no
            #     vehicle_details.is_available = True
            #     vehicle_details.save()


            self.object = form.save()
            if not form.is_valid():
                return self.render_to_response(context)
            
        return response


    def form_invalid(self, form):
        error_message = 'Error saving the Form, check fields below.'
        messages.error(self.request, error_message)
        return super().form_invalid(form)
	




class SupplierExitUpdateView(LoginRequiredMixin, CommonMixin, UpdateView):
    """
     Suppllier Exit Update View
    """
    model = SupplierExit
    form_class = SupplierExitForm
    template_name = 'supplier_exit/create_form.html'

    def get(self, request, *args, **kwargs):
    # checking if the user is customer
        # try:
        #     if not self.request.user.is_superuser:
        #         messages.error(self.request, 'You have no permission to access the requested resource!')
        #         return redirect(reverse('index'))
        # except AttributeError as error:
        #     messages.error(self.request, 'You have no permission to access the requested resource!')
        #     return redirect(reverse('index'))

        return super().get(request, *args, **kwargs)

    def get_success_url(self, **kwargs):
        messages.success(self.request, 'Updated Successfully')
        return reverse('weighment_supply:supplier_exit_list')

    def get_object(self):
        weighment_details = SupplierExit.cmobjects.filter(pk=self.kwargs['pk'], created_by_id=self.request.user.id).first()
        return weighment_details

    def form_valid(self, form):

        self.pk = self.kwargs['pk']
        context = self.get_context_data()

        with transaction.atomic():
            self.object = form.save()

            if not form.is_valid():
                return self.render_to_response(context)

        return super(SupplierExitUpdateView, self).form_valid(form)


    def form_invalid(self, form):
        error_message = 'Error saving the Form, check fields below.'
        messages.error(self.request, error_message)
        return super().form_invalid(form)



class SupplierExitDetailsView(LoginRequiredMixin, CommonMixin, DetailView):
    """
     Suppllier Exit Details View
    """
    model = SupplierExit
    context_object_name = 'supplier_exit_details'
    form_class = SupplierExitForm
    template_name = 'supplier_exit/details_view.html'

    def get(self, request, *args, **kwargs):
    # checking if the user is customer
        # try:
        #     if not self.request.user.is_superuser:
        #         messages.error(self.request, 'You have no permission to access the requested resource!')
        #         return redirect(reverse('index'))
        # except AttributeError as error:
        #     messages.error(self.request, 'You have no permission to access the requested resource!')
        #     return redirect(reverse('index'))

        return super().get(request, *args, **kwargs)

    def get_object(self):
        supplier_exit_details = SupplierExit.cmobjects.filter(pk=self.kwargs['pk'], created_by_id=self.request.user.id).first()
        return supplier_exit_details
    
    def get_context_data(self, **kwargs):
        context = super(SupplierExitDetailsView, self).get_context_data(**kwargs)
        supplier_exit_details = self.get_object()
        context['form'] = SupplierExitForm(instance=supplier_exit_details)
        context['supplier_exit_details'] = supplier_exit_details
        context['supply_mode'] = supplier_exit_details.weighment_txn.mode_of_supply
        return context


    def form_invalid(self, form):
        error_message = 'Error saving the Form, check fields below.'
        messages.error(self.request, error_message)
        return super().form_invalid(form)





def supplier_exit_delete(request, pk):
    """ Supplier Exit Delete View """
    supplier_details = SupplierExit.objects.filter(pk=pk).first()
    supplier_details.is_deleted = True
    supplier_details.save()
    message = 'Deleted Successfully'
    messages.success(request, message)

    return redirect(request.META['HTTP_REFERER'])




#####################  Reports View ######################
from easy_weight_master.models import *

def lef_collection_reports_register_admin(request):
    mode = request.GET.get('mode', '')
    entity_id = request.GET.get('entity_id', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    region = request.GET.get('region', '')
    month_year = request.GET.get('month_year', '')
    
    region_list = Region.objects.all()

    distinct_regions = None
    if mode:
        mode_wise_profiles = BlfProfile.objects.filter(easy_weight_system=mode).exclude(region__isnull=True)
        # Get a list of distinct region names
        distinct_regions = mode_wise_profiles.values_list('region__region_name', flat=True).distinct()

    print("region", region)
    blf_list = None
    if region:
        blf_list = BlfProfile.cmobjects.filter(region__region_name=region)




    
    if region:
        if region.isdigit():
            region_details = Region.cmobjects.filter(pk=region).first()
        else:
            id_str = str(region) 
            region_id_details = Region.cmobjects.filter(region_name__exact=id_str).first()
            id = region_id_details.pk
            region_details = Region.cmobjects.filter(pk=id).first()
    else:
        region_details = ""

    if entity_id:
        blf_details = BlfProfile.cmobjects.filter(pk=entity_id).first()
        blf_name = blf_details.entity_unit
        blf_phone = blf_details.ho_contact_number
        blf_username = blf_details.user
        user_details = User.objects.filter(username=blf_username).first()

    else:
        blf_phone = ""
        blf_name = ""
        blf_username = ""
        user_details = ""




    print("mode #####", mode)

    if str(mode) == "Enable":

        result = EasyWeightMaster.objects.filter(created_by=user_details, month_year=month_year)
    elif str(mode) == "Disable":
        if start_date and end_date:
            result = LeafCollection.objects.filter(supply_date__range=(start_date, end_date), created_by_id=user_details.pk, nt_wght__isnull=False).exclude(nt_wght=0).order_by('-supply_date')
        else:
            result = LeafCollection.objects.filter(created_by_id=user_details.pk, nt_wght__isnull=False).exclude(nt_wght=0).order_by('-supply_date')
    else:
        result = []	



    # Paginate the result
    page = request.GET.get('page', 1)
    paginator = Paginator(result, 10)

    try:
        result = paginator.page(page)
    except PageNotAnInteger:
        result = paginator.page(1)
    except EmptyPage:
        result = paginator.page(paginator.num_pages)

    total_items = paginator.count  # Get the total count of item

    context = {
        "mode": mode,
        "start_date": start_date,
        "end_date": end_date,
        "result": result,
        "region_list" : region_list,
        "region" : region,
        "entity_id" : entity_id,
        "blf_name" : blf_name,
        "blf_username" : blf_username,
        "region_details" : region_details,
        "blf_phone" : blf_phone,
        "blf_list" : blf_list,
        "distinct_regions" : distinct_regions,
    }
    
    return CommonMixin.render(request, "reports/leaf_collections_list_admin.html", context)






#  EXCEL GENERATE LEAF COLLECTION REPORTS

def lef_collection_reports_register_admin_excel(request):
    # try:
        mode = request.GET.get('mode', '')
        entity_id = request.GET.get('entity_id', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        region = request.GET.get('region', '')


        # print(f"region #### = {region}")
        # print(f"entity_id #### = {entity_id}")
        # print(f"start_date #### = {start_date}")
        # print(f"end_date #### = {end_date}")
        # print(f"mode #### = {mode}")


        if region:
            if region.isdigit():
                region_details = Region.cmobjects.filter(pk=region).first()
            else:
                id_str = str(region) 
                region_id_details = Region.cmobjects.filter(region_name__exact=id_str).first()
                id = region_id_details.pk
                region_details = Region.cmobjects.filter(pk=id).first()
        else:
            region_details = ""

        if entity_id:
            blf_details = BlfProfile.cmobjects.filter(pk=entity_id).first()
            blf_name = blf_details.entity_unit
            blf_phone = blf_details.ho_contact_number
            blf_username = blf_details.user
            user_details = User.objects.filter(username=blf_username).first()

        else:
            blf_phone = ""
            blf_name = ""
            blf_username = ""

        if start_date and end_date and mode:
            result = LeafCollection.objects.filter(supply_date__range=(start_date, end_date), created_by_id=user_details.pk, nt_wght__isnull=False).exclude(nt_wght=0).order_by('-supply_date')
        else:
            result = []

        # Create a workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leaf Collection Report"

        # Define headers for the Excel file
        if mode == 'Disable': # Normal
            headers = ["SL No.", "Weighment Supply ID", "Grower ID", "Grower Name", "Supply Date", "Supply Qty", "FLC %", "Deduction %", "By Aggregator (if any)"]
        else:  # Easy Weight
            headers = ["SL No.", "Date", "Supplier Code", "Grower", "Phone No.", "Net Weight (Kg.)", "Gross Weight (Kg.)", "Rate(INR)"]


        # Add headers to the worksheet
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header

        

        if mode == 'Enable': # Easy Weight
            # Populate data into the worksheet
            row_num = 2  # Start from the second row
            for index, item in enumerate(result, start=1):
                print("item", item)
                ws[f"A{row_num}"] = index  # SL No.
                ws[f"B{row_num}"] = item.supply_date or "NA"
                ws[f"C{row_num}"] = item.supplier.username if item.supplier else "NA"
                if str(item.supplier_type) == "grower":
                    ws[f"D{row_num}"] = item.grower.name or "NA"
                else:
                    ws[f"D{row_num}"] = "NA"
                if str(item.supplier_type) == "grower":
                    ws[f"E{row_num}"] = item.grower.mobile_number or "NA"
                else:
                    ws[f"E{row_num}"] = item.aggregator.mobile_number or "NA"
                ws[f"F{row_num}"] = item.nt_wght or 'NA'
                ws[f"G{row_num}"] = item.weighment_supply_id.total_gross_weight_kg or 'NA'
                ws[f"H{row_num}"] = item.leaf_receipt_id.rate or 'NA'

                # Increment row_num for the next iteration
                row_num += 1
        else:
            # Populate data into the worksheet
            row_num = 2  # Start from the second row
            for index, item in enumerate(result, start=1):
                print("item", item)
                ws[f"A{row_num}"] = index  # SL No.
                ws[f"B{row_num}"] = item.weighment_supply_id.weighment_txn_id or "NA"
                ws[f"C{row_num}"] = item.supplier.username if item.supplier else "NA"
                if str(item.supplier_type) == "grower":
                    ws[f"D{row_num}"] = item.grower.name or "NA"
                else:
                    ws[f"D{row_num}"] = item.aggregator.name or "NA"
                ws[f"E{row_num}"] = item.supply_date or "NA"
                ws[f"F{row_num}"] = item.nt_wght or "NA"
                ws[f"G{row_num}"] = item.leaf_receipt_id.final_leaf_count or "NA"
                ws[f"H{row_num}"] = item.leaf_receipt_id.deduction or "NA"
                ws[f"I{row_num}"] = item.aggregator.name if item.aggregator else "NA"

                # Increment row_num for the next iteration
                row_num += 1

        # Create an HTTP response with the Excel file as an attachment
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"leaf_collection_report.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Save the workbook content to the response
        wb.save(response)

        return response

    # except (ValueError, BlfProfile.DoesNotExist, Region.DoesNotExist) as e:
    #     error_message = "Something went wrong."
    #     return render(request, 'error.html', {'error_message': error_message})

    # except Exception as e:
    #     error_message = "Something went wrong."
    #     return render(request, 'error.html', {'error_message': error_message})
        


# END EXCEL
from django.db.models import Q

# def lef_collection_reports(request):
#     supplier_type = request.GET.get('supplier_type', '')
#     start_date = request.GET.get('start_date', '')
#     end_date = request.GET.get('end_date', '')

#     if str(supplier_type) == "grower":
#         # Fetch data for 'grower' supplier_type

#         grower_list = LeafCollection.objects.filter(supplier_type__iexact='grower', created_by_id=request.user.id)
#         agg_list = LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id)

#         grower_set = set()

#         for item in agg_list:
#             # Assuming supply_id is a ForeignKey in GrowerDetailsSupply pointing to the supply in LeafCollection
#             grower_data = GrowerDetailsSupply.objects.filter(supply_id=item.supply_id_id)

#             for grower_entry in grower_data:
#                 # Assuming grower_name is the field you want to fetch
#                 grower_set.add(grower_entry.grower)

#         # Now grower_set contains unique grower names related to the aggregator LeafCollection items
#         grower_list = list(grower_set)


#         result = supplier_list

#     elif str(supplier_type) == "aggregator":
#         # Fetch data for 'aggregator' supplier_type
#         result = LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id)

#     else:
#         # Handle other cases or set a default behavior
#         result = LeafCollection.objects.none()

#     # Apply date range filtering if start_date and end_date are provided
#     if start_date and end_date:
#         result = result.filter(supply_date__range=(start_date, end_date))

#     # Paginate the result
#     page = request.GET.get('page', 1)
#     paginator = Paginator(result, 10)

#     try:
#         result = paginator.page(page)
#     except PageNotAnInteger:
#         result = paginator.page(1)
#     except EmptyPage:
#         result = paginator.page(paginator.num_pages)

#     total_items = paginator.count  # Get the total count of items

#     context = {
#         "supplier_type": supplier_type,
#         "start_date": start_date,
#         "end_date": end_date,
#         "result": result,
#         "total_items" : total_items,
#     }
#     return CommonMixin.render(request, "reports/leaf_collections_list.html", context)



@login_required
def leaf_collection_reports(request):
    supplier_type = request.GET.get('supplier_type', '').lower()
    start_date_str = request.GET.get('start_date', None)
    end_date_str = request.GET.get('end_date', None)

    start_date = None
    end_date = None
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            start_date = end_date = None

    user = request.user
    blf_profile = BlfProfile.cmobjects.filter(user=user).first()
    result = []

    # Initialize date bounds
    minimum_date = None
    maximum_date = None

    if supplier_type == "grower":
        # Fetch grower and aggregator collections
        grower_collections = LeafCollection.objects.filter(
            supplier_type__iexact='grower',
            created_by=user,
            nt_wght__isnull=False
        ).exclude(nt_wght=0).select_related('aggregator', 'grower').order_by('-supply_date')

        aggregator_collections = LeafCollection.objects.filter(
            supplier_type__iexact='aggregator',
            created_by=user,
            nt_wght__isnull=False
        ).exclude(nt_wght=0).select_related('aggregator').order_by('-supply_date')

        # Get date range if no filter
        if not start_date or not end_date:
            grower_dates = grower_collections.aggregate(
                min_date=Min('supply_date'), max_date=Max('supply_date')
            )
            agg_dates = aggregator_collections.aggregate(
                min_date=Min('supply_date'), max_date=Max('supply_date')
            )
            min_dates = [d['min_date'] for d in [grower_dates, agg_dates] if d['min_date']]
            max_dates = [d['max_date'] for d in [grower_dates, agg_dates] if d['max_date']]
            minimum_date = min(min_dates) if min_dates else None
            maximum_date = max(max_dates) if max_dates else None

        # Apply date filter if provided
        if start_date and end_date:
            grower_collections = grower_collections.filter(supply_date__range=(start_date, end_date))
            aggregator_collections = aggregator_collections.filter(supply_date__range=(start_date, end_date))

        # Build result list
        for item in grower_collections:
            result.append({
                "weighment_supply_id": item.weighment_supply_id,
                "leaf_receipt_id": item.leaf_receipt_id,
                "collected_date": item.supply_date,
                "collected_quantity": item.nt_wght,
                "net_qty": item.nt_wght,
                "grower": item.grower,
                "supplier_type": item.supplier_type,
                "aggregator": item.aggregator.name if item.aggregator else None,
            })
        
        # Handle aggregator-supplied grower details
        for item in aggregator_collections:
            grower_supplies = GrowerDetailsSupply.cmobjects.filter(supply_id=item.supply_id)
            for detail in grower_supplies:
                leaf_item = LeafCollection.objects.filter(supply_id=detail.supply).first()
                leaf_receipt_id = leaf_item.leaf_receipt_id if leaf_item else None
                weighment_supply_id = leaf_item.weighment_supply_id if leaf_item else None

                result.append({
                    "weighment_supply_id": weighment_supply_id,
                    "leaf_receipt_id": leaf_receipt_id,
                    "collected_date": detail.collected_date,
                    "collected_quantity": detail.collected_quantity,
                    "net_qty": item.nt_wght,
                    "grower": detail.grower,
                    "supplier_type": None,
                    "aggregator": item.aggregator.name if item.aggregator else None,
                })

        # Final date filtering on dictionary list if needed
        if start_date and end_date:
            result = [
                r for r in result
                if start_date <= r['collected_date'] <= end_date
            ]
    
    elif supplier_type == "aggregator":
        # Fetch only aggregator records
        qs = LeafCollection.objects.filter(
            supplier_type__iexact='aggregator',
            created_by=user,
            nt_wght__isnull=False
        ).exclude(nt_wght=0).order_by('-supply_date')

        # Get date range if no filter
        if not start_date or not end_date:
            date_range = qs.aggregate(min_date=Min('supply_date'), max_date=Max('supply_date'))
            minimum_date = date_range['min_date']
            maximum_date = date_range['max_date']

        # Apply date filter
        if start_date and end_date:
            qs = qs.filter(supply_date__range=(start_date, end_date))
            minimum_date = start_date
            maximum_date = end_date

        result = list(qs)

    else:
        result = []
        minimum_date = maximum_date = None

    # Use parsed or computed dates as fallback
    if not start_date and minimum_date:
        start_date = minimum_date
    if not end_date and maximum_date:
        end_date = maximum_date

    # Pagination
    paginator = Paginator(result, 10)
    page = request.GET.get('page', 1)
    try:
        paginated_result = paginator.page(page)
    except PageNotAnInteger:
        paginated_result = paginator.page(1)
    except EmptyPage:
        paginated_result = paginator.page(paginator.num_pages)

    context = {
        "supplier_type": supplier_type,
        "start_date": start_date,
        "end_date": end_date,
        "result": paginated_result,
        "blf_profile": blf_profile,
        "grower_dict": {},  # You can populate this if needed
        "total_items": paginator.count,
    }

    return CommonMixin.render(request, "reports/leaf_collections_list.html", context)


@login_required
def leaf_collection_reports__last(request):
    supplier_type = request.GET.get('supplier_type', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    request_user_type = str(request.GET.get('supplier_type', ''))

    grower_dict = {}  # Initialize grower_dict outside the conditional blocks

    if request_user_type == "grower":

        # Fetch data for 'grower' supplier_type
        grower_list = LeafCollection.objects.filter(supplier_type__iexact='grower', created_by_id=request.user.id, nt_wght__isnull=False).exclude(nt_wght=0).order_by('-supply_date')
        agg_list = LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id,nt_wght__isnull=False).exclude(nt_wght=0)

        # Ensure grower_list_dict contains model instances
        grower_list_dict = []
        for item in grower_list:
            # Extracting fields from LeafCollection
            challan_id = item.supply_id
            weighment_supply_id = item.weighment_supply_id
            supplier_type = item.supplier_type
            leaf_receipt_id = item.leaf_receipt_id

            supply_date = item.supply_date
            supply_qty = item.nt_wght
            grower = item.grower
            aggregator = item.aggregator.name if item.aggregator else None

            # grower_collection_details = GrowerDetailsSupply.objects.filter(supply=challan_id)

            print("challan_id", grower)

            grower_data = {
                "weighment_supply_id": weighment_supply_id,
                "leaf_receipt_id": leaf_receipt_id,
                "collected_date": supply_date,
                "collected_quantity": supply_qty,
                "net_qty" : supply_qty,
                "grower": grower,
                "supplier_type": supplier_type,
                "aggregator": aggregator


                # "grower_collection_details" : grower_collection_details,
            }

            grower_list_dict.append(grower_data)

        for item in agg_list:
            supply_id = item.supply_id
            net_qty = item.nt_wght
            grower_supply_details = GrowerDetailsSupply.objects.filter(supply_id=supply_id)
            aggregator = item.aggregator.name if item.aggregator else None
            
            for grower_item in grower_supply_details:
                # Extracting fields from GrowerDetailsSupply

                challan_id = grower_item.supply
                collected_date = grower_item.collected_date
                collected_quantity = grower_item.collected_quantity
                grower = grower_item.grower

                collection_details = LeafCollection.objects.filter(supply_id=challan_id)

                for leaf_item in collection_details:
                    leaf_receipt_id = leaf_item.leaf_receipt_id
                    weighment_supply_id = leaf_item.weighment_supply_id

                print("leaf_receipt_id ===**", leaf_receipt_id)
                
                grower_data = {
                    "weighment_supply_id": weighment_supply_id,
                    "leaf_receipt_id": leaf_receipt_id,
                    "collected_date": collected_date,
                    "collected_quantity": collected_quantity,
                    "net_qty" : net_qty,
                    "grower": grower,
                    "supplier_type": None,
                    "collection_details" : collection_details,
                    "aggregator": aggregator
                }

                grower_list_dict.append(grower_data)

        # The grower_list_dict now contains the combined data from LeafCollection and GrowerDetailsSupply
        final_list = grower_list_dict

        result = final_list

    elif request_user_type == "aggregator":
        # print(f"supplier_type-2=={supplier_type}")
        # print("YES IAM AGGREGATOR HA HA HA ******", supplier_type)
        result = LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id, nt_wght__isnull=False).exclude(nt_wght=0).order_by('-supply_date')

    else:
        # print(f"supplier_type-3=={supplier_type}")
        result = LeafCollection.objects.none()


    # Apply date range filtering if start_date and end_date are provided
    if start_date and end_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

        # Initialize minimum_date and maximum_date to None
        minimum_date = None
        maximum_date = None

        if request_user_type == "grower":
            # Convert supply_date to datetime.date for comparison
            result = [
                item for item in result
                if ('supply_date' in item and start_date <= item['supply_date'] <= end_date)
                or ('collected_date' in item and start_date <= item['collected_date'] <= end_date)
            ]
        else:
            # Apply date range filtering on the queryset
            result = result.filter(supply_date__range=(start_date, end_date))


    blf_profile = BlfProfile.objects.filter(user=request.user).first()
    
    minimum_date = datetime.strptime(str(start_date), "%Y-%m-%d") if start_date else start_date
    maximum_date = datetime.strptime(str(end_date), "%Y-%m-%d") if end_date else end_date

    if request_user_type == "grower":
        if not start_date or not end_date:
            grower_date = grower_list.aggregate(min_date=Min('supply_date'), max_date=Max('supply_date'))
            agg_date = agg_list.aggregate(min_date=Min('supply_date'), max_date=Max('supply_date'))
            # Check if the dates are not None before using them
            if grower_date['min_date'] is not None and agg_date['min_date'] is not None:
                minimum_date = min(grower_date['min_date'], agg_date['min_date'])
            if grower_date['max_date'] is not None and agg_date['max_date'] is not None:
                maximum_date = max(grower_date['max_date'], agg_date['max_date'])


    if request_user_type == "aggregator":
        if not start_date or not end_date:
            result_date = result.aggregate(min_date=Min('supply_date'), max_date=Max('supply_date'))
            # Check if the dates are not None before using them
            if result_date['min_date'] is not None:
                minimum_date = result_date['min_date']
            if result_date['max_date'] is not None:
                maximum_date = result_date['max_date']

    # print(f"FINAL LIST ==={result}")

    # Paginate the result
    page = request.GET.get('page', 1)
    paginator = Paginator(result, 10)


    try:
        result = paginator.page(page)
    except PageNotAnInteger:
        result = paginator.page(1)
    except EmptyPage:
        result = paginator.page(paginator.num_pages)

    total_items = paginator.count  # Get the total count of items

    # print(f"request_user_type-4=={request_user_type}")

    context = {
        "supplier_type": request_user_type,
        "start_date": minimum_date,
        "end_date": maximum_date,
        "result": result,
        "blf_profile": blf_profile,
        "grower_dict": grower_dict,  # Pass the grower_dict to the template
        "total_items": total_items,
    }
    
    # context = {
    #     "supplier_type": request_user_type,
    #     "start_date": start_date,
    #     "end_date": end_date,
    #     "result": result,
    #     "total_items": total_items,
    #     "grower_dict": grower_dict,  # Pass the grower_dict to the template
    # }
    return CommonMixin.render(request, "reports/leaf_collections_list.html", context)












def leaf_collection_reports__back(request):
    supplier_type = request.GET.get('supplier_type', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    request_user_type = str(request.GET.get('supplier_type', ''))

    # supplier_type = str(supplier_type)
    # print(f"request_user_type-1=={request_user_type}")

    grower_dict = {}  # Initialize grower_dict outside the conditional blocks

    if request_user_type == "grower":

        print("GROWER DATA ########### ")


        # Fetch data for 'grower' supplier_type
        grower_list = LeafCollection.objects.filter(
            supplier_type__iexact='grower',
            created_by_id=request.user.id,
            nt_wght__isnull=False,
        ).exclude(nt_wght=0).order_by('-supply_date')

        # print("grower_list", grower_list)
        print(request.user.id)
        agg_list = LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id, nt_wght__isnull=False).exclude(nt_wght=0).order_by('-supply_date')

        print("agg_list", LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id, nt_wght__isnull=False, supply_date="2024-08-29", id=4678).exclude(nt_wght=0).values('supply_date').order_by('-supply_date').count())

        # Ensure grower_list_dict contains model instances
        grower_list_dict = []
        for item in grower_list:
            # Extracting fields from LeafCollection
            challan_id = item.supply_id
            weighment_supply_id = item.weighment_supply_id
            supplier_type = item.supplier_type
            leaf_receipt_id = item.leaf_receipt_id

            supply_date = item.supply_date
            supply_qty = item.nt_wght
            grower = item.grower
            aggregator = item.aggregator.name if item.aggregator else None

            print("############# SUpply DATE", supply_date)

            # grower_collection_details = GrowerDetailsSupply.objects.filter(supply=challan_id)

            grower_data = {
                "weighment_supply_id": weighment_supply_id,
                "leaf_receipt_id": leaf_receipt_id,
                "collected_date": supply_date,
                "collected_quantity": supply_qty,
                "net_qty" : supply_qty,
                "grower": grower,
                "supplier_type": supplier_type,
                "aggregator": aggregator
                # "grower_collection_details" : grower_collection_details,
            }

            grower_list_dict.append(grower_data)


        # print(f"## SUPPLY QTY KG grower_list_dict =={grower_list_dict}")

        for item in agg_list:
            supply_id = item.supply_id
            grower_supply_details = GrowerDetailsSupply.objects.filter(supply_id=supply_id)

            net_qty = item.nt_wght

            # print("net_qty", net_qty)
            
            for grower_item in grower_supply_details:
                # Extracting fields from GrowerDetailsSupply

                challan_id = grower_item.supply
                collected_date = grower_item.collected_date
                collected_quantity = grower_item.collected_quantity
                grower = grower_item.grower
                aggregator = item.aggregator.name if item.aggregator else None


                collection_details = LeafCollection.objects.filter(supply_id=challan_id)

                for leaf_item in collection_details:
                    leaf_receipt_id = leaf_item.leaf_receipt_id
                    weighment_supply_id = leaf_item.weighment_supply_id

                # print("leaf_receipt_id ===**", leaf_receipt_id)
                
                grower_data = {
                    "weighment_supply_id": weighment_supply_id,
                    "leaf_receipt_id": leaf_receipt_id,
                    "collected_date": collected_date,
                    "collected_quantity": collected_quantity,
                    "net_qty" : net_qty,
                    "grower": grower,
                    "supplier_type": None,
                    "collection_details" : collection_details,
                    "aggregator": aggregator
                }

                grower_list_dict.append(grower_data)

        # The grower_list_dict now contains the combined data from LeafCollection and GrowerDetailsSupply
        final_list = grower_list_dict

        result = final_list

        # print(final_list)


        # Fetch data for 'grower' supplier_type
        # supplier_list = LeafCollection.objects.filter(created_by_id=request.user.id).order_by('-supply_date')
        # grower_list = LeafCollection.objects.filter(supplier_type__iexact='grower',\
        #                                              created_by_id=request.user.id).order_by('-supply_date').values()
        # agg_list = LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id)


        # # result = []
        # grower_list_dict = []
        # for item in agg_list:
        #     supply_id = item.supply_id
        #     grower_supply_details = list(GrowerDetailsSupply.objects.filter(supply_id=supply_id).values())  # Convert QuerySet to list
        #     grower_list_dict.extend(grower_supply_details)




        # Fetch data for 'grower' supplier_type
        # grower_list = LeafCollection.objects.filter(supplier_type__iexact='grower', created_by_id=request.user.id).order_by('-supply_date')
        # agg_list = LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id)

        # # Ensure grower_list_dict contains model instances
        # grower_list_dict = []
        # for item in agg_list:
        #     supply_id = item.supply_id
        #     grower_supply_details = GrowerDetailsSupply.objects.filter(supply_id=supply_id)
        #     grower_list_dict.extend(grower_supply_details)

        # # Combine grower_list and grower_list_dict
        # combined_list = list(chain(grower_list, grower_list_dict))


        # print(combined_list)
        # final_list = []
        # for item in combined_list:
        #     challan_id = item.supply_id

        #     grower_id = item.grower
        #     grower_name = item.grower.name

        
        #     # supply_date = item.supply_date


        #     leaf_collection_data = LeafCollection.objects.filter(supply_id=challan_id).first()

        #     # if item.supply_date



        #     supplier_type = leaf_collection_data.supplier_type if leaf_collection_data else None
        #     supply_date = leaf_collection_data.supply_date if leaf_collection_data else None
        #     nt_wght = leaf_collection_data.nt_wght if leaf_collection_data else None
        #     final_leaf_count = leaf_collection_data.leaf_receipt_id.final_leaf_count if leaf_collection_data else None
        #     deduction = leaf_collection_data.leaf_receipt_id.deduction if leaf_collection_data else None
        #     weighment_supply_id = leaf_collection_data.weighment_supply_id if leaf_collection_data else None
        #     supplier_type = str(supplier_type)

        #     # Fetch other details
        #     # supplier = item.supplier
        #     collection_details = LeafCollection.objects.filter(supply_id=challan_id)
        #     grower_supply_list = GrowerDetailsSupply.objects.filter(supply=challan_id)

            
        #     # print(f"grower_supply_list={grower_supply_list}")

        #     grower_data ={
        #         "weighment_supply_id" : weighment_supply_id,
        #         # "supplier" : supplier,
        #         # "supply_date" : supply_date,
        #         "nt_wght" : nt_wght,
        #         "final_leaf_count" : final_leaf_count,
        #         "deduction" : deduction,
        #         # "grower_name" : grower_name,
        #         # "grower_supplier_list" : grower_supplier_list,
        #         # "supplier_type" : supplier_type,
        #         # "deduction" : deduction,
        #         # "supplier" : supplier,

        #         "supply_date" : supply_date,
        #         "grower_name" : grower_name,
        #         "grower_id" : grower_id,
        #         "supplier_type" : supplier_type,
        #         "collection_details": collection_details,
        #         "grower_supply_list" : grower_supply_list,

        #     }

        #     final_list.append(grower_data)



             


           




        # print(combined_list)





        # for gr_item in grower_list_dict:
        #     grower_name = gr_item.supply

            # print(gr_item.grower)
            
            # weighment_supply_id = item.weighment_supply_id
            # supplier = item.supplier
            # supplier_type = str(item.supplier_type)
            # supply_date = item.supply_date
            # nt_wght  = item.nt_wght
            # final_leaf_count = item.leaf_receipt_id.final_leaf_count
            # deduction = item.leaf_receipt_id.deduction

            # grower_name = item.grower
            # grower_supplier_list = GrowerDetailsSupply.cmobjects.filter(supply_id=item.supply_id_id).order_by('-id')


            # if str(item.supplier_type) == "grower":
            #     grower_name = item.grower.name
            # else:
            #     grower_name = ""
            # if str(item.supplier_type) == "aggregator":

            #     grower_supplier_list = GrowerDetailsSupply.cmobjects.filter(supply_id=item.supply_id_id).order_by('-id')


                # print("grower_name", grower_name)

            # grower_data ={
            #     "weighment_supply_id" : weighment_supply_id,
            #     "supplier" : supplier,
            #     "supply_date" : supply_date,
            #     "nt_wght" : nt_wght,
            #     "final_leaf_count" : final_leaf_count,
            #     "grower_name" : grower_name,
            #     "grower_supplier_list" : grower_supplier_list,
            #     "supplier_type" : supplier_type,
            #     "deduction" : deduction,
            # }

            # grower_list.append(grower_supplier_list)

    


        # grower_result = []
        # for i in grower_supplier_list:
        #     supplier = i.supply

        #     weighment_details = {}

        #     weighment_supply_id = item.weighment_supply_id



            # print(f"grower_supplier_list=={supply_id}")

        # combined_list = list(chain(grower_list, grower_list_dict))

        # final_result = []
        # for i in combined_list:

        #     weighment_supply_id = i.get('weighment_supply_id_id')
           


        # print("result=====", result)

        

        # print(f"result={grower_list}")

    elif request_user_type == "aggregator":
        # print(f"supplier_type-2=={supplier_type}")
        # print("YES IAM AGGREGATOR HA HA HA ******", supplier_type)
        result = LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id, nt_wght__isnull=False,).exclude(nt_wght=0).order_by('-supply_date')

    else:
        # print(f"supplier_type-3=={supplier_type}")
        result = LeafCollection.objects.none()


    # Apply date range filtering if start_date and end_date are provided
    if start_date and end_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

        if request_user_type == "grower":
            # Convert supply_date to datetime.date for comparison
            result = [
                item for item in result
                if ('supply_date' in item and start_date <= item['supply_date'] <= end_date)
                or ('collected_date' in item and start_date <= item['collected_date'] <= end_date)
            ]
        else:
            # Apply date range filtering on the queryset
            result = result.filter(supply_date__range=(start_date, end_date))


    # print(f"FINAL LIST ==={result}")

    # Paginate the result
    page = request.GET.get('page', 1)
    paginator = Paginator(result, 10)


    try:
        result = paginator.page(page)
    except PageNotAnInteger:
        result = paginator.page(1)
    except EmptyPage:
        result = paginator.page(paginator.num_pages)

    total_items = paginator.count  # Get the total count of items

    # print(f"request_user_type-4=={request_user_type}")


    

    context = {
        "supplier_type": request_user_type,
        "start_date": start_date,
        "end_date": end_date,
        "result": result,
        "total_items": total_items,
        "grower_dict": grower_dict,  # Pass the grower_dict to the template
    }
    return CommonMixin.render(request, "reports/leaf_collections_list.html", context)
























# def lef_collection_reports(request):
#     supplier_type = request.GET.get('supplier_type', '')
#     start_date = request.GET.get('start_date', '')
#     end_date = request.GET.get('end_date', '')


#     if str(supplier_type) == "grower":
        
#         if start_date and end_date and supplier_type:
#             result = LeafCollection.objects.filter(supply_date__range=(start_date, end_date), supplier_type__icontains=supplier_type, created_by_id=request.user.id)
#         elif supplier_type:
#             result = LeafCollection.objects.filter(supplier_type__icontains=supplier_type, created_by_id=request.user.id).order_by('-id')
#         else:
#             result = []


#     elif str(supplier_type) == "aggregator":

#         if start_date and end_date and supplier_type:
#             result = LeafCollection.objects.filter(supply_date__range=(start_date, end_date), supplier_type__icontains=supplier_type, created_by_id=request.user.id)
#         elif supplier_type:
#             result = LeafCollection.objects.filter(supplier_type__icontains=supplier_type, created_by_id=request.user.id).order_by('-id')
#         else:
#             result = []



#     # Paginate the result
#     page = request.GET.get('page', 1)
#     paginator = Paginator(result, 10)

#     try:
#         result = paginator.page(page)
#     except PageNotAnInteger:
#         result = paginator.page(1)
#     except EmptyPage:
#         result = paginator.page(paginator.num_pages)

#     total_items = paginator.count  # Get the total count of item

#     context = {
#         "supplier_type": supplier_type,
#         "start_date": start_date,
#         "end_date": end_date,
#         "result": result,
#         "total_items" : total_items,
#     }

#     return CommonMixin.render(request, "reports/leaf_collections_list.html", context)


# REPORTS PDF GENERATE
def lef_collection_supplier_wise_export(request):
    export_format = request.GET.get('export_format', '')
    supplier_type = request.GET.get('supplier_type', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    request_user_type = str(request.GET.get('supplier_type', ''))
    print("export_format", export_format)

    # supplier_type = str(supplier_type)
    # print(f"request_user_type-1=={request_user_type}")

    grower_dict = {}  # Initialize grower_dict outside the conditional blocks

    if request_user_type == "grower":

        # Fetch data for 'grower' supplier_type
        grower_list = LeafCollection.objects.filter(supplier_type__iexact='grower', created_by_id=request.user.id, nt_wght__isnull=False).exclude(nt_wght=0).order_by('-supply_date')
        agg_list = LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id,nt_wght__isnull=False).exclude(nt_wght=0)

        # Ensure grower_list_dict contains model instances
        grower_list_dict = []
        for item in grower_list:
            # Extracting fields from LeafCollection
            challan_id = item.supply_id
            weighment_supply_id = item.weighment_supply_id
            supplier_type = item.supplier_type
            leaf_receipt_id = item.leaf_receipt_id

            supply_date = item.supply_date
            supply_qty = item.nt_wght
            grower = item.grower
            aggregator = item.aggregator.name if item.aggregator else None

            # grower_collection_details = GrowerDetailsSupply.objects.filter(supply=challan_id)

            print("challan_id", grower)

            grower_data = {
                "weighment_supply_id": weighment_supply_id,
                "leaf_receipt_id": leaf_receipt_id,
                "collected_date": supply_date,
                "collected_quantity": supply_qty,
                "net_qty" : supply_qty,
                "grower": grower,
                "supplier_type": supplier_type,
                "aggregator": aggregator


                # "grower_collection_details" : grower_collection_details,
            }

            grower_list_dict.append(grower_data)

        for item in agg_list:
            supply_id = item.supply_id
            net_qty = item.nt_wght
            grower_supply_details = GrowerDetailsSupply.objects.filter(supply_id=supply_id)
            aggregator = item.aggregator.name if item.aggregator else None
            
            for grower_item in grower_supply_details:
                # Extracting fields from GrowerDetailsSupply

                challan_id = grower_item.supply
                collected_date = grower_item.collected_date
                collected_quantity = grower_item.collected_quantity
                grower = grower_item.grower

                collection_details = LeafCollection.objects.filter(supply_id=challan_id)

                for leaf_item in collection_details:
                    leaf_receipt_id = leaf_item.leaf_receipt_id
                    weighment_supply_id = leaf_item.weighment_supply_id

                print("leaf_receipt_id ===**", leaf_receipt_id)
                
                grower_data = {
                    "weighment_supply_id": weighment_supply_id,
                    "leaf_receipt_id": leaf_receipt_id,
                    "collected_date": collected_date,
                    "collected_quantity": collected_quantity,
                    "net_qty" : net_qty,
                    "grower": grower,
                    "supplier_type": None,
                    "collection_details" : collection_details,
                    "aggregator": aggregator
                }

                grower_list_dict.append(grower_data)

        # The grower_list_dict now contains the combined data from LeafCollection and GrowerDetailsSupply
        final_list = grower_list_dict

        result = final_list

    elif request_user_type == "aggregator":
        # print(f"supplier_type-2=={supplier_type}")
        # print("YES IAM AGGREGATOR HA HA HA ******", supplier_type)
        result = LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id, nt_wght__isnull=False).exclude(nt_wght=0).order_by('-supply_date')

    else:
        # print(f"supplier_type-3=={supplier_type}")
        result = LeafCollection.objects.none()


    # Apply date range filtering if start_date and end_date are provided
    if start_date and end_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

        # Initialize minimum_date and maximum_date to None
        minimum_date = None
        maximum_date = None

        if request_user_type == "grower":
            # Convert supply_date to datetime.date for comparison
            result = [
                item for item in result
                if ('supply_date' in item and start_date <= item['supply_date'] <= end_date)
                or ('collected_date' in item and start_date <= item['collected_date'] <= end_date)
            ]
        else:
            # Apply date range filtering on the queryset
            result = result.filter(supply_date__range=(start_date, end_date))


    blf_profile = BlfProfile.objects.filter(user=request.user).first()
    
    minimum_date = datetime.strptime(str(start_date), "%Y-%m-%d") if start_date else start_date
    maximum_date = datetime.strptime(str(end_date), "%Y-%m-%d") if end_date else end_date

    if request_user_type == "grower":
        if not start_date or not end_date:
            grower_date = grower_list.aggregate(min_date=Min('supply_date'), max_date=Max('supply_date'))
            agg_date = agg_list.aggregate(min_date=Min('supply_date'), max_date=Max('supply_date'))
            # Check if the dates are not None before using them
            if grower_date['min_date'] is not None and agg_date['min_date'] is not None:
                minimum_date = min(grower_date['min_date'], agg_date['min_date'])
            if grower_date['max_date'] is not None and agg_date['max_date'] is not None:
                maximum_date = max(grower_date['max_date'], agg_date['max_date'])


    if request_user_type == "aggregator":

        if not start_date or not end_date:
            result_date = result.aggregate(min_date=Min('supply_date'), max_date=Max('supply_date'))
            # Check if the dates are not None before using them
            if result_date['min_date'] is not None:
                minimum_date = result_date['min_date']
            if result_date['max_date'] is not None:
                maximum_date = result_date['max_date']

    context = {
        "supplier_type": request_user_type,
        "start_date": minimum_date,
        "end_date": maximum_date,
        "result": result,
        "blf_profile": blf_profile,
        "grower_dict": grower_dict,  # Pass the grower_dict to the template
    }

    if export_format == 'excel':
        # Create a workbook and add a worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leaf Collection Report"

        # Define headers for the Excel file
        if supplier_type == "aggregator":
            headers = ["SL No.", "TXN ID", "Aggregator ID", "Aggregator Name", "Supply Date", "Supply Qty (in Kg.)", "FLC %", "Deduction %", "Vehicle No. (if any)"]
        else:
            headers = ["SL No.", "TXN ID", "Grower ID", "Grower Name", "Supply Date", "Supply Qty (in Kg.)", "FLC %", "Deduction %", "By Aggregator (if any)"]


        # Add headers to the worksheet
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
        print("supplier_type", supplier_type)
        if supplier_type == "aggregator":
            # Populate data into the worksheet
            row_num = 2  # Start from the second row
            print("result", result)
            for index, item in enumerate(result, start=1):
                ws[f"A{row_num}"] = index  # SL No.
                ws[f"B{row_num}"] = str(item.weighment_supply_id) or "NA"
                ws[f"C{row_num}"] = str(item.supplier) or "NA"
                if item.aggregator:
                    ws[f"D{row_num}"] = item.aggregator.name or "NA"
                else:
                    ws[f"D{row_num}"] = item.grower.name or "NA"
                ws[f"E{row_num}"] = item.supply_date or "NA"
                ws[f"F{row_num}"] = item.nt_wght or "NA"
                ws[f"G{row_num}"] = str(item.leaf_receipt_id.final_leaf_count) + "%" or "NA"
                ws[f"H{row_num}"] = str(item.leaf_receipt_id.deduction) + "%" or "NA"
                ws[f"I{row_num}"] = str(item.supply_id.alloted_vehicle) or "NA"

                # Increment row_num for the next iteration
                row_num += 1
        else:
            # Populate data into the worksheet
            row_num = 2  # Start from the second row
            for index, item in enumerate(result, start=1):
                print("grower")
                print("item", item)
                ws[f"A{row_num}"] = index  # SL No.
                ws[f"B{row_num}"] = str(item["weighment_supply_id"]) or "NA"
                ws[f"C{row_num}"] = str(item["grower"]) or "NA"
                ws[f"D{row_num}"] = item["grower"].name or "NA"
                ws[f"E{row_num}"] = item["collected_date"] or "NA"
                ws[f"F{row_num}"] = str(item["net_qty"]) or "NA"
                ws[f"G{row_num}"] = str(item["leaf_receipt_id"].final_leaf_count) + "%" or "NA"
                ws[f"H{row_num}"] = str(item["leaf_receipt_id"].deduction) + "%" or "NA"
                ws[f"I{row_num}"] = item["aggregator"] or "NA"

                # Increment row_num for the next iteration
                row_num += 1
        
    
        # Create an HTTP response with the Excel file as an attachment
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"leaf_collection_report.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Save the workbook content to the response
        wb.save(response)

        return response

    else:

        # return CommonMixin.render(request, "blf_reports_pdf/supplier_wise_lear_reports_pdf.html", context)

        # Configure pdfkit options
        options = {
            'page-size': 'letter',
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
        }

        # Set the path to the wkhtmltopdf executable
        
        path_wkhtmltopdf = str(os.getenv('WKHTMLTOPDF_PATH',''))

        # Create an HTML template
        template_name = 'blf_reports_pdf/supplier_wise_lear_reports_pdf.html'
        html = render_to_string(template_name, context)

        config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)

        # Generate PDF
        pdf = pdfkit.from_string(html, False, options, configuration=config)

        # Create HTTP response
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="supplier_wise_leaf_collection_pdf.pdf"'

        return response













# Backup
def lef_collection_supplier_wise_export_backup(request):
    """ Labour Data generate pdf View """
    from django.http import HttpResponse
    from xhtml2pdf import pisa
    from django.template.loader import get_template

    supplier_type= request.GET.get('supplier_type')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    print("PDF GENERATE #####################", supplier_type)

    # if start_date and end_date and supplier_type:
    #     result = LeafCollection.objects.filter(
    #         supply_date__range=(start_date,end_date),
    #         supplier_type__icontains=supplier_type,
    #         created_by_id=request.user.id)
    # elif supplier_type:
    #     result = LeafCollection.objects.filter(supplier_type__icontains=supplier_type, created_by_id=request.user.id).order_by('-id')
    # else:
    #     result = ""


    request_user_type = str(request.GET.get('supplier_type', ''))

    # supplier_type = str(supplier_type)
    print(f"request_user_type-1=={request_user_type}")

    grower_dict = {}  # Initialize grower_dict outside the conditional blocks
    if request_user_type == "grower":
        print("YES IAM GROWER ******", supplier_type)

        # Fetch data for 'grower' supplier_type
        supplier_list = LeafCollection.objects.filter(created_by_id=request.user.id).order_by('-supply_date')
        agg_list = LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id)

        result = []
        for item in supplier_list:
            
            weighment_supply_id = item.weighment_supply_id
            supplier = item.supplier
            supplier_type = str(item.supplier_type)
            supply_date = item.supply_date
            nt_wght  = item.nt_wght
            final_leaf_count = item.leaf_receipt_id.final_leaf_count
            deduction = item.leaf_receipt_id.deduction

            if str(item.supplier_type) == "grower":
                grower_name = item.grower.name
            else:
                grower_name = GrowerDetailsSupply.cmobjects.filter(supply_id=item.supply_id_id).order_by('-id')

            grower_data ={
                "weighment_supply_id" : weighment_supply_id,
                "supplier" : supplier,
                "supply_date" : supply_date,
                "nt_wght" : nt_wght,
                "final_leaf_count" : final_leaf_count,
                "grower_name" : grower_name,
                "supplier_type" : supplier_type,
                "deduction" : deduction,
            }

            result.append(grower_data)

    elif request_user_type == "aggregator":
        print(f"supplier_type-2=={supplier_type}")
        print("YES IAM AGGREGATOR HA HA HA ******", supplier_type)
        result = LeafCollection.objects.filter(supplier_type__iexact='aggregator', created_by_id=request.user.id).order_by('-supply_date')

    else:
        print(f"supplier_type-3=={supplier_type}")
        result = LeafCollection.objects.none()


    # Apply date range filtering if start_date and end_date are provided
    if start_date and end_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

        if request_user_type == "grower":
            # Convert supply_date to datetime.date for comparison
            result = [item for item in result if start_date <= item['supply_date'] <= end_date]
        else:
            # Apply date range filtering on the queryset
            result = result.filter(supply_date__range=(start_date, end_date))


    context = {
        "supplier_type": request_user_type,
        "start_date": start_date,
        "end_date": end_date,
        "result": result,
        "grower_dict": grower_dict,  # Pass the grower_dict to the template
    }



    return CommonMixin.render(request, "blf_reports_pdf/supplier_wise_lear_reports_pdf.html", context)

    # # Create a response with PDF content
    # response = HttpResponse(content_type='application/pdf')
    # response['Content-Disposition'] = f'attachment; filename="supplier_wise_leaf_collection_pdf.pdf"'

    # # Create an HTML tem    plate
    # template = get_template('blf_reports_pdf/supplier_wise_lear_reports_pdf.html')
    # html = template.render(context)  # Replace with your template context data

    # # Generate PDF
    # pisa.CreatePDF(html, dest=response) 
    # return response
    
    # start pdf-kit 

    # # Configure pdfkit options
    # options = {
    #     'page-size': 'letter',
    #     'margin-top': '0.75in',
    #     'margin-right': '0.75in',
    #     'margin-bottom': '0.75in',
    #     'margin-left': '0.75in',
    # }

    # # Set the path to the wkhtmltopdf executable
    # path_wkhtmltopdf = '/usr/bin/wkhtmltopdf'

    # # Create an HTML template
    # template_name = 'blf_reports_pdf/supplier_wise_lear_reports_pdf.html'
    # html = render_to_string(template_name, context)

    # config = pdfkit.configuration(wkhtmltopdf=path_wkhtmltopdf)

    # # Generate PDF
    # pdf = pdfkit.from_string(html, False, options, configuration=config)

    # # Create HTTP response
    # response = HttpResponse(pdf, content_type='application/pdf')
    # response['Content-Disposition'] = 'attachment; filename="supplier_wise_leaf_collection_pdf.pdf"'

    # return response

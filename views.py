from django.shortcuts import render, redirect
from django.contrib import messages
from django.urls import reverse
from django import forms as django_forms
from django.db.models import Q
from django.http import JsonResponse
from django.db.models.functions import Coalesce
from django.db.models import Sum, Count, Value, OuterRef, Subquery, IntegerField
from django.template.loader import render_to_string
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect
from django.contrib.auth import get_user_model
User = get_user_model()
from django.contrib.auth.decorators import login_required
from django.views.generic import ListView, DetailView, UpdateView, CreateView
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.paginator import Paginator , EmptyPage, PageNotAnInteger
import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from django.db import transaction
from django.http import HttpResponse
from datetime import datetime
from master.common import CommonMixin
from .models import *
from .forms import *
from accounts.forms import *
from gardens_managment.models import *
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser, IsAuthenticatedOrReadOnly, AllowAny
from django.core.paginator import Paginator
from knox.auth import TokenAuthentication
from weighment_supply.models import *
from master.utils import *
from django.template import loader
from django.utils import timezone
import pytz
from master.decorators import *
from django.template.loader import get_template
from xhtml2pdf import pisa   
from .helpers import *
import datetime
import requests
from django.urls import resolve
import pandas as pd
import fuzzymatcher
from django.db.models import Q
import math
import numpy as np
import requests
import certifi
import fuzzymatcher
from .serializers import *
from .aggregator_api_models import SupplyManagement
from user_profile.blf_api_models import SupplierExit
from leaf_receipt.models import LeafManagement, LeafCollection
from vehicle_management.models import VehicleManagement


def _can_access_challan_process(request):
	if not request.user.is_authenticated:
		return False
	if request.user.is_superuser:
		return True
	profile = Profile.cmobjects.filter(user=request.user).select_related("user_type").first()
	return bool(profile and profile.user_type and str(profile.user_type.name).lower() == "blf")


def _challan_process_blf_list():
	challan_count = (
		WeighmentSupply.cmobjects
		.filter(
			created_by_id=OuterRef("user_id"),
			weignment_supply_txn_id__isnull=False,
			available_weighment_id__isnull=False,
		)
		.values("created_by_id")
		.annotate(total=Count("id"))
		.values("total")
	)
	return (
		BlfProfile.cmobjects
		.select_related("user")
		.only("id", "entity_unit", "entity_name", "email", "user_id", "user__username")
		.annotate(challan_count=Coalesce(Subquery(challan_count[:1], output_field=IntegerField()), Value(0)))
		.order_by("entity_unit")
	)


def _selected_challan_process_blf(request):
	if request.user.is_superuser:
		blf_id = request.GET.get("blf_user") or request.POST.get("blf_user")
		if not blf_id:
			return None
		return BlfProfile.cmobjects.select_related("user").filter(pk=blf_id).first()
	return BlfProfile.cmobjects.select_related("user").filter(user=request.user).first()


def _challan_process_owner_user(request):
	selected_blf = _selected_challan_process_blf(request)
	if selected_blf and selected_blf.user_id:
		return selected_blf.user
	if request.user.is_superuser:
		return None
	return request.user


def _challan_process_context_query(request):
	blf_id = request.GET.get("blf_user") or request.POST.get("blf_user")
	return f"?blf_user={blf_id}" if blf_id else ""


def _challan_process_queryset(request, owner_user=None):
	owner_user = owner_user if owner_user is not None else _challan_process_owner_user(request)
	queryset = WeighmentSupply.cmobjects.select_related(
		"supplier",
		"supply_challan",
		"vehicle_no",
		"created_by",
	).filter(
		weignment_supply_txn_id__isnull=False,
		available_weighment_id__isnull=False,
	).distinct().order_by("-id")
	if owner_user:
		queryset = queryset.filter(created_by=owner_user)
	else:
		queryset = queryset.none()
	return queryset


def _supplier_choices_for_challan_process(request, current_user_id=None, owner_blf=None, limit=80):
	owner_blf = owner_blf if owner_blf is not None else _selected_challan_process_blf(request)
	if owner_blf:
		aggregators = AggregatorProfile.cmobjects.select_related("user").filter(associated_entity=owner_blf).only("user_id", "name", "user__username").distinct().order_by("name")
		growers = GrowerProfile.cmobjects.select_related("user").filter(associated_entity=owner_blf).only("user_id", "name", "user__username").distinct().order_by("name")
	elif request.user.is_superuser:
		aggregators = AggregatorProfile.cmobjects.none()
		growers = GrowerProfile.cmobjects.none()
	else:
		aggregators = AggregatorProfile.cmobjects.select_related("user").filter(associated_entity__user=request.user).only("user_id", "name", "user__username").distinct().order_by("name")
		growers = GrowerProfile.cmobjects.select_related("user").filter(associated_entity__user=request.user).only("user_id", "name", "user__username").distinct().order_by("name")
	choices = []
	per_group_limit = max(int(limit / 2), 1) if limit else None
	if per_group_limit:
		aggregators = aggregators[:per_group_limit]
		growers = growers[:per_group_limit]
	for profile in aggregators:
		if profile.user_id:
			choices.append((profile.user_id, f"Aggregator - {profile.name or profile.user.username}", "aggregator"))
	for profile in growers:
		if profile.user_id:
			choices.append((profile.user_id, f"Grower - {profile.name or profile.user.username}", "grower"))
	if current_user_id and str(current_user_id) not in {str(user_id) for user_id, _label, _supplier_type in choices}:
		aggregator = AggregatorProfile.cmobjects.select_related("user").only("user_id", "name", "user__username").filter(user_id=current_user_id).first()
		grower = None if aggregator else GrowerProfile.cmobjects.select_related("user").only("user_id", "name", "user__username").filter(user_id=current_user_id).first()
		if aggregator and aggregator.user_id:
			choices.append((aggregator.user_id, f"Aggregator - {aggregator.name or aggregator.user.username}", "aggregator"))
		elif grower and grower.user_id:
			choices.append((grower.user_id, f"Grower - {grower.name or grower.user.username}", "grower"))
	return choices


class ChallanProcessForm(django_forms.Form):
	supplier = django_forms.ChoiceField(required=True)
	supplier_type = django_forms.ChoiceField(choices=(("aggregator", "Aggregator"), ("grower", "Grower")), required=True)
	supply_challan = django_forms.ModelChoiceField(queryset=SupplyManagement.cmobjects.none(), required=False, empty_label="Select challan")
	mode_of_supply = django_forms.ChoiceField(choices=WeighmentSupply.supply_option, required=False)
	supply_date = django_forms.DateField(required=True, widget=django_forms.DateInput(attrs={"type": "date"}))
	vehicle_no = django_forms.ModelChoiceField(queryset=VehicleManagement.cmobjects.none(), required=False, empty_label="Select vehicle")
	total_gross_weight_kg = django_forms.FloatField(required=True, min_value=0)
	tare_weight_kg = django_forms.FloatField(required=True, min_value=0)
	mobile_number = django_forms.CharField(required=False, max_length=17)
	deduction = django_forms.FloatField(required=True, min_value=0, max_value=100)
	final_leaf_count = django_forms.FloatField(required=True, min_value=0, max_value=100)
	net_leaf_weight = django_forms.FloatField(required=True, min_value=0)
	rate = django_forms.CharField(required=False, max_length=200)
	quality_standard = django_forms.ChoiceField(choices=LeafManagement.QUALITY_STANDARD, required=True)
	acknowledge_status = django_forms.ChoiceField(choices=LeafManagement.ACKNOWLEDGE_STATUS, required=True)
	unloaded_vehicle_weight = django_forms.FloatField(required=True, min_value=0)
	is_released = django_forms.ChoiceField(choices=SupplierExit.released_option, required=True)

	def __init__(self, *args, **kwargs):
		self.request = kwargs.pop("request")
		self.owner_blf = kwargs.pop("owner_blf", None)
		self.owner_user = kwargs.pop("owner_user", None)
		super().__init__(*args, **kwargs)
		self.fields["supply_date"].input_formats = ["%Y-%m-%d"]
		self.fields["supply_date"].widget.format = "%Y-%m-%d"
		self.fields["supplier"].label = "Supplier"
		self.fields["supplier_type"].label = "Supplier Type"
		self.fields["supply_challan"].label = "Supply Challan"
		self.fields["mode_of_supply"].label = "Mode of Supply"
		self.fields["supply_date"].label = "Date of Supply"
		self.fields["vehicle_no"].label = "Vehicle Number"
		self.fields["total_gross_weight_kg"].label = "Gross Weight (Kgs)"
		self.fields["tare_weight_kg"].label = "Tare Weight (Kgs)"
		self.fields["mobile_number"].label = "Mobile Number"
		self.fields["deduction"].label = "Deduction %"
		self.fields["final_leaf_count"].label = "Final Leaf %"
		self.fields["net_leaf_weight"].label = "Net Weight (Kgs)"
		self.fields["rate"].label = "Rate"
		self.fields["quality_standard"].label = "Quality Standard"
		self.fields["acknowledge_status"].label = "Acknowledgement Status"
		self.fields["unloaded_vehicle_weight"].label = "Unloaded Vehicle Weight"
		self.fields["is_released"].label = "Is Released?"
		current_supplier_id = self.data.get("supplier") or self.initial.get("supplier")
		self.supplier_choices = _supplier_choices_for_challan_process(self.request, current_user_id=current_supplier_id, owner_blf=self.owner_blf)
		self.valid_supplier_types = {
			str(user_id): choice_type for user_id, _label, choice_type in self.supplier_choices
		}
		self.fields["supplier"].choices = [("", "Select supplier")] + [
			(str(user_id), label) for user_id, label, _supplier_type in self.supplier_choices
		]
		current_challan_id = self.data.get("supply_challan") or self.initial.get("supply_challan")
		current_vehicle_id = self.data.get("vehicle_no") or self.initial.get("vehicle_no")
		if self.owner_user:
			supply_qs = SupplyManagement.cmobjects.filter(consumer=self.owner_user, is_weighment_proceed=False)
			vehicle_qs = VehicleManagement.cmobjects.filter(created_by=self.owner_user, is_available=True)
		elif self.request.user.is_superuser:
			supply_qs = SupplyManagement.cmobjects.none()
			vehicle_qs = VehicleManagement.cmobjects.none()
		else:
			supply_qs = SupplyManagement.cmobjects.filter(consumer=self.request.user, is_weighment_proceed=False)
			vehicle_qs = VehicleManagement.cmobjects.filter(created_by=self.request.user, is_available=True)
		supply_ids = list(supply_qs.order_by("-id").values_list("id", flat=True)[:80])
		if current_challan_id:
			supply_ids.append(current_challan_id)
		vehicle_ids = list(vehicle_qs.order_by("vehicle_number").values_list("id", flat=True)[:80])
		if current_vehicle_id:
			vehicle_ids.append(current_vehicle_id)
		self.fields["supply_challan"].queryset = SupplyManagement.cmobjects.filter(id__in=supply_ids).only("id", "supply_challan_id").order_by("-id")
		self.fields["vehicle_no"].queryset = VehicleManagement.cmobjects.filter(id__in=vehicle_ids).only("id", "vehicle_number").order_by("vehicle_number")
		for field in self.fields.values():
			existing_class = field.widget.attrs.get("class", "")
			field.widget.attrs["class"] = f"{existing_class} form-control".strip()
			field.widget.attrs.setdefault("placeholder", field.label)
			if field.required:
				field.widget.attrs["required"] = "required"

	def clean(self):
		cleaned_data = super().clean()
		supplier = cleaned_data.get("supplier")
		supplier_type = cleaned_data.get("supplier_type")
		if supplier and self.valid_supplier_types.get(str(supplier)) != supplier_type:
			raise django_forms.ValidationError("Selected supplier does not match supplier type.")
		total_gross = cleaned_data.get("total_gross_weight_kg") or 0
		unloaded = cleaned_data.get("unloaded_vehicle_weight") or 0
		if unloaded > total_gross:
			raise django_forms.ValidationError("Unloaded vehicle weight cannot be more than total gross weight.")
		return cleaned_data

	def api_payload(self):
		data = self.cleaned_data.copy()
		if data.get("supply_challan"):
			data["supply_challan"] = data["supply_challan"].id
		if data.get("vehicle_no"):
			data["vehicle_no"] = data["vehicle_no"].id
		return data


def _challan_process_form_initial(weighment):
	leaf_receipt = LeafManagement.cmobjects.filter(weighment_txn=weighment).order_by("-id").first()
	supplier_exit = SupplierExit.cmobjects.filter(weighment_txn=weighment).order_by("-id").first()
	return {
		"supplier": str(weighment.supplier_id) if weighment.supplier_id else "",
		"supplier_type": weighment.supplier_type,
		"supply_challan": weighment.supply_challan_id,
		"mode_of_supply": weighment.mode_of_supply,
		"supply_date": weighment.supply_date,
		"vehicle_no": weighment.vehicle_no_id,
		"total_gross_weight_kg": weighment.total_gross_weight_kg,
		"tare_weight_kg": weighment.tare_weight_kg,
		"mobile_number": weighment.mobile_number,
		"deduction": leaf_receipt.deduction if leaf_receipt else None,
		"final_leaf_count": leaf_receipt.final_leaf_count if leaf_receipt else None,
		"net_leaf_weight": leaf_receipt.net_leaf_weight if leaf_receipt else None,
		"rate": leaf_receipt.rate if leaf_receipt else None,
		"quality_standard": leaf_receipt.quality_standard if leaf_receipt else None,
		"acknowledge_status": leaf_receipt.acknowledge_status if leaf_receipt else None,
		"unloaded_vehicle_weight": supplier_exit.unloaded_vehicle_weight if supplier_exit else None,
		"is_released": supplier_exit.is_released if supplier_exit else None,
	}


def _challan_process_grower_rows(weighment):
	if not weighment or not weighment.supply_challan_id:
		return []
	rows = []
	for item in GrowerDetailsSupply.objects.filter(supply_id=weighment.supply_challan_id).select_related("grower").order_by("id"):
		rows.append({
			"name": item.grower.name if item.grower else "NA",
			"initial_weight": item.collected_quantity or item.supply_quantity or "NA",
			"final_weight": item.supply_quantity or item.collected_quantity or "NA",
		})
	return rows


def _challan_process_list_url(request):
	query = request.GET.copy()
	query.pop("next", None)
	url = reverse("user_profile:challan_process_list")
	if query.urlencode():
		return f"{url}?{query.urlencode()}"
	return url


def _challan_process_records_url(request):
	query = request.GET.copy()
	query.pop("next", None)
	url = reverse("user_profile:challan_process_records")
	if query.urlencode():
		return f"{url}?{query.urlencode()}"
	return url


def _supply_management_report_url(request):
	query = request.GET.copy()
	query.pop("next", None)
	url = reverse("user_profile:supply_management_report")
	if query.urlencode():
		return f"{url}?{query.urlencode()}"
	return url


def _supply_management_report_queryset(search_query, date_of_supply_from=None, date_of_supply_to=None):
	queryset = SupplyManagement.cmobjects.select_related(
		"consumer",
		"alloted_vehicle",
		"created_by",
		"created_by__profile",
		"created_by__profile__user_type",
	).order_by("-created_at", "-id")

	if search_query:
		queryset = queryset.filter(
			Q(supply_challan_id__icontains=search_query) |
			Q(supply_to__icontains=search_query) |
			Q(vehicle_option__icontains=search_query) |
			Q(alloted_vehicle__vehicle_number__icontains=search_query) |
			Q(gross_leaf__icontains=search_query) |
			Q(quantity__icontains=search_query) |
			Q(supply_bag_id__icontains=search_query) |
			Q(driver_name__icontains=search_query) |
			Q(mobile_number__icontains=search_query) |
			Q(consumer__username__icontains=search_query) |
			Q(created_by__username__icontains=search_query) |
			Q(created_by__profile__user_type__name__icontains=search_query)
		)
	if date_of_supply_from:
		queryset = queryset.filter(date_of_supply__gte=date_of_supply_from)
	if date_of_supply_to:
		queryset = queryset.filter(date_of_supply__lte=date_of_supply_to)
	return queryset


@permission_required_admin
def supply_management_report(request):
	search_query = request.GET.get("q", "").strip()
	date_of_supply_from = request.GET.get("date_of_supply_from", "").strip()
	date_of_supply_to = request.GET.get("date_of_supply_to", "").strip()
	queryset = _supply_management_report_queryset(search_query, date_of_supply_from, date_of_supply_to)
	paginator = Paginator(queryset, 10)
	page_number = request.GET.get("page", 1)
	page_obj = paginator.get_page(page_number)

	context = {
		"supply_list": page_obj.object_list,
		"page_obj": page_obj,
		"paginator": paginator,
		"is_paginated": page_obj.has_other_pages(),
		"search_query": search_query,
		"date_of_supply_from": date_of_supply_from,
		"date_of_supply_to": date_of_supply_to,
		"list_url": reverse("user_profile:supply_management_report"),
		"current_list_url": _supply_management_report_url(request),
		"total_count": paginator.count,
	}
	if _is_ajax_request(request):
		return CommonMixin.render(request, "reports/_supply_management_table.html", context)
	return CommonMixin.render(request, "reports/supply_management_list.html", context)


@permission_required_admin
def supply_management_report_detail(request, id):
	next_url = get_safe_next_url(request, reverse("user_profile:supply_management_report"))
	supply = SupplyManagement.cmobjects.select_related(
		"consumer",
		"alloted_vehicle",
		"created_by",
		"created_by__profile",
		"created_by__profile__user_type",
	).filter(id=id).first()

	if not supply:
		messages.error(request, "Supply management data not found.")
		return redirect(next_url)

	return CommonMixin.render(request, "reports/supply_management_detail.html", {
		"supply": supply,
		"next_url": next_url,
	})


@login_required
def challan_process_list(request):
	if not _can_access_challan_process(request):
		messages.error(request, "You have no permission to access the requested resource!")
		return redirect(reverse("index"))
	if not request.user.is_superuser:
		return redirect(reverse("user_profile:challan_process_records"))

	search_query = request.GET.get("q", "").strip()
	queryset = _challan_process_blf_list()
	if search_query:
		queryset = queryset.filter(
			Q(entity_unit__icontains=search_query) |
			Q(entity_name__icontains=search_query) |
			Q(email__icontains=search_query) |
			Q(user__username__icontains=search_query)
		)
	paginator = Paginator(queryset, 10)
	page_obj = paginator.get_page(request.GET.get("page", 1))
	context = {
		"page_obj": page_obj,
		"paginator": paginator,
		"is_paginated": page_obj.has_other_pages(),
		"blf_results": page_obj.object_list,
		"search_query": search_query,
		"list_url": reverse("user_profile:challan_process_list"),
		"total_count": paginator.count,
	}
	if _is_ajax_request(request):
		return CommonMixin.render(request, "challan_process/_blf_results.html", context)
	return CommonMixin.render(request, "challan_process/list.html", context)


@login_required
def challan_process_records(request):
	if not _can_access_challan_process(request):
		messages.error(request, "You have no permission to access the requested resource!")
		return redirect(reverse("index"))
	selected_blf = _selected_challan_process_blf(request)
	owner_user = selected_blf.user if selected_blf and selected_blf.user_id else None
	if request.user.is_superuser and not owner_user:
		messages.error(request, "Please select BLF user first.")
		return redirect(reverse("user_profile:challan_process_list"))
	queryset = _challan_process_queryset(request, owner_user=owner_user)
	search_query = request.GET.get("q", "").strip()
	date_from = request.GET.get("date_from", "").strip()
	date_to = request.GET.get("date_to", "").strip()
	if search_query:
		queryset = queryset.filter(
			Q(weighment_txn_id__icontains=search_query) |
			Q(supplier__username__icontains=search_query) |
			Q(supplier__profile__full_name__icontains=search_query) |
			Q(supply_challan__supply_challan_id__icontains=search_query) |
			Q(vehicle_no__vehicle_number__icontains=search_query)
		)
	if date_from:
		queryset = queryset.filter(supply_date__gte=date_from)
	if date_to:
		queryset = queryset.filter(supply_date__lte=date_to)
	paginator = Paginator(queryset, 10)
	page_obj = paginator.get_page(request.GET.get("page", 1))
	context = {
		"page_obj": page_obj,
		"paginator": paginator,
		"is_paginated": page_obj.has_other_pages(),
		"process_list": page_obj.object_list,
		"search_query": search_query,
		"date_from": date_from,
		"date_to": date_to,
		"list_url": reverse("user_profile:challan_process_records"),
		"current_list_url": _challan_process_records_url(request),
		"total_count": paginator.count,
		"selected_blf": selected_blf,
		"selected_blf_id": str(selected_blf.id) if selected_blf else "",
		"blf_query": _challan_process_context_query(request),
	}
	if _is_ajax_request(request):
		return CommonMixin.render(request, "challan_process/_table.html", context)
	return CommonMixin.render(request, "challan_process/records.html", context)


@login_required
def challan_process_detail(request, pk):
	if not _can_access_challan_process(request):
		messages.error(request, "You have no permission to access the requested resource!")
		return redirect(reverse("index"))
	selected_blf = _selected_challan_process_blf(request)
	owner_user = selected_blf.user if selected_blf and selected_blf.user_id else None
	weighment = _challan_process_queryset(request, owner_user=owner_user).filter(pk=pk).first()
	if not weighment:
		messages.error(request, "Challan process record not found.")
		return redirect(reverse("user_profile:challan_process_records") + _challan_process_context_query(request))
	leaf_receipt = LeafManagement.cmobjects.filter(weighment_txn=weighment).order_by("-id").first()
	supplier_exit = SupplierExit.cmobjects.filter(weighment_txn=weighment).order_by("-id").first()
	leaf_collection = LeafCollection.objects.filter(weighment_supply_id=weighment).order_by("-id").first()
	return CommonMixin.render(request, "challan_process/detail.html", {
		"weighment": weighment,
		"leaf_receipt": leaf_receipt,
		"supplier_exit": supplier_exit,
		"leaf_collection": leaf_collection,
		"blf_query": _challan_process_context_query(request),
		"back_url": get_safe_next_url(request, reverse("user_profile:challan_process_records") + _challan_process_context_query(request)),
	})


@login_required
def challan_process_create(request):
	if not _can_access_challan_process(request):
		messages.error(request, "You have no permission to access the requested resource!")
		return redirect(reverse("index"))
	selected_blf = _selected_challan_process_blf(request)
	owner_user = selected_blf.user if selected_blf and selected_blf.user_id else None
	if request.user.is_superuser and not owner_user:
		messages.error(request, "Please select BLF user first.")
		return redirect(reverse("user_profile:challan_process_list"))
	if request.method == "POST":
		form = ChallanProcessForm(request.POST, request=request, owner_blf=selected_blf, owner_user=owner_user)
		if form.is_valid():
			from user_profile.signals import create_challan_process_records
			try:
				result = create_challan_process_records(form.api_payload(), owner_user or request.user)
			except Exception as error:
				messages.error(request, str(error))
			else:
				messages.success(request, "Challan processed successfully.")
				return redirect(reverse("user_profile:challan_process_detail", kwargs={"pk": result["weighment"].pk}) + _challan_process_context_query(request))
		else:
			messages.error(request, "Please correct the errors below.")
	else:
		form = ChallanProcessForm(request=request, owner_blf=selected_blf, owner_user=owner_user)
	return CommonMixin.render(request, "challan_process/form.html", {
		"form": form,
		"form_title": "Create Process Challan",
		"submit_label": "Finalize Challan",
		"grower_rows": [],
		"selected_blf_id": str(selected_blf.id) if selected_blf else "",
		"back_url": reverse("user_profile:challan_process_records") + _challan_process_context_query(request),
		"supplier_choices_json": [
			{"id": str(user_id), "type": supplier_type} for user_id, _label, supplier_type in form.supplier_choices
		],
	})


@login_required
def challan_process_edit(request, pk):
	if not _can_access_challan_process(request):
		messages.error(request, "You have no permission to access the requested resource!")
		return redirect(reverse("index"))
	selected_blf = _selected_challan_process_blf(request)
	owner_user = selected_blf.user if selected_blf and selected_blf.user_id else None
	weighment = _challan_process_queryset(request, owner_user=owner_user).filter(pk=pk).first()
	if not weighment:
		messages.error(request, "Challan process record not found.")
		return redirect(reverse("user_profile:challan_process_records") + _challan_process_context_query(request))
	if request.method == "POST":
		form = ChallanProcessForm(request.POST, request=request, owner_blf=selected_blf, owner_user=owner_user)
		if form.is_valid():
			from user_profile.signals import update_challan_process_records
			try:
				update_challan_process_records(pk, form.api_payload(), owner_user or request.user)
			except Exception as error:
				messages.error(request, str(error))
			else:
				messages.success(request, "Challan process updated successfully.")
				return redirect(reverse("user_profile:challan_process_detail", kwargs={"pk": pk}) + _challan_process_context_query(request))
		else:
			messages.error(request, "Please correct the errors below.")
	else:
		form = ChallanProcessForm(initial=_challan_process_form_initial(weighment), request=request, owner_blf=selected_blf, owner_user=owner_user)
	return CommonMixin.render(request, "challan_process/form.html", {
		"form": form,
		"form_title": "Edit Process Challan",
		"submit_label": "Update Challan",
		"weighment": weighment,
		"grower_rows": _challan_process_grower_rows(weighment),
		"selected_blf_id": str(selected_blf.id) if selected_blf else "",
		"back_url": reverse("user_profile:challan_process_records") + _challan_process_context_query(request),
		"supplier_choices_json": [
			{"id": str(user_id), "type": supplier_type} for user_id, _label, supplier_type in form.supplier_choices
		],
	})


@login_required
def challan_process_delete(request, pk):
	if not _can_access_challan_process(request):
		messages.error(request, "You have no permission to access the requested resource!")
		return redirect(reverse("index"))
	next_url = get_safe_next_url(request, reverse("user_profile:challan_process_records") + _challan_process_context_query(request))
	if request.method != "POST":
		return redirect(reverse("user_profile:challan_process_detail", kwargs={"pk": pk}))
	from user_profile.signals import delete_challan_process_records
	try:
		delete_challan_process_records(pk, request.user)
	except Exception as error:
		messages.error(request, str(error))
		return redirect(reverse("user_profile:challan_process_detail", kwargs={"pk": pk}))
	messages.success(request, "Challan process deleted successfully.")
	return redirect(next_url)


def _is_ajax_request(request):
	return request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest"


def _display_value(value):
	if value is None or value == "":
		return "NA"
	return str(value)


def _join_values(values):
	values = [str(value) for value in values if value]
	return ", ".join(values) if values else "NA"


def _garden_cell(item):
	names = [garden.name for garden in item.gardens_grower.all()]
	if any(names):
		return _join_values(names)
	if item.user_id:
		url = reverse("user_profile:grower_profile_edit", kwargs={"user_create_pk": item.user_id})
		return f'<a href="{url}" class="btn bg-warning btn-xs text-white pl-3 pr-3"><i class="fa fa-edit mr-1"></i>Edit Profile</a>'
	return "NA"


def _related_names(manager, attr="name"):
	return _join_values(getattr(obj, attr, None) for obj in manager.all())


def _related_usernames(manager):
	return _join_values(obj.user.username for obj in manager.all() if getattr(obj, "user", None))


def _profile_list_config():
	return {
		"grower": {
			"model": GrowerProfile,
			"edit_url": "user_profile:grower_profile_edit",
			"select_related": ["user", "region", "profile_type"],
			"prefetch_related": ["associated_entity__user", "gardens_grower"],
			"page_size": 10,
			"filters": [
				{"name": "user_id", "label": "User Id/ Username", "type": "text", "lookup": "user__username__icontains"},
				{"name": "name", "label": "Name", "type": "text", "lookup": "name__icontains"},
				{"name": "associated_entity", "label": "Asso.. Entity", "type": "select", "lookup": "associated_entity__user__username__icontains"},
				{"name": "mobile_number", "label": "Contact no", "type": "tel", "lookup": "mobile_number__icontains"},
			],
			"columns": ["Sr. No", "Username/ID", "Name", "Grower Type", "Associated Entity", "Gardens", "Visiting Card", "User Status", "Action"],
			"row": lambda item, index: [
				index,
				_display_value(item.user.username if item.user else None),
				_display_value(item.name),
				_display_value(item.grower_type),
				_related_usernames(item.associated_entity),
				_garden_cell(item),
			],
			"visiting_card": True,
		},
		"aggregator": {
			"model": AggregatorProfile,
			"edit_url": "user_profile:aggregator_profile_edit",
			"select_related": ["user", "region", "profile_type"],
			"prefetch_related": ["associated_entity__user"],
			"page_size": 10,
			"filters": [
				{"name": "user_id", "label": "User Id/ Username", "type": "text", "lookup": "user__username__icontains"},
				{"name": "name", "label": "Name", "type": "text", "lookup": "name__icontains"},
				{"name": "mobile_number", "label": "Contact no", "type": "tel", "lookup": "mobile_number__icontains"},
			],
			"columns": ["Sr. No", "Username/ID", "Name", "Aggregator Type", "Associated Entity", "Contact No", "Visiting Card", "User Status", "Action"],
			"row": lambda item, index: [
				index,
				_display_value(item.user.username if item.user else None),
				_display_value(item.name),
				_display_value(item.aggregator_type),
				_related_usernames(item.associated_entity),
				_display_value(item.mobile_number),
			],
			"visiting_card": True,
		},
		"blf": {
			"model": BlfProfile,
			"edit_url": "user_profile:blf_profile_edit",
			"select_related": ["user", "region", "profile_type"],
			"prefetch_related": [],
			"page_size": 50,
			"filters": [
				{"name": "user_id", "label": "User Id/ Username", "type": "text", "lookup": "user__username__icontains"},
				{"name": "entity_unit", "label": "Entity Unit", "type": "text", "lookup": "entity_unit__icontains"},
				{"name": "is_tcms_user", "label": "TCMS Verified", "type": "select", "lookup": "is_tcms_user", "choices": [("True", "YES"), ("False", "NO")]},
			],
			"columns": ["Sr. No", "Username/ID", "Entity Unit", "Region", "TCMS Verified", "User Status", "Action"],
			"row": lambda item, index: [
				index,
				_display_value(item.user.username if item.user else None),
				_display_value(item.entity_unit),
				_display_value(item.region),
				"YES" if item.is_tcms_user else "NO",
			],
		},
		"estate": {
			"model": EstateProfile,
			"edit_url": "user_profile:estate_profile_edit",
			"select_related": ["user", "region", "profile_type"],
			"prefetch_related": ["estate_gardens"],
			"page_size": 10,
			"filters": [
				{"name": "user_id", "label": "User Id/ Username", "type": "text", "lookup": "user__username__icontains"},
				{"name": "entity_unit", "label": "Garden/Entity Unit", "type": "text", "lookup": "entity_unit__icontains"},
			],
			"columns": ["Sr. No", "Username/ID", "Garden/Entity Unit", "Region", "Gardens", "Trough Details", "User Status", "Action"],
			"row": lambda item, index: [
				index,
				_display_value(item.user.username if item.user else None),
				_display_value(item.entity_unit),
				_display_value(item.region),
				_related_names(item.estate_gardens),
			],
			"trough_details": True,
		},
		"advisory": {
			"model": AdvisoryProfile,
			"edit_url": "user_profile:advisory_profile_edit",
			"select_related": ["user", "region", "profile_type"],
			"prefetch_related": [],
			"page_size": 10,
			"filters": [
				{"name": "user_id", "label": "User Id/ Username", "type": "text", "lookup": "user__username__icontains"},
				{"name": "organization_name", "label": "Organization Name", "type": "text", "lookup": "organization_name__icontains"},
			],
			"columns": ["Sr. No", "Username/ID", "Organization", "Expert Name", "Region", "User Status", "Action"],
			"row": lambda item, index: [
				index,
				_display_value(item.user.username if item.user else None),
				_display_value(item.organization_name),
				_display_value(item.expert_name),
				_display_value(item.region),
			],
		},
		"consignee": {
			"model": ConsigneeProfile,
			"edit_url": "user_profile:consignee_profile_edit",
			"select_related": ["user", "region", "profile_type"],
			"prefetch_related": [],
			"page_size": 10,
			"filters": [
				{"name": "user_id", "label": "User Id/ Username", "type": "text", "lookup": "user__username__icontains"},
				{"name": "mobile_number", "label": "Contact no", "type": "tel", "lookup": "mobile_number__icontains"},
			],
			"columns": ["Sr. No", "Username/ID", "Organization", "Buyer Name", "Contact No", "Region", "User Status", "Action"],
			"row": lambda item, index: [
				index,
				_display_value(item.user.username if item.user else None),
				_display_value(item.organization_name),
				_display_value(item.buyer_name),
				_display_value(item.mobile_number),
				_display_value(item.region),
			],
		},
		"shg": {
			"model": ShgCooperativeProfile,
			"edit_url": "user_profile:shg_profile_edit",
			"select_related": ["user", "region", "shg_cooperative_type", "profile_type"],
			"prefetch_related": [],
			"page_size": 10,
			"filters": [
				{"name": "user_id", "label": "User Id/ Username", "type": "text", "lookup": "user__username__icontains"},
				{"name": "name", "label": "Name", "type": "text", "lookup": "name__icontains"},
				{"name": "mobile_number", "label": "Contact no", "type": "tel", "lookup": "mobile_number__icontains"},
			],
			"columns": ["Sr. No", "Username/ID", "Name", "SHG/Cooperative Type", "Contact No", "Region", "User Status", "Action"],
			"row": lambda item, index: [
				index,
				_display_value(item.user.username if item.user else None),
				_display_value(item.name),
				_display_value(item.shg_cooperative_type),
				_display_value(item.mobile_number),
				_display_value(item.region),
			],
		},
	}


def _profile_model_for_type(user_type):
	config = _profile_list_config().get(str(user_type).lower())
	return config["model"] if config else None


def _append_next(url, next_url):
	if not next_url:
		return url
	separator = "&" if "?" in url else "?"
	return f"{url}{separator}next={quote(next_url)}"


def _profile_list_default(user_type):
	return reverse("user_profile:all_users", kwargs={"usertype_slug": user_type})


def _apply_profile_filters(queryset, filters, request):
	query = request.GET.get("q", "").strip()
	needs_distinct = False
	if query:
		search = Q(user__username__icontains=query)
		for field in filters:
			lookup = field.get("lookup")
			if not lookup or field["type"] == "select" and lookup == "is_tcms_user":
				continue
			if lookup.startswith("associated_entity__") or lookup.startswith("associated_"):
				needs_distinct = True
			search |= Q(**{lookup: query})
		queryset = queryset.filter(search)

	for field in filters:
		value = request.GET.get(field["name"])
		if value in (None, "", field["label"]):
			continue
		if field["type"] == "select" and field.get("lookup") == "is_tcms_user":
			queryset = queryset.filter(is_tcms_user=(value == "True"))
			continue
		if field["lookup"].startswith("associated_entity__") or field["lookup"].startswith("associated_"):
			needs_distinct = True
		queryset = queryset.filter(**{field["lookup"]: value.strip()})
	return queryset.distinct() if needs_distinct else queryset


def _build_profile_rows(page_obj, config, user_type, start_index, current_list_url=""):
	rows = []
	for offset, item in enumerate(page_obj.object_list, start=start_index):
		user = item.user
		user_id = user.id if user else None
		is_active = bool(user and user.is_active)
		edit_url = reverse(config["edit_url"], kwargs={"user_create_pk": user_id}) if user_id and is_active else None
		view_url = reverse("user_profile:user_profile_view", kwargs={"user_create_pk": user_id, "user_type": user_type}) if user_id else None
		delete_url = reverse("user_profile:user_profile_delete", kwargs={"user_create_pk": user_id, "user_type": user_type}) if user_id else None
		row = {
			"cells": config["row"](item, offset),
			"is_active": is_active,
			"view_url": _append_next(view_url, current_list_url) if view_url else None,
			"edit_url": _append_next(edit_url, current_list_url) if edit_url else None,
			"delete_url": _append_next(delete_url, current_list_url) if delete_url else None,
			"password_url": _append_next(reverse("accounts:users_change_password", kwargs={"user_id": user_id}), current_list_url) if user_id else None,
			"activate_url": _append_next(reverse("user_profile:user_profile_active", kwargs={"user_create_pk": user_id, "user_type": user_type}), current_list_url) if user_id and not is_active else None,
			"deactivate_url": _append_next(reverse("user_profile:user_profile_deactive", kwargs={"user_create_pk": user_id, "user_type": user_type}), current_list_url) if user_id and is_active else None,
			"visiting_card_url": None,
			"trough_details_url": None,
		}
		if config.get("visiting_card") and user_id:
			row["visiting_card_url"] = reverse("user_profile:visting_card_download", kwargs={"user_pk": user_id, "user_type": user_type})
		if config.get("trough_details") and user_id:
			row["trough_details_url"] = reverse("user_profile:estate_trough_details_list", kwargs={"user_create_pk": user_id})
		rows.append(row)
	return rows


@permission_required_admin
def users_list(request, usertype_slug):
	configs = _profile_list_config()
	config = configs.get(usertype_slug)
	if not config:
		messages.error(request, "Invalid user type.")
		return redirect(reverse("index"))

	queryset = config["model"].cmobjects.select_related(*config["select_related"]).prefetch_related(*config["prefetch_related"]).order_by("-id")
	queryset = _apply_profile_filters(queryset, config["filters"], request)

	paginator = Paginator(queryset, config["page_size"])
	page_number = request.GET.get("page", 1)
	try:
		profile_list = paginator.page(page_number)
	except PageNotAnInteger:
		profile_list = paginator.page(1)
	except EmptyPage:
		profile_list = paginator.page(paginator.num_pages)

	query_params = request.GET.copy()
	query_params.pop("page", None)
	context = {
		"profile_list": profile_list,
		"rows": _build_profile_rows(profile_list, config, usertype_slug, profile_list.start_index(), request.get_full_path()),
		"columns": config["columns"],
		"filters": config["filters"],
		"user_type": usertype_slug,
		"profile_type": usertype_slug,
		"tot_count": paginator.count,
		"querystring": query_params.urlencode(),
		"search_query": request.GET.get("q", "").strip(),
		"list_url": reverse("user_profile:all_users", kwargs={"usertype_slug": usertype_slug}),
		"current_list_url": request.get_full_path(),
		"has_filters": any(request.GET.get(field["name"]) for field in config["filters"]),
		"current_url": True,
		"url": request.path,
	}
	if _is_ajax_request(request):
		return CommonMixin.render(request, "profile/_user_table.html", context)
	return CommonMixin.render(request, "profile/user_list.html", context)

def aggregator_list(request, usertype_slug):

	profile_details = BlfProfile.cmobjects.filter(user_id=request.user.id).first()
	agg_list = AggregatorProfile.cmobjects.filter(created_by_id=request.user.id).order_by('-id')
	print(profile_details)
	user_type = "aggregator"
	context={
		'user_type' : usertype_slug,
		'agg_list' : agg_list, 
		'profile_type' : usertype_slug,

	}
	return CommonMixin.render(request, 'profile/aggregator_list.html', context)


@login_required
def search_profile(request):
	profile_type = request.GET.get('user_type')
	if not profile_type:
		messages.error(request, "User type is required.")
		return redirect(reverse("index"))
	return users_list(request, profile_type)


@login_required
@permission_required_admin
def profile_detail(request, user_create_pk, user_type):
	model = _profile_model_for_type(user_type)
	if not model:
		messages.error(request, "Invalid user type.")
		return redirect(reverse("index"))

	queryset = model.cmobjects.select_related("user", "region", "profile_type")
	if user_type == "grower":
		queryset = queryset.select_related("state", "district", "associated_unit").prefetch_related(
			"associated_entity",
			"associated_aggregator",
			"grower_proof_id_attachments",
		)
	profile_details = queryset.filter(user_id=user_create_pk).first()
	if not profile_details:
		messages.error(request, "Profile not found.")
		return redirect(reverse("user_profile:all_users", kwargs={"usertype_slug": user_type}))

	proof_attachments = []
	if user_type == "grower":
		proof_attachments = UserProfileAttachments.cmobjects.filter(
			grower=profile_details,
			doc_type="id_proof",
		).order_by("id")

	context = {
		"profile_details": profile_details,
		"proof_attachments": proof_attachments,
		"user_type": user_type,
		"next_url": get_safe_next_url(request, reverse("user_profile:all_users", kwargs={"usertype_slug": user_type})),
	}
	return CommonMixin.render(request, "profile/user_profile_detail.html", context)



class EstateUserList(LoginRequiredMixin, CommonMixin, ListView):
    """
    Tea Grade List View
    """
    model = EstateProfile
    context_object_name = 'profile_list'
    template_name = 'estate_list.html'
    paginate_by = 5

    def get(self, request, *args, **kwargs):
    # checking if the user is Admin
        try:
            if not self.request.user.is_superuser:
                messages.error(self.request, 'You have no permission to access the requested resource!')
                return redirect(reverse('index'))
        except AttributeError as error:
            messages.error(self.request, 'You have no permission to access the requested resource!')
            return redirect(reverse('index'))

        return super().get(request, *args, **kwargs)

    def get_queryset(self, *args, **kwargs):

        qs = super().get_queryset(*args, **kwargs)
        tea_type=self.request.GET.get('tea_type')
        grade=self.request.GET.get('grade')
        if tea_type:
            return qs.filter(tea_type__name__icontains=tea_type)
        if grade:
            return qs.filter(grade__icontains=grade)
       
        return TeaGradeDetails.objects.all().order_by('-id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        context['tea_type_list'] = TeaType.objects.all().order_by('id')
        return context







@login_required
def delete_profile(request, user_create_pk, user_type):

	try:
		if not request.user.is_superuser:
			messages.error(request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))
	except AttributeError as error:
		messages.error(request, 'You have no permission to access the requested resource!')
		return redirect(reverse('index'))

	profile_model = _profile_model_for_type(user_type)
	if not profile_model:
		messages.error(request, "Invalid user type.")
		return redirect(reverse("index"))

	
	# profile_details.delete()

	with transaction.atomic():
		profile_details = profile_model.cmobjects.filter(user_id=user_create_pk).first()
		if not profile_details:
			messages.error(request, "Profile not found.")
			return redirect(get_safe_next_url(request, reverse('user_profile:all_users', kwargs={"usertype_slug": user_type })))
		profile_details.is_deleted = True
		try:
			profile_details.save()
		except APIException as exception:
			messages.error(request, get_api_exception_message(exception))
			return redirect(get_safe_next_url(request, reverse('user_profile:all_users', kwargs={"usertype_slug": user_type })))
		# profile_details.delete()

		User.objects.filter(id=profile_details.user_id).update(is_active=False) 
		messages.success(
			request, 'User Profile Deleted Successfully')

	return redirect(get_safe_next_url(request, reverse('user_profile:all_users', kwargs={"usertype_slug": user_type })))



@login_required
def user_profile_active(request, user_create_pk, user_type):

	try:
		if not request.user.is_superuser:
			messages.error(request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))
	except AttributeError as error:
		messages.error(request, 'You have no permission to access the requested resource!')
		return redirect(reverse('index'))

	if  user_type == "grower":
		profile_model = GrowerProfile
	elif user_type == "aggregator":
		profile_model = AggregatorProfile
	elif user_type == "blf":
		profile_model = BlfProfile
	elif user_type == "advisory":
		profile_model = AdvisoryProfile
	elif user_type == "estate":
		profile_model = EstateProfile
	elif user_type == "consignee":
		profile_model = ConsigneeProfile
	elif user_type == "shg":
		profile_model = ShgCooperativeProfile

	# profile_model.objects.filter(user_id=user_create_pk).update(is_active=True)
	User.objects.filter(id=user_create_pk).update(is_active=True)
	messages.success(request, 'User Profile is Successfully Activated!')
	# return HttpResponseRedirect(request.META.get("HTTP_REFERER"))
	
	return redirect(get_safe_next_url(request, reverse('user_profile:all_users', kwargs={"usertype_slug": user_type })))


@login_required
def user_profile_deactive(request, user_create_pk, user_type):

	try:
		if not request.user.is_superuser:
			messages.error(request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))
	except AttributeError as error:
		messages.error(request, 'You have no permission to access the requested resource!')
		return redirect(reverse('index'))

	if  user_type == "grower":
		profile_model = GrowerProfile
	elif user_type == "aggregator":
		profile_model = AggregatorProfile
	elif user_type == "blf":
		profile_model = BlfProfile
	elif user_type == "advisory":
		profile_model = AdvisoryProfile
	elif user_type == "estate":
		profile_model = EstateProfile
	elif user_type == "consignee":
		profile_model = ConsigneeProfile
	elif user_type == "shg":
		profile_model = ShgCooperativeProfile

	print(" dwedefe ")

	# profile_model.objects.filter(user_id=user_create_pk).update(is_active=False)
	User.objects.filter(id=user_create_pk).update(is_active=False)
	messages.success(request, 'User Profile is Successfully Deactivated!')
	# return HttpResponseRedirect(request.META.get("HTTP_REFERER"))

	return redirect(get_safe_next_url(request, reverse('user_profile:all_users', kwargs={"usertype_slug": user_type })))




class ProfileListView(LoginRequiredMixin, ListView, CommonMixin):
	"""
	Profile List View
	"""
	model = GrowerProfile
	context_object_name = "user_list"
	template_name = "profile/user_list.html"
	paginate_by = 10

	def get_queryset(self):
		return self.model.cmobjects.filter(user__is_active=True).order_by('user__username')

	def get_context_data(self, **kwargs):   
		context = super(ProfileListView, self).get_context_data(**kwargs)

		# context['admin_list'] = Profile.objects.filter(user_type = 'ADMIN', user__is_active=True).order_by('user__username')

		# context['employee_list'] = Profile.objects.filter(user_type = 'EMPLOYEE', user__is_active=True).order_by('user__username')

		# context['manager_list'] = Profile.objects.filter(user_type = 'REPORTING MANAGER', user__is_active=True).order_by('user__username')

        # context['']

		return context


@login_required
def user_create(request, usertype_slug):
	# User create view for all Users Type 
	profile_type = ProfileType.objects.filter(
		slug=usertype_slug).first()
	next_url = get_safe_next_url(request, reverse('user_profile:all_users', kwargs={"usertype_slug": usertype_slug}))
	
	if request.method == 'POST':
		form = CreateUserForm(request.POST)
		if form.is_valid():
			# import pdb
			# pdb.set_trace()
			username = form.cleaned_data.get('username')
			email = form.cleaned_data.get('email')
			password = form.cleaned_data.get('password1')
			password2 = form.cleaned_data.get('password2')
			request.session['emp_pass'] = str(password)

			create = form.save()
			user_id = create.pk

			# print(profile_type)
			
			if str(profile_type) == "aggregator":
				AggregatorProfile.cmobjects.update_or_create(user=create, defaults={
						'email': email, 'profile_type': profile_type, 'created_by' : request.user,})
				Profile.objects.update_or_create(user=create, defaults={
						'user_type': profile_type})
				print("aggregator check")
				messages.success(request, 'User Created Successfully')
				return redirect(_append_next(reverse('user_profile:aggregator_profile_create', kwargs={"user_create_pk": user_id }), next_url))
	
			elif str(profile_type) == "grower":
				GrowerProfile.cmobjects.update_or_create(user=create, defaults={
						'email': email, 'profile_type': profile_type, 'created_by' : request.user})
				Profile.objects.update_or_create(user=create, defaults={
						'user_type': profile_type})
				gr_id = GrowerProfile.objects.filter(user_id=user_id).first()
				Gardens.objects.update_or_create(grower=gr_id, defaults={'is_division': False})
				messages.success(request, 'User Created Successfully')
				return redirect(_append_next(reverse('user_profile:grower_profile_create', kwargs={"user_create_pk": user_id }), next_url))	
			

			elif str(profile_type) == "blf":
				BlfProfile.cmobjects.update_or_create(user=create, defaults={
						'email': email, 'profile_type': profile_type, 'created_by' : request.user})
				Profile.objects.update_or_create(user=create, defaults={
						'user_type': profile_type})
				messages.success(request, 'User Created Successfully')
				return redirect(_append_next(reverse('user_profile:blf_profile_create', kwargs={"user_create_pk": user_id }), next_url))
			
			elif str(profile_type) == "advisory":
				AdvisoryProfile.cmobjects.update_or_create(user=create, defaults={
						'email': email, 'profile_type': profile_type, 'created_by' : request.user})
				Profile.objects.update_or_create(user=create, defaults={
						'user_type': profile_type})
				messages.success(request, 'User Created Successfully')
				return redirect(_append_next(reverse('user_profile:advisory_profile_create', kwargs={"user_create_pk": user_id }), next_url))
			

			elif str(profile_type) == "estate":
				EstateProfile.cmobjects.update_or_create(user=create, defaults={
						'email': email, 'profile_type': profile_type, 'created_by' : request.user})
				Profile.objects.update_or_create(user=create, defaults={
						'user_type': profile_type})
				est_id = EstateProfile.objects.filter(user_id=user_id).first()
				# print(est_id)
				Gardens.objects.update_or_create(user=create, defaults={'is_division': False, 'estate' : est_id })
				messages.success(request, 'User Created Successfully')
				return redirect(_append_next(reverse('user_profile:estate_profile_create', kwargs={"user_create_pk": user_id }), next_url))
			

			elif str(profile_type) == "consignee":
				ConsigneeProfile.cmobjects.update_or_create(user=create, defaults={
						'email': email, 'profile_type': profile_type, 'created_by' : request.user})
				Profile.objects.update_or_create(user=create, defaults={
						'user_type': profile_type})
				messages.success(request, 'User Created Successfully')
				return redirect(_append_next(reverse('user_profile:consignee_profile_create', kwargs={"user_create_pk": user_id }), next_url))
			elif str(profile_type) == "SHG":
				ShgCooperativeProfile.cmobjects.update_or_create(user=create, defaults={
						'email': email, 'profile_type': profile_type, 'created_by' : request.user})
				Profile.objects.update_or_create(user=create, defaults={
						'user_type': profile_type})
				messages.success(request, 'User Created Successfully')
				return redirect(_append_next(reverse('user_profile:shg_profile_create', kwargs={"user_create_pk": user_id }), next_url))
		else:
			for field, errors in form.errors.items():
				field_label = form.fields[field].label if field in form.fields else None
				for error in errors:
					if field_label:
						messages.error(request, f"{field_label}: {error}")
					else:
						messages.error(request, error)
	else:
		form = CreateUserForm()
	context={
		'profile_type' : profile_type,
		'form' : form,
		'current_url': True,
		'url': reverse('user_profile:all_users', kwargs={"usertype_slug": usertype_slug}),
		'next_url': next_url,
	}
	return CommonMixin.render(request, 'profile/user_create.html', context)


@login_required
def validate_user_create_fields(request):
	username = request.GET.get('username', '').strip()
	email = request.GET.get('email', '').strip()
	errors = {}

	if username and User.objects.filter(username__iexact=username).exists():
		errors['username'] = 'This username/user ID already exists.'

	if email and User.objects.filter(email__iexact=email).exists():
		errors['email'] = 'This email address is already registered.'

	return JsonResponse({
		'is_valid': not bool(errors),
		'errors': errors,
	})


class GrowerProfileCreateView(LoginRequiredMixin, CommonMixin, CreateView):

	model = GrowerProfile
	form_class = GrowerProfileForm
	second_form_class = CreateUserForm
	template_name = 'profile/grower_profile_create.html'

	def get(self, request, *args, **kwargs):
	# checking if the user is Admin
		try:
			if not self.request.user.is_superuser:
				messages.error(self.request, 'You have no permission to access the requested resource!')
				return redirect(reverse('index'))
		except AttributeError as error:
			messages.error(self.request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)
	

	def get_success_url(self, **kwargs):
		messages.success(self.request, 'Grower Profile Updated Successfully')
		# return reverse('user_profile:grower_profile_create', kwargs={"user_create_pk": self.kwargs['user_create_pk'] }  )
		return reverse('user_profile:all_users',kwargs={"usertype_slug": 'grower' })

	def get_object(self):
		profile_type = ProfileType.objects.filter(slug=self.kwargs['usertype_slug']).first()
		return profile_type

	def get_context_data(self, **kwargs):
		kwargs['active_client'] = True
		if 'form' not in kwargs:
			kwargs['form'] = self.form_class(instance=self.get_object())
		if 'form2' not in kwargs:
			kwargs['form2'] = self.second_form_class(instance=self.get_object())


		

		context = super(GrowerProfileCreateView, self).get_context_data(**kwargs)
		context['profile_type'] = self.get_object()
		return context
	
	def form_valid(self, form):
		self.object = form.save(commit=False)
		self.object.is_active = False
		self.object.save()

		return super().form_valid(form)
	
	# def post(self, request, *args, **kwargs):
	# 	user_form = CreateUserForm(data = request.POST)
	# 	profile_form = GrowerProfileForm(data = request.POST)
	# 	if user_form.is_valid() and profile_form.is_valid():
	# 		user = user_form.save()
	# 		user.set_password(user.password)
	# 		user.save()
	# 		profile = profile_form.save(commit=False)
	# 		profile.user = user
	# 		profile.save()
	# 		registered = True
		
	# 	else:
	# 		user_form = CreateUserForm()
	# 		profile_form = GrowerProfileForm()

	# 	return reverse('user_profile:all_users',kwargs={"usertype_slug": 'grower' })	
	
	# def form_valid(self, form):
		
	# 	print("###########  Print here  $$$$")
	# 	# print(self.form2.instance.id)

	# 	# self.user_id = self.object
	# 	print('form_valid')
	# 	context = self.get_context_data()

	# 	with transaction.atomic():
	# 		self.object = form.save()

		# return super(GrowerProfileCreateView, self).form_valid(form)

	
#######################################################################################
############  Grower Profile Section ################


class GrowerCreateView(LoginRequiredMixin, CommonMixin, UpdateView):
	"""
	Grower Profile Create View 
	"""
	model = GrowerProfile
	form_class = GrowerProfileForm
	template_name = 'profile/grower_profile_create.html'
	import datetime
	print(datetime.date.today())
	
	def get(self, request, *args, **kwargs):
	# checking if the user is Admin
		try:
			if not self.request.user.is_superuser:
				messages.error(self.request, 'You have no permission to access the requested resource!')
				return redirect(reverse('index'))
		except AttributeError as error:
			messages.error(self.request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)


	def get_success_url(self, **kwargs):
		messages.success(self.request, 'Grower Profile Updated Successfully')
		# return reverse('user_profile:grower_profile_create', kwargs={"user_create_pk": self.kwargs['user_create_pk'] }  )
		profile_details = GrowerProfile.objects.filter(user_id=self.kwargs['user_create_pk']).first()

		gardens_details = Gardens.objects.filter(grower_id=profile_details.pk).first()

		print("grower ID")
		print(profile_details.pk)

		print("garden ID")
		print(gardens_details.pk)

		# garden_id = gardens_details.pk

		if str(profile_details.grower_type) == "LTG":
			return reverse('gardens_managment:garden_update', kwargs={"garden_pk": gardens_details.pk, "grower_pk": profile_details.pk })
			# return reverse('index')

		if str(profile_details.grower_type) == "STG":
			return reverse('gardens_managment:garden_update', kwargs={"garden_pk": gardens_details.pk, "grower_pk": profile_details.pk })
		else:
			return reverse('user_profile:all_users', kwargs={"usertype_slug": 'grower' }) 
			# return reverse('index')


	def get_object(self):
		profile_details = GrowerProfile.objects.filter(user_id=self.kwargs['user_create_pk']).first()
		
		return profile_details

	def get_context_data(self, **kwargs):
		context = super(GrowerCreateView, self).get_context_data(**kwargs)

		context['profile_details'] = self.get_object()

		context['aggregator_list'] = AggregatorProfile.cmobjects.filter(user__is_active=True).select_related('user').order_by('-id')[:100]
		context['next_url'] = get_safe_next_url(self.request, _profile_list_default("grower"))

		return context
	
	def form_valid(self, form):
		
		self.user_id = self.kwargs['user_create_pk']

		context = self.get_context_data()

		profile_details = GrowerProfile.cmobjects.filter(user_id=self.kwargs['user_create_pk']).first()
		garden_details = Gardens.objects.filter(grower_id=profile_details.pk).first()
		Gardens.objects.filter(grower_id=profile_details.pk).update(name = form.instance.name)
		Plot.objects.filter(garden_id=garden_details.pk).update_or_create(garden_id=garden_details.pk, defaults={
			'name': form.instance.name,})
		
		# BlfAssociatedUsers.objects.update_or_create(grower_id=form.instance,\
		# 			       defaults={ 'associated_entity' : form.instance.associated_entity})
		
		GrowerQrCode.objects.update_or_create(grower_id=profile_details.pk, \
					defaults={'name': form.instance.name})
		


		# new_mymodels = []

		# for i in GrowerProfile.objects.all():
		# 	new_mymodels = BlfAssociatedUsers(
		# 		blf=i['blf'],
		# 		grower=i['grower'],
		# 		aggregator=i['aggregator'],
		# 	)
		with transaction.atomic():
			self.object = form.save()
		return super(GrowerCreateView, self).form_valid(form)


def _grower_id_proof_queryset(profile):
	if not profile:
		return UserProfileAttachments.cmobjects.none()
	return UserProfileAttachments.cmobjects.filter(
		grower=profile,
		doc_type="id_proof",
	).order_by("id")


def _save_grower_id_proof_formset(formset, profile, request):
	for form in formset.deleted_forms:
		if form.instance and form.instance.pk:
			form.instance.is_deleted = True
			form.instance.updated_by = request.user
			form.instance.save(update_fields=["is_deleted", "updated_by", "updated_at"])

	for form in formset.forms:
		if form in formset.deleted_forms or not form.has_changed():
			continue
		attachment = form.save(commit=False)
		attachment.grower = profile
		attachment.doc_type = "id_proof"
		if not attachment.created_by_id:
			attachment.created_by = request.user
		attachment.updated_by = request.user
		attachment.save()


def _add_formset_errors_to_messages(request, formset):
	for error in formset.non_form_errors():
		messages.error(request, error)
	for form in formset.forms:
		for field_name, errors in form.errors.items():
			label = form.fields[field_name].label if field_name in form.fields else "ID proof"
			for error in errors:
				messages.error(request, f"{label}: {error}")


@login_required
def update_profile(request, user_create_pk):
	profile = GrowerProfile.cmobjects.select_related('user').filter(
				user_id=user_create_pk).first()
	if not profile:
		messages.error(request, "Profile not found.")
		return redirect(get_safe_next_url(request, _profile_list_default("grower")))

	if request.method == "POST":
		form = GrowerProfileForm(request.POST, request.FILES, instance=profile)
		attachment_formset = GrowerIDProofAttachmentFormSet(
			request.POST,
			request.FILES,
			instance=profile,
			queryset=_grower_id_proof_queryset(profile),
			prefix="idproof",
		)
		if form.is_valid() and attachment_formset.is_valid():
			with transaction.atomic():
				profile = form.save()
				_save_grower_id_proof_formset(attachment_formset, profile, request)
			messages.success(request, "Profile updated successfully!")
			return redirect(get_safe_next_url(request, _profile_list_default("grower")))
		else:
			messages.error(request, "Please correct the errors below.")
			if not attachment_formset.is_valid():
				_add_formset_errors_to_messages(request, attachment_formset)
	else:
		form = GrowerProfileForm(instance=profile)
		attachment_formset = GrowerIDProofAttachmentFormSet(
			instance=profile,
			queryset=_grower_id_proof_queryset(profile),
			prefix="idproof",
		)

	context = {
		"form": form,
		"attachment_formset": attachment_formset,
		"profile_details": profile,
		"user_id" : user_create_pk,
		"current_url" : True,
		"next_url": get_safe_next_url(request, _profile_list_default("grower")),
	}
	return CommonMixin.render(request, "profile/grower_profile_update.html", context)

class GrowerProfileUpdateView(LoginRequiredMixin, CommonMixin, UpdateView):
    """
    Grower Profile Update View (Optimized)
    """
    model = GrowerProfile
    form_class = GrowerProfileForm
    template_name = 'profile/grower_profile_update.html'

    def dispatch(self, request, *args, **kwargs):
        """
        Handle permission in one place instead of overriding get()
        """
        if not getattr(request.user, "is_superuser", False):
            messages.error(request, 'You have no permission to access the requested resource!')
            return redirect(reverse('index'))
        return super().dispatch(request, *args, **kwargs)

    def get_object(self):
        """
        Fetch object with related user in single query
        """
        return GrowerProfile.cmobjects.select_related('user').filter(
            user_id=self.kwargs['user_create_pk']
        ).first()

    def get_success_url(self):
        messages.success(self.request, 'Grower Profile Updated Successfully')
        return get_safe_next_url(self.request, _profile_list_default("grower"))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        profile = self.object or self.get_object()
        context.update({
            'active_client': True,
            'form': kwargs.get('form') or self.form_class(instance=profile),
            'profile_details': profile,
            'next_url': get_safe_next_url(self.request, _profile_list_default("grower")),
        })
        return context

    def form_valid(self, form):
        profile = self.get_object()
        if not profile:
            messages.error(self.request, "Profile not found")
            return redirect(reverse('index'))
        with transaction.atomic():
            # Save main profile
            self.object = form.save()
            # Update Garden (single query)
            garden = Gardens.objects.filter(grower_id=profile.pk).first()
            if garden:
                garden.name = form.instance.name
                garden.save(update_fields=['name'])
                Plot.objects.update_or_create(
                    garden_id=garden.pk,
                    defaults={'name': form.instance.name}
                )
            # Update/Create QR Code
            GrowerQrCode.objects.update_or_create(
                grower_id=profile.pk,
                defaults={'name': form.instance.name}
            )
        return super().form_valid(form)



class AggregatorProfileUpdateView(LoginRequiredMixin, CommonMixin, UpdateView):
	"""
	Aggregator Profile Update View @vivek
	"""
	model = AggregatorProfile
	form_class = AggregatorProfileForm
	template_name = 'profile/aggregator_profile_create.html'

	def get(self, request, *args, **kwargs):
	# checking if the user is Admin
		# try:
		# 	if not self.request.user.is_superuser:
		# 		messages.error(self.request, 'You have no permission to access the requested resource!')
		# 		return redirect(reverse('index'))
		# except AttributeError as error:
		# 	messages.error(self.request, 'You have no permission to access the requested resource!')
		# 	return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)

	def get_success_url(self, **kwargs):
		messages.success(self.request, 'Aggregator Profile Updated Successfully')
		return get_safe_next_url(self.request, _profile_list_default("aggregator"))

	def get_object(self):
		profile_details = AggregatorProfile.cmobjects.filter(user_id=self.kwargs['user_create_pk']).first()
		return profile_details

	def get_context_data(self, **kwargs):
		context = super(AggregatorProfileUpdateView, self).get_context_data(**kwargs)
		context['profile_details'] = self.get_object()
		context['usertype_slug'] = "aggregator"
		context["current_url"] = True
		context["next_url"] = get_safe_next_url(self.request, _profile_list_default("aggregator"))
		return context

	def form_valid(self, form):
		
		self.user_id = self.kwargs['user_create_pk']
		context = self.get_context_data()

		profile_details = AggregatorProfile.cmobjects.filter(user_id=self.kwargs['user_create_pk']).first()
		print("Aggregator")
		print(profile_details.pk)

		AggregatorQrCode.objects.update_or_create(aggregator_id=profile_details.pk, \
					defaults={'name': form.instance.name })
		
		# BlfSupplier.objects.update_or_create(aggregator_id=profile_details.pk, \
		# 		        defaults={'supplier_type': "aggregator", 'blf' : form.instance })

		print(self.kwargs['user_create_pk'])

		with transaction.atomic():
			self.object = form.save()

		return super(AggregatorProfileUpdateView, self).form_valid(form)




class AdvisoryProfileUpdateView(LoginRequiredMixin, CommonMixin, UpdateView):
	"""
	Advisory Profile Update View @vivek
	"""
	model = AdvisoryProfile
	form_class = AdvisoryProfileForm
	template_name = 'profile/advisory_profile_create.html'

	def get(self, request, *args, **kwargs):
	# checking if the user is Admin
		try:
			if not self.request.user.is_superuser:
				messages.error(self.request, 'You have no permission to access the requested resource!')
				return redirect(reverse('index'))
		except AttributeError as error:
			messages.error(self.request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)

	def get_success_url(self, **kwargs):
		messages.success(self.request, 'Advisory Profile Updated Successfully')
		return get_safe_next_url(self.request, _profile_list_default("advisory"))

	def get_object(self):
		profile_details = AdvisoryProfile.objects.filter(user_id=self.kwargs['user_create_pk']).first()
		return profile_details

	def get_context_data(self, **kwargs):
		context = super(AdvisoryProfileUpdateView, self).get_context_data(**kwargs)
		context['profile_details'] = self.get_object()
		context["current_url"] = True
		context["next_url"] = get_safe_next_url(self.request, _profile_list_default("advisory"))
		return context
	def form_valid(self, form):
		
		self.user_id = self.kwargs['user_create_pk']
		context = self.get_context_data()


		with transaction.atomic():
			self.object = form.save()


		return super(AdvisoryProfileUpdateView, self).form_valid(form)



class ConsigneeProfileUpdateView(LoginRequiredMixin, CommonMixin, UpdateView):
	"""
	Consignee Profile Update View @vivek
	"""
	model = ConsigneeProfile
	form_class = ConsigneeProfileForm
	template_name = 'profile/consignee_profile_create.html'

	def get(self, request, *args, **kwargs):
	# checking if the user is Admin
		try:
			if not self.request.user.is_superuser:
				messages.error(self.request, 'You have no permission to access the requested resource!')
				return redirect(reverse('index'))
		except AttributeError as error:
			messages.error(self.request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)

	def get_success_url(self, **kwargs):
		messages.success(self.request, 'Consignee Profile Updated Successfully')
		return get_safe_next_url(self.request, _profile_list_default("consignee"))

	def get_object(self):
		profile_details = ConsigneeProfile.objects.filter(user_id=self.kwargs['user_create_pk']).first()
		return profile_details

	def get_context_data(self, **kwargs):
		context = super(ConsigneeProfileUpdateView, self).get_context_data(**kwargs)
		context['profile_details'] = self.get_object()
		context["current_url"] = True
		context["next_url"] = get_safe_next_url(self.request, _profile_list_default("consignee"))
		return context
	def form_valid(self, form):
		self.user_id = self.kwargs['user_create_pk']
		context = self.get_context_data()
		with transaction.atomic():
			self.object = form.save()
		return super(ConsigneeProfileUpdateView, self).form_valid(form)
	
class ShgProfileUpdateView(LoginRequiredMixin, CommonMixin, UpdateView):
	"""
	SHG Pofile Update View @vivek
	"""
	model = ShgCooperativeProfile
	form_class = ShgCooperativeProfileForm
	template_name = 'profile/shg_profile_create.html'

	def get(self, request, *args, **kwargs):
	# checking if the user is Admin
		try:
			if not self.request.user.is_superuser:
				messages.error(self.request, 'You have no permission to access the requested resource!')
				return redirect(reverse('index'))
		except AttributeError as error:
			messages.error(self.request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)

	def get_success_url(self, **kwargs):
		messages.success(self.request, 'Shg Profile Updated Successfully')
		return get_safe_next_url(self.request, _profile_list_default("shg"))

	def get_object(self):
		profile_details = ShgCooperativeProfile.objects.filter(user_id=self.kwargs['user_create_pk']).first()
		return profile_details

	def get_context_data(self, **kwargs):
		context = super(ShgProfileUpdateView, self).get_context_data(**kwargs)

		context['profile_details'] = self.get_object()
		context["next_url"] = get_safe_next_url(self.request, _profile_list_default("shg"))
		return context	
	def form_valid(self, form):
		
		self.user_id = self.kwargs['user_create_pk']
		context = self.get_context_data()


		with transaction.atomic():
			self.object = form.save()


		return super(ShgProfileUpdateView, self).form_valid(form)


########## BLF PROFILE UPDATE VIEW ###############


class BlfProfileUpdateView(LoginRequiredMixin, CommonMixin, UpdateView):
	"""
	BLF PROFILE Update View
	"""
	model = BlfProfile
	form_class = BlfProfileForm
	template_name = 'profile/blf_profile_create.html'

	def get(self, request, *args, **kwargs):
		# checking if the user is Admin
		# try:
		# 	if request.user.profile.user_type == 'ADMIN':
		# 		messages.error(request, 'You have no permission to access the requested resource!')
		# 		return redirect(reverse('hr:employee_dashboard'))
		# 	elif request.user.profile.user_type == 'CUSTOMER':
		# 		messages.error(request, 'You have no permission to access the requested resource!')
		# 		return redirect(reverse('index'))
		# except AttributeError as error:
		# 	messages.error(request, 'You have no permission to access the requested resource!')
		# 	return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)

	def get_success_url(self, **kwargs):

		if (self.request.user.id == 1):
			messages.success(self.request, 'BLF Profile Updated Successfully')
			return get_safe_next_url(self.request, _profile_list_default("blf"))
		else:
			messages.success(self.request, 'BLF Profile Updated Successfully')
			return reverse('user_profile:blf_profile_create', kwargs={"user_create_pk": self.kwargs['user_create_pk'] } )
			

	def get_object(self):
		profile_details = BlfProfile.objects.filter(user_id=self.kwargs['user_create_pk']).first()
		return profile_details

	def get_context_data(self, **kwargs):
		context = super(BlfProfileUpdateView, self).get_context_data(**kwargs)
		context['profile_details'] = self.get_object()
		if self.request.POST:
			context['blf_tea_production_row_formset'] = BLF_TEA_PRODUCTION_FORM_SET(self.request.POST, instance=self.get_object())
		else:
			context['blf_tea_production_row_formset'] = BLF_TEA_PRODUCTION_FORM_SET(instance=self.get_object())

		if self.request.POST:
			context['blf_through_formset'] = BLF_TROUGH_FORMSET(self.request.POST, instance=self.get_object())
		else:
			context['blf_through_formset'] = BLF_TROUGH_FORMSET(instance=self.get_object())
		context["current_url"] = True
		context["next_url"] = get_safe_next_url(self.request, _profile_list_default("blf"))
		return context


	def form_valid(self, form):
		self.user_id = self.kwargs['user_create_pk']
		context = self.get_context_data()
		blf_tea_production_row_formset = context['blf_tea_production_row_formset']
		blf_through_formset = context['blf_through_formset']

		with transaction.atomic():
			self.object = form.save()

			if blf_tea_production_row_formset.is_valid():
				blf_tea_production_row_formset.instance = self.object
				blf_tea_production_row_formset.save()

			if blf_through_formset.is_valid():
				blf_through_formset.instance = self.object
				blf_through_formset.save()

			if not form.is_valid()  or not blf_tea_production_row_formset.is_valid() or not blf_through_formset.is_valid():
				# formsets or form is invalid; render the form with error
				return self.render_to_response(context)

		return super(BlfProfileUpdateView, self).form_valid(form)


class BlfProfileDetailView(LoginRequiredMixin, CommonMixin, DetailView):
	""" Blf Profile Details View """
	context_object_name = 'blf_detail'
	model = BlfProfile
	template_name = "profile/blf_profile_view.html"

	def get_object(self):
		blf_detail = BlfProfile.objects.filter(
			user_id=self.kwargs['user_create_pk']).first()
		return blf_detail

	def get_context_data(self, **kwargs):
		context = super(BlfProfileDetailView,
						self).get_context_data(**kwargs)

		blf_details = self.get_object()

		if self.request.POST:
			context['blf_tea_production_row_formset'] = BLF_TEA_PRODUCTION_FORM_SET(self.request.POST, instance=self.get_object())
		else:
			context['blf_tea_production_row_formset'] = BLF_TEA_PRODUCTION_FORM_SET(instance=self.get_object())

		if self.request.POST:
			context['blf_through_formset'] = BLF_TROUGH_FORMSET(self.request.POST, instance=self.get_object())
		else:
			context['blf_through_formset'] = BLF_TROUGH_FORMSET(instance=self.get_object())

		context['form'] = BlfProfileForm(instance=blf_details)
		context["next_url"] = get_safe_next_url(self.request, _profile_list_default("blf"))

		return context



# API views to fetch Entity details

def get_entity_list_view(request):

	entity_list = get_entity_list_from_api()
	unit_list = get_unit_list_from_api()

	context = {
		'entity_list' : entity_list,
		'unit_list' : unit_list
	}
	return CommonMixin.render(request, 'profile/api_entity_details.html', context)



def ajax_entity_unit_list(request):
	unit_list = get_unit_list_from_api()
	# cities = City.objects.filter(country_id=country_id).order_by('name')
	# print("Subhashis")
	context = {
		
	}
	return render(request, 'profile/unit_list.html', context)
	

def unit_details(request):

	id = request.GET.get('id', None)


	value = get_unit_list_details_from_api(id)

	data = {
		"value" : value
	}

	return JsonResponse(data)
	# return render(request, 'profile/unit_list.html', JsonResponse(data))


def entity_details(request):
	id = request.GET.get('id', None)
	value = get_entity_list_details_from_api(id)
	data = {
		"value" : value
	}
	return JsonResponse(data)



# def create_factory_marks(request):

# 	print("############## Here isb post ############")

# 	if request.method == "POST":
# 		form = FactoryDetailsMarksForm(request.POST)
# 		if form.is_valid():
# 			name = request.POST.get('name')
# 			created = FactoryDetailsMarks.objects.create(name=name)
# 			created.save()
			
# 			print("Frm Submited ############## ")
# 			# response_data = {}
# 			# created = FactoryDetailsMarks.objects.create(name=name)
# 			# created.save()
# 			# post = FactoryDetailsMarks(name=name)
# 			# post.save()

# 			# response_data['result'] = 'Create post successful!'
# 			# response_data['name'] = post.name

# 			messages.success(request, 'Successfully Created!')
# 		else:
# 			messages.error(request, 'Please Correct The Error Below.')
# 	else:
# 		form = FactoryDetailsMarksForm()

# 	# return HttpResponse(
# 	# 		json.dumps(response_data),
# 	# 		content_type="application/json"
# 	# 	)
# 	return HttpResponseRedirect(request.META.get("HTTP_REFERER"))




########## Estate PROFILE VIEW ###############

class EstateCreateView(LoginRequiredMixin, CommonMixin, UpdateView):
	"""
	Estate Profile Create View 
	"""
	model = EstateProfile
	form_class = EstateProfileForm
	template_name = 'estate_create.html'

	def get(self, request, *args, **kwargs):
	# checking if the user is Admin
		try:
			if not self.request.user.is_superuser:
				messages.error(self.request, 'You have no permission to access the requested resource!')
				return redirect(reverse('index'))
		except AttributeError as error:
			messages.error(self.request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)


	def get_success_url(self, **kwargs):
		messages.success(self.request, 'Estate Profile Updated Successfully')
		factory_details = FactoryTroughDetails.objects.filter(user_id=self.kwargs['user_create_pk']).first()

		return reverse('user_profile:estate_trough_details_update', kwargs={"user_create_pk": self.kwargs['user_create_pk'], \
								       "factory_pk": factory_details.pk})

	def get_object(self):	
		profile_details = EstateProfile.objects.filter(user_id=self.kwargs['user_create_pk']).first()
		return profile_details

	def get_context_data(self, **kwargs):
		context = super(EstateCreateView, self).get_context_data(**kwargs)
		context['profile_details'] = self.get_object()
		if self.request.POST:
			context['estate_production_row_formset'] = ESTATE_PRODUCTIONS_FORM_SET(self.request.POST, instance=self.get_object())
		else:
			context['estate_production_row_formset'] = ESTATE_PRODUCTIONS_FORM_SET(instance=self.get_object())

		return context
	
	def form_invalid(self, form):
		error_message = 'Error saving the Form, check fields below.'
		messages.error(self.request, error_message)
		return super().form_invalid(form)

	def form_valid(self, form):

		self.user_id = self.kwargs['user_create_pk']
		context = self.get_context_data()

		estate_production_row_formset = context['estate_production_row_formset']

		FactoryTroughDetails.objects.update_or_create(user_id=self.kwargs['user_create_pk'], defaults={
			'estate': form.instance,})

		with transaction.atomic():
			self.object = form.save()

			if estate_production_row_formset.is_valid():
				estate_production_row_formset.instance = self.object
				estate_production_row_formset.save()

			if not form.is_valid() or not estate_production_row_formset.is_valid():
				return self.render_to_response(context)

		return super(EstateCreateView, self).form_valid(form)
	


class EstateProfileUpdateView(LoginRequiredMixin, CommonMixin, UpdateView):
	"""
	Estate PROFILE Update View
	"""
	model = EstateProfile
	form_class = EstateProfileForm
	template_name = 'profile/estate_profile_create.html'

	def get(self, request, *args, **kwargs):
	# checking if the user is Admin
		try:
			if not self.request.user.is_superuser:
				messages.error(self.request, 'You have no permission to access the requested resource!')
				return redirect(reverse('index'))
		except AttributeError as error:
			messages.error(self.request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)

	def get_success_url(self, **kwargs):
		messages.success(self.request, 'Estate Profile Updated Successfully')
		return get_safe_next_url(self.request, _profile_list_default("estate"))

	def get_object(self):
		profile_details = EstateProfile.objects.filter(user_id=self.kwargs['user_create_pk']).first()
		return profile_details

	def get_context_data(self, **kwargs):
		context = super(EstateProfileUpdateView, self).get_context_data(**kwargs)

		context['profile_details'] = self.get_object()

		if self.request.POST:
			context['estate_production_row_formset'] = ESTATE_PRODUCTIONS_FORM_SET(self.request.POST, instance=self.get_object())
		else:
			context['estate_production_row_formset'] = ESTATE_PRODUCTIONS_FORM_SET(instance=self.get_object())
		context["current_url"] = True
		context["next_url"] = get_safe_next_url(self.request, _profile_list_default("estate"))
		return context
	
	
	def form_invalid(self, form):
		error_message = 'Error saving the Form, check fields below.'
		messages.error(self.request, error_message)
		return super().form_invalid(form)

	def form_valid(self, form):

		self.user_id = self.kwargs['user_create_pk']
		context = self.get_context_data()

		estate_production_row_formset = context['estate_production_row_formset']

		# FactoryTroughDetails.objects.update_or_create(user_id=self.kwargs['user_create_pk'], defaults={
		# 	'estate': form.instance,})

		with transaction.atomic():
			self.object = form.save()

			if estate_production_row_formset.is_valid():
				estate_production_row_formset.instance = self.object
				estate_production_row_formset.save()

			if not form.is_valid() or not estate_production_row_formset.is_valid():
				return self.render_to_response(context)

		return super(EstateProfileUpdateView, self).form_valid(form)
	


class EstateTroughList(LoginRequiredMixin, CommonMixin, ListView):
	"""
	Estate Trough Details List View
	"""
	model = FactoryTroughDetails
	context_object_name = 'factory_trough_list'
	template_name = 'factory_trough_list.html'
	paginate_by = 10

	def get(self, request, *args, **kwargs):
	# checking if the user is customer
		try:
			if not self.request.user.is_superuser:
				messages.error(self.request, 'You have no permission to access the requested resource!')
				return redirect(reverse('index'))
		except AttributeError as error:
			messages.error(self.request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)

	def get_queryset(self):
		print(self.kwargs['user_create_pk'])
		return FactoryTroughDetails.objects.filter(user_id=self.kwargs['user_create_pk']).order_by('-id')

	def get_context_data(self, **kwargs):
		context = super().get_context_data(**kwargs)
		context['profile_type'] = "estate"
		context['user_id'] = self.kwargs['user_create_pk']
		print(self.kwargs['user_create_pk'])
		return context




class EstateTroughUpdate(LoginRequiredMixin, CommonMixin, UpdateView):

	model = FactoryTroughDetails
	form_class = FactoryTroughDetailsForm
	template_name = 'estate_trough_form.html'

	def get(self, request, *args, **kwargs):
	# checking if the user is customer
		try:
			if not self.request.user.is_superuser:
				messages.error(self.request, 'You have no permission to access the requested resource!')
				return redirect(reverse('index'))
		except AttributeError as error:
			messages.error(self.request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)

	def get_success_url(self, **kwargs):
		messages.success(self.request, 'Updated Successfully')
		profile_details = EstateProfile.objects.filter(user_id=self.kwargs['user_create_pk']).first()
		return reverse('user_profile:estate_trough_details_list', kwargs={"user_create_pk": self.kwargs['user_create_pk'] })
		# return reverse('gardens_managment:garden_details', kwargs={"grower_pk": grower_details })

	def get_object(self):
		profile_details = EstateProfile.cmobjects.filter(user_id=self.kwargs['user_create_pk']).first()
		factory_through_details = FactoryTroughDetails.objects.filter(pk=self.kwargs['factory_pk']).first()
		return factory_through_details

	def get_context_data(self, **kwargs):
		context = super(EstateTroughUpdate, self).get_context_data(**kwargs)
		context['factory_through_details'] = self.get_object()

		context['profile_type'] = "estate"
		context['user_id'] = self.kwargs['user_create_pk']

		if self.request.POST:
			context['estate_trough_form_set'] = ESTATE_TROUGH_FORM_SET(self.request.POST, instance=self.get_object())	
		else:
			context['estate_trough_form_set'] = ESTATE_TROUGH_FORM_SET(instance=self.get_object())

		return context

	def form_valid(self, form):
		profile_details = EstateProfile.cmobjects.filter(user_id=self.kwargs['user_create_pk']).first()
		print(profile_details.pk)
		self.estate_id = profile_details.pk
		context = self.get_context_data()
		estate_trough_form_set = context['estate_trough_form_set']

		with transaction.atomic():
			self.object = form.save()

			if estate_trough_form_set.is_valid():
				estate_trough_form_set.instance = self.object
				estate_trough_form_set.save()

			if not form.is_valid() or not estate_trough_form_set.is_valid():
				# formsets or form is invalid; render the form with error
				return self.render_to_response(context)

		return super(EstateTroughUpdate, self).form_valid(form)



class EstateTroughCreate(LoginRequiredMixin, CommonMixin, CreateView):

	model = FactoryTroughDetails
	form_class = FactoryTroughDetailsForm
	template_name = 'estate_trough_form.html'

	def get(self, request, *args, **kwargs):
	# checking if the user is customer
		try:
			if not self.request.user.is_superuser:
				messages.error(self.request, 'You have no permission to access the requested resource!')
				return redirect(reverse('index'))
		except AttributeError as error:
			messages.error(self.request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)

	def get_success_url(self, **kwargs):
		messages.success(self.request, 'Updated Successfully')
		profile_details = EstateProfile.objects.filter(user_id=self.kwargs['user_create_pk']).first()
		return reverse('user_profile:estate_trough_details_list', kwargs={"user_create_pk": self.kwargs['user_create_pk'] })
		# return reverse('gardens_managment:garden_details', kwargs={"grower_pk": grower_details })

	def get_context_data(self, **kwargs):
		context = super(EstateTroughCreate, self).get_context_data(**kwargs)
		context['profile_type'] = "estate"
		context['user_id'] = self.kwargs['user_create_pk']

		if self.request.POST:
			context['estate_trough_form_set'] = ESTATE_TROUGH_FORM_SET(self.request.POST)	
		else:
			context['estate_trough_form_set'] = ESTATE_TROUGH_FORM_SET()
		return context

	def form_valid(self, form):
		profile_details = EstateProfile.cmobjects.filter(user_id=self.kwargs['user_create_pk']).first()
		context = self.get_context_data()
		estate_trough_form_set = context['estate_trough_form_set']

		with transaction.atomic():
			form.instance.user_id = self.kwargs['user_create_pk']
			form.instance.estate_id = profile_details.pk

			self.object = form.save()
			if estate_trough_form_set.is_valid():
				estate_trough_form_set.instance = self.object
				estate_trough_form_set.save()

			if not form.is_valid() or not estate_trough_form_set.is_valid():
				# formsets or form is invalid; render the form with error
				return self.render_to_response(context)

		return super(EstateTroughCreate, self).form_valid(form)


def estate_factory_trough_details(request, user_create_pk, factory_pk):
	factory_details = FactoryTroughDetails.objects.filter(pk=factory_pk).first()
	trough_list = EstateTroughDetails.objects.filter(factory_id=factory_pk).order_by('-id')
	context={
		'factory_details' : factory_details,
		'user_id' : user_create_pk, 
		'trough_list' : trough_list,
	}

	return CommonMixin.render(request, 'factory_trough_details.html', context)


def estate_factory_delete(request, factory_pk, user_create_pk):
	with transaction.atomic():
		factory_details = FactoryTroughDetails.objects.filter(pk=factory_pk).first()
		factory_details.delete()
		messages.success(
			request, 'Deleted Successfully')

	return redirect(reverse('user_profile:estate_trough_details_list', kwargs={"user_create_pk": user_create_pk }))

	
# PROFILE UPDATE VIEW

class ProfileUpdateView(LoginRequiredMixin, CommonMixin, UpdateView):
	"""
	PROFILE Update View
	"""
	model = Profile
	form_class = ProfileForm
	template_name = 'edit_profile.html'

	def get(self, request, *args, **kwargs):
	# checking if the user is Admin
		try:
			if not self.request.user.is_superuser:
				messages.error(self.request, 'You have no permission to access the requested resource!')
				return redirect(reverse('index'))
		except AttributeError as error:
			messages.error(self.request, 'You have no permission to access the requested resource!')
			return redirect(reverse('index'))

		return super().get(request, *args, **kwargs)

	def get_success_url(self, **kwargs):
		messages.success(self.request, 'Profile Updated Successfully')
		return reverse('user_profile:profile_edit', kwargs={"user_create_pk": self.kwargs['user_create_pk'] }  )


	def get_object(self):
		profile_details = Profile.objects.filter(user_id=self.kwargs['user_create_pk']).first()
		return profile_details

	def get_context_data(self, **kwargs):
		context = super(ProfileUpdateView, self).get_context_data(**kwargs)

		context['profile_details'] = self.get_object()

		return context
	
	def form_invalid(self, form):
		error_message = 'Error saving the Form, check fields below.'
		messages.error(self.request, error_message)
		return super().form_invalid(form)

	def form_valid(self, form):

		self.user_id = self.kwargs['user_create_pk']
		context = self.get_context_data()

		with transaction.atomic():
			self.object = form.save()

			if not form.is_valid():
				return self.render_to_response(context)

		return super(ProfileUpdateView, self).form_valid(form)
	



def qr_code_generator(request):
	import pyqrcode
	import png
	from pyqrcode import QRCode

	s = "www.geeksforgeeks.org"

	url = pyqrcode.create(s)

	# Create and save the svg file naming "myqr.svg"
	url.svg("myqr.svg", scale = 8)

	# Create and save the png file naming "myqr.png"
	url.png('myqr.png', scale = 6)


	context = {

	}

	return render(request, 'qr_code.html', context)



# QR CODE VIEWS

def grower_qr_code(request, grower_pk):
   
   qr_code=GrowerQrCode.objects.filter(grower_id=grower_pk).first()
   print()
   context ={
		'qr_code':qr_code,
   }
   return render(request,"qrcode.html", context)



def aggregator_qr_code(request, aggragator_pk):
   
   qr_code=AggregatorQrCode.objects.filter(aggregator_id=aggragator_pk).first()
   
   context ={
		'qr_code':qr_code,
   }

   return render(request,"qrcode.html", context)


###############   Grower Form    ################

# def blf_list_region_wise(request):
# 	""" Region   """
# 	print("HEllo Region wise Entity")
# 	id = request.GET.get('id', None)
# 	print(id)
# 	# supply_date = weighment_details.supply_date
# 	blf_list_region_wise = BlfProfile.cmobjects.filter(region__region_name__exact=id).values('id', 'user')
# 	print(blf_list_region_wise)

# 	options = '<option value="">---------</option>'
# 	for subcategory in blf_list_region_wise:
# 		options += f'<option value="{subcategory.id}">{subcategory.user}</option>'

# 	# value = list(blf_list_region_wise)

# 	data = {'options': options}

# 	# data = {
# 	# 	"value" : value,
# 	# }
# 	return JsonResponse(data)


# def entity_details(request):
# 	id = request.GET.get('id', None)
# 	value = get_entity_list_details_from_api(id)
# 	data = {
# 		"value" : value
# 	}
# 	return JsonResponse(data)


# def region_list_easy_mode_wise(request):
# 	mode = request.GET.get('mode', None)
# 	mode_wise_region = BlfProfile.objects.filter(easy_weight_system=mode)
# 	# Get a list of distinct values for the easy_weight_system field
# 	# distinct_values = mode_wise_region.values_list('easy_weight_system', flat=True).distinct()
# 	# distinct_values = list(set(mode_wise_region.values_list('easy_weight_system', flat=True)))

# 	# print(f"VALUES = {distinct_values}")

# 	context = {
# 		'mode_wise_region': mode_wise_region
# 	}
# 	return render(request, 'mode_wise_region.html', context)


def region_list_easy_mode_wise(request):
	mode = request.GET.get('mode', None)
	mode_wise_profiles = BlfProfile.objects.filter(easy_weight_system=mode).exclude(region__isnull=True)
	# Get a list of distinct region names
	distinct_regions = mode_wise_profiles.values_list('region__region_name', flat=True).distinct()

	context = {
		'mode_wise_region': distinct_regions,
		'mode_wise_profiles' : mode_wise_profiles,
		'selected_mode': mode,
	}
	return render(request, 'mode_wise_region.html', context)




def blf_list_region_wise_back(request):
	
	id = request.GET.get('id', None)
	mode = str(request.GET.get('mode', None))
	if not id.isdigit():
		# id is a positive integer
		id_str = str(id)
		region_details = Region.cmobjects.filter(region_name__exact=id_str).first()
		id = region_details.pk

	if mode:
		blf_list_region_wise = BlfProfile.cmobjects.filter(region_id=id, easy_weight_system__exact=mode)
	else:
		blf_list_region_wise = BlfProfile.cmobjects.filter(region_id=id)

	context = {
		'blf_list_region_wise': blf_list_region_wise
	}

	return render(request, 'blf_ajax_list.html', context)




def blf_list_region_wise(request):
	id = request.GET.get('id', None)
	if not id.isdigit():
		# id is a positive integer
		id_str = str(id)
		region_details = Region.cmobjects.filter(region_name__exact=id_str).first()
		id = region_details.pk
	
	blf_list_region_wise = BlfProfile.cmobjects.filter(region_id=id)
	context = {
		'blf_list_region_wise': blf_list_region_wise
	}
	return render(request, 'blf_ajax_list.html', context)





def aggregator_list_blf_wise(request):
	id = request.GET.get('id', None)
	agg_list_blf_wise = AggregatorProfile.objects.filter(associated_entity=id).order_by('-id')
	context = {
		'agg_list_blf_wise': agg_list_blf_wise
	}
	return render(request, 'agg_list_blf_wise.html', context)

###### Visiting Card #################################################
def generate_qr_code(qr_code, user_id,user_type):
		qr = qrcode.QRCode(
		    version=1,
		    error_correction=qrcode.constants.ERROR_CORRECT_L,
		    box_size=10,
		    border=4,
		)
		qr.add_data(qr_code)

		qr.make(fit=True)
		qr_code_image = qr.make_image(fill_color="black", back_color="white")	
		# Specify the directory where you want to save the QR code image
		qr_code_directory = os.path.join(settings.MEDIA_DIR, "qrcode")  # Create a 'qrcodes' directory
		os.makedirs(qr_code_directory, exist_ok=True)		
		# Specify the filename for the QR code image
		qr_code_image_filename = f"{user_type}_{user_id}_qrcode.png"	
		# Save the QR code image using the file storage
		qr_code_image_path = os.path.join(qr_code_directory, qr_code_image_filename)
		qr_code_image.save(qr_code_image_path)		
		
		return qr_code_image_path
def GrowerVistingCard(request, grower_pk):
   
	details=GrowerProfile.objects.filter(id=grower_pk).first()
	if details:
		user_type=details.profile_type.name
		# qr_code=GrowerProfile.objects.filter(user_id=user).first()	
		user_id=details.user
		# qr_code= "Grower ID : " + str(details.user) + "\n" + "Name : " + str(details.name)
		# print("print qr",qr_code)
		qr_code = {
			'userid': details.user.id,
			'name': details.name,
			'id': details.id,
			'username': details.user.username
					}
		qr_code_image_path = generate_qr_code(qr_code,user_id ,user_type)

		media_url = settings.MEDIA_URL.rstrip('/')  # Remove trailing slash if present
		absolute_qr_code_image_url = request.build_absolute_uri(qr_code_image_path.replace(settings.MEDIA_DIR, media_url).replace("\\", "/"))
		if details is not None:
			if details.photo:
				details_photo_url = request.build_absolute_uri(details.photo.url)
			else:
				details_photo_url = None
		else:
			details_photo_url = None
		context ={
			'details':details,
			'user_type':user_type,
			'details_photo_url':details_photo_url,
			'absolute_qr_code_image_url':absolute_qr_code_image_url

		}
		return render(request,"visiting.html", context)	
	else:
		messages.error(request, "Grower details not found")
		# Return an appropriate response in case details are not found
		return HttpResponse("Grower details not found")
def AggregatorVistingCard(request, aggregator_pk):
   
	details=AggregatorProfile.objects.filter(id=aggregator_pk).first()
	if details:
		user_type=details.profile_type.name
		# qr_code=GrowerProfile.objects.filter(user_id=user).first()	
		user_id=details.user
		# qr_code= "Aggregator ID : " + str(details.user) + "\n" + "Name : " + str(details.name)
		qr_code = {
		'userid': details.user.id,
		'name': details.name,
		'id': details.id,
		'username': details.user.username
				}
		qr_code_image_path = generate_qr_code(qr_code,user_id ,user_type)

		media_url = settings.MEDIA_URL.rstrip('/')  # Remove trailing slash if present
		absolute_qr_code_image_url = request.build_absolute_uri(qr_code_image_path.replace(settings.MEDIA_DIR, media_url).replace("\\", "/"))
		if details is not None:
			if details.user_file:
				details_photo_url = request.build_absolute_uri(details.user_file.url)
			else:
				details_photo_url = None
		else:
			details_photo_url = None
		context ={
			'details':details,
			'user_type':user_type,
			'details_photo_url':details_photo_url,
			'absolute_qr_code_image_url':absolute_qr_code_image_url

	}
	return render(request,"visiting.html", context)	
	    


#### AJAX TEA GRADE LOAD

def blf_tea_grade_ajax(request):
	tea_type = request.GET.get('tea_type', None)
	print("TEA GRADE dwdwd ########################",  tea_type)
	tea_type = request.GET.get('tea_type', None)
	print("Tea Type ###################gggw", tea_type)
	tea_grade_list = TeaGradeDetails.objects.filter(tea_type_id=tea_type).order_by('-id')
	print("tea_grade_listt", tea_grade_list)
	context = {
		'tea_grade_list': tea_grade_list
	}
	return render(request, 'tea_type_ajax.html', context)



	# logged_user_type = Profile.objects.filter(user_id=request.user.id).first()
	# user_type = logged_user_type.user_type
	# try:
	# 	if not str(user_type.name) == "blf" and request.user.is_authenticated:
	# 		messages.error(request, 'You have no permission to access the requested resource!')
	# 		return redirect(reverse('index'))
	# except AttributeError as error:
	# 	messages.error(request, 'You have no permission to access the requested resource!')
	# 	return redirect(reverse('index'))



	# range = range_details.bag_sl_no_range
	# print(range)

	# if request.is_ajax():
	# 	html = render_to_string('tea_type_ajax.html',
	# 							context, request=request)
	# data = {
	# 	'html': html
	# }

	# data = {
	# 	'value': tea_type
	# }

	# return JsonResponse(data)

	# context = {
	# 	'tea_grade_list': tea_grade_list
	# }

	# return render(request, 'tea_type_ajax.html', context)
	# 	}
	# 	return render(request,"visiting.html", context)	
	# else:
	# 	messages.error(request, "Aggregator details not found")
	# 	# Return an appropriate response in case details are not found
	# 	return HttpResponse("Aggregator details not found")

from django.shortcuts import render
def DownloadVistingCard(request, user_pk, user_type):
    print('user_type', user_type)

    if user_type.lower() == 'grower':
        details = GrowerProfile.objects.filter(user_id=user_pk).first()
        if details:
            print("details", details)
            user_type = details.profile_type.name
            user_id = details.user
            
            qr_code = {
                'userid': details.user.id,
                'name': details.name,
                'id': details.id,
                'username': details.user.username
            }
            
            qr_code_image_path = generate_qr_code(qr_code, user_id, user_type)
            media_url = settings.MEDIA_URL.rstrip('/')  # Remove trailing slash if present
            absolute_qr_code_image_url = request.build_absolute_uri(qr_code_image_path.replace(settings.MEDIA_DIR, media_url).replace("\\", "/"))
            
            if details is not None:
                if details.photo:
                    details_photo_url = request.build_absolute_uri(details.photo.url)
                else:
                    details_photo_url = None
            else:
                details_photo_url = None
            
            context = {
                'details': details,
                'user_type': user_type,
                'details_photo_url': details_photo_url,
                'absolute_qr_code_image_url': absolute_qr_code_image_url
            }
            
            template_path = 'visiting_download.html'
            html = get_template(template_path).render(context)
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="visiting_card.pdf"'
            
            # Generate PDF from HTML content
            pisa_status = pisa.CreatePDF(html, dest=response)
            if pisa_status.err:
                error_message = f"Something went wrong"
                return render(request, 'error.html', {'error_message': error_message})
            
            return response
            
        else:
            error_message = f"Something went wrong"
            return render(request, 'error.html', {'error_message': error_message})

    elif user_type.lower() == 'aggregator':
        details = AggregatorProfile.objects.filter(user_id=user_pk).first()
        if details:
            user_type = details.profile_type.name
            user_id = details.user

            qr_code = {
                'userid': details.user.id,
                'name': details.name,
                'id': details.id,
                'username': details.user.username
            }

            qr_code_image_path = generate_qr_code(qr_code, user_id, user_type)
            media_url = settings.MEDIA_URL.rstrip('/')  # Remove trailing slash if present
            absolute_qr_code_image_url = request.build_absolute_uri(qr_code_image_path.replace(settings.MEDIA_DIR, media_url).replace("\\", "/"))

            if details is not None:
                if details.user_file:
                    details_photo_url = request.build_absolute_uri(details.user_file.url)
                else:
                    details_photo_url = None
            else:
                details_photo_url = None

            context = {
                'details': details,
                'user_type': user_type,
                'details_photo_url': details_photo_url,
                'absolute_qr_code_image_url': absolute_qr_code_image_url
            }
            template_path = 'visiting_download.html'
            html = get_template(template_path).render(context)

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="visiting_card.pdf"'

            # Generate PDF from HTML content
            pisa_status = pisa.CreatePDF(html, dest=response)
            if pisa_status.err:
                error_message = f"Something went wrong"
                return render(request, 'error.html', {'error_message': error_message})

            return response

        else:
            error_message = f"Something went wrong"
            return render(request, 'error.html', {'error_message': error_message})


		









#####################  Reports View ######################

def aggregator_register_list_blf_report(request):
	
		regions = Region.objects.all()  # Fetch all regions
		selected_region_id = request.GET.get('region')  # Get the selected region ID from the request
		selected_blf_id = request.GET.get('entity_id') 
    	# Fetch all entities initially
		region_name=None
		entities=None
    	# If a region is selected, filter entities based on that region

		blf_name=None
		blf_username=None
		result = []
		print()
		
    	# If both region and BLF are selected, filter Aggregator details
		if selected_region_id and selected_blf_id:
			selected_blf = BlfProfile.objects.get(pk=selected_blf_id)
			region_name=Region.objects.filter(id=selected_region_id).first()
			blf_name=selected_blf.entity_unit
			blf_username=selected_blf.user.username
			result = AggregatorProfile.cmobjects.filter(associated_entity=selected_blf)
		# Paginate the result
		paginator = Paginator(result, 10)
		page = request.GET.get('page', 1)

		try:
			result = paginator.page(page)
		except PageNotAnInteger:
			result = paginator.page(1)
		except EmptyPage:
			result = paginator.page(paginator.num_pages)	
		 # Calculate the serial number adjusted for pagination
		serial_number_offset = (result.number - 1) * paginator.per_page
		for index, item in enumerate(result, start=1):
			item.serial_number = index + serial_number_offset		
		total_items = paginator.count


		blf_list = BlfProfile.cmobjects.filter(region=selected_region_id)


		context = {
		    "region_list": regions,
			"region_name":region_name,
		    "blf_name": blf_name ,
			"blf_username":blf_username,
		    "result": result,
		    "total_items": total_items,
			"selected_blf_id": selected_blf_id,
			'entity_id':selected_blf_id,
			'region':selected_region_id,
			"selected_region_id":selected_region_id,
			"blf_list":blf_list,
		}

		return CommonMixin.render(request, "reports/report_aggregator_register_list.html", context)



from django.utils import timezone

def report_aggregator_register_pdf(request):
	""" Register Aggregator Report generate pdf View """
	try:
	
		regions = Region.objects.all()  # Fetch all regions
		selected_region_id = request.GET.get('selected_region_id',None)
		selected_blf_id = request.GET.get('selected_blf_id',None)
		print("##########test",request.GET.get('selected_region_id'))
    	# Fetch all entities initially
		region_name=None
		entities=None
    	# If a region is selected, filter entities based on that region

		blf_name=None
		blf_username=None
		blf_address=None
		result = []
		selected_blf=None
    	 # Check if both region and BLF IDs are not None
		if selected_region_id is not None and selected_blf_id is not None:
			try:
				# Convert IDs to integers
				selected_region_id = int(selected_region_id)
				selected_blf_id = int(selected_blf_id)
				# Fetch the selected BLF and region
				selected_blf = BlfProfile.objects.get(pk=selected_blf_id)
				region_name = Region.objects.filter(id=selected_region_id).first()
				# Assign values to variables
				blf_name = selected_blf.entity_unit
				blf_username = selected_blf.user.username
				blf_address=selected_blf.address
				# Filter Aggregator details
				result = AggregatorProfile.cmobjects.filter(associated_entity=selected_blf)
			except (ValueError, BlfProfile.DoesNotExist, Region.DoesNotExist):
				# Handle exceptions or incorrect IDs here
				pass
		 # Get Indian Standard Time (IST)
		tz = pytz.timezone('Asia/Kolkata')  # Indian Standard Time
		current_date = timezone.now().astimezone(tz).strftime("%d-%m-%Y") 
		for item in result:
			if len(item.name) > 8 and ' ' not in item.name:
				item.name = ' '.join([item.name[i:i+8] for i in range(0, len(item.name), 8)]) 
		
			if len(item.gstin_no or '') > 8 and ' ' not in item.gstin_no:
				item.gstin_no = ' '.join([item.gstin_no[i:i + 8] for i in range(0, len(item.gstin_no), 8)])

			if len(item.address or '') > 8 and ' ' not in item.address:
				item.address = ' '.join([item.address[i:i + 8] for i in range(0, len(item.address), 8)])	 
			if len(item.user.username) > 8 and ' ' not in item.user.username:
				item.user.username = ' '.join([item.user.username[i:i + 8] for i in range(0, len(item.user.username), 8)])
		context = {
		    "regions": regions,
			"region_name":region_name,
		    "blf_name": blf_name ,
			"blf_username":blf_username,
		    "result": result,
			'blf_address':blf_address,
			"selected_blf_id": selected_blf_id,
			'selected_blf':selected_blf,
			"selected_region_id":selected_region_id,
			'current_date': current_date,

		}
		# Create a response with PDF content
		response = HttpResponse(content_type='application/pdf')
		blf_name_for_filename = selected_blf.entity_unit.replace(' ', '_') if selected_blf else 'aggregator_lists'
		filename = f"{blf_name_for_filename}_aggregator_lists.pdf"
		response['Content-Disposition'] = f'attachment; filename="{filename}"'
		# response['Content-Disposition'] = f'attachment; filename="aggregator_registered.pdf"'

		# Create an HTML template
		template = get_template('reports/report_aggregator_register_pdf.html')
		# template = get_template('reports/report_test_pdf.html')
		html = template.render(context)  # Replace with your template context data

		# Generate PDF
		pisa.CreatePDF(html, dest=response)
		return response
	except (ValueError, BlfProfile.DoesNotExist, Region.DoesNotExist) as e:
		error_message = "Something went wrong."
		return render(request, 'error.html', {'error_message': error_message}) 
	except Exception as e:
		error_message = "Something went wrong."
		return render(request, 'error.html', {'error_message': error_message}) 




def report_aggregator_register_excel(request):
	try:
		selected_region_id = request.GET.get('selected_region_id', None)
		selected_blf_id = request.GET.get('selected_blf_id', None)
	
		result = []
		selected_blf=None
	
		# Fetch Aggregator details
		if selected_region_id is not None and selected_blf_id is not None:
			try:
				# Convert IDs to integers
				selected_region_id = int(selected_region_id)
				selected_blf_id = int(selected_blf_id)
				# Fetch the selected BLF and region
				selected_blf = BlfProfile.objects.get(pk=selected_blf_id)
				
				# Filter Aggregator details
				result = AggregatorProfile.objects.filter(associated_entity=selected_blf)
			except (ValueError, BlfProfile.DoesNotExist, Region.DoesNotExist):
				# Handle exceptions or incorrect IDs here
				pass
		# Create a workbook and add a worksheet
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "Aggregator Register Report"
		# Define headers for the Excel file
		headers = ["SL No.", "Tracetea ID", "Name", "Type", "Region", "State", "District", "Address", "Mobile No.", "GISTIN ID"]
		# Add headers to the worksheet
		for col_num, header in enumerate(headers, 1):
			col_letter = get_column_letter(col_num)
			ws[f"{col_letter}1"] = header
		# Populate data into the worksheet
		row_num = 2  # Start from the second row
		for index, item in enumerate(result, start=1):
			ws[f"A{row_num}"] = index  # SL No.
			ws[f"B{row_num}"] = item.user.username or "NA"  
			ws[f"C{row_num}"] = item.name or "NA" 
			ws[f"D{row_num}"] = item.aggregator_type or "NA" 
			ws[f"E{row_num}"] = item.region.region_name if item.region else "NA"  
			ws[f"F{row_num}"] = item.state.name if item.state else "NA" 
			ws[f"G{row_num}"] = item.district.name if item.district else "NA" 
			ws[f"H{row_num}"] = item.address or "NA" 
			ws[f"I{row_num}"] = item.mobile_number or "NA" 
			ws[f"J{row_num}"] = item.gstin_no or "NA"  
			row_num += 1  # Move to the next row for the next data item	
			# Create an HTTP response with the Excel file as an attachment
			response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
			blf_name_for_filename = selected_blf.entity_unit.replace(' ', '_') if selected_blf else 'aggregator_lists'
			filename = f"{blf_name_for_filename}_aggregator_lists.xlsx"
			response['Content-Disposition'] = f'attachment; filename="{filename}"'
			# response['Content-Disposition'] = 'attachment; filename="aggregator_registered_report.xlsx"'	
		# Save the workbook content to the response
		wb.save(response)
		return response
	except (ValueError, BlfProfile.DoesNotExist, Region.DoesNotExist) as e:
		error_message = "Something went wrong."
		return render(request, 'error.html', {'error_message': error_message}) 
	except Exception as e:
		error_message = "Something went wrong."
		return render(request, 'error.html', {'error_message': error_message}) 

from django.shortcuts import get_object_or_404



@login_required
def grower_register_list_blf_report(request):
	"""Grower register of BLF or entity report"""
	regions = Region.objects.all()  # Fetch all regions
	selected_region_id = request.GET.get('region')  # Get the selected region ID from the request
	selected_blf_id = request.GET.get('entity_id')    
	region_name = None
	blf_name = None
	blf_username = None
	result = []
	serial_number = 1

	blf_list = BlfProfile.cmobjects.filter(region=selected_region_id)

    # If both region and BLF are selected, filter Aggregator details
	if selected_region_id and selected_blf_id:
		selected_blf = BlfProfile.objects.get(pk=selected_blf_id)
		region_name = Region.objects.filter(id=selected_region_id).first()
		blf_name = selected_blf.entity_unit
		blf_username = selected_blf.user.username
		# Retrieve GrowerProfiles associated with the selected BlfProfile
		grower_profiles = GrowerProfile.objects.filter(associated_entity=selected_blf)
		for grower_profile in grower_profiles:
			# Retrieve Gardens associated with each GrowerProfile
		
			gardens = Gardens.objects.filter(grower=grower_profile)
			for garden in gardens:
				# Access land_type and production_area associated with each Gardens entry
				land_type = garden.land_type
				production_area = garden.production_area
				# Perform operations with land_type and production_area as needed
				# For example, print or store these values in a list/dictionary
				print(f"Land Type: {land_type}, Production Area: {production_area}")
				# Append a tuple containing GrowerProfile and associated land_type and production_area
				result.append((grower_profile, land_type, production_area, serial_number))
				serial_number += 1

	# Pagination logic
	paginator = Paginator(result, 10)
	page = request.GET.get('page', 1)
	try:
		result = paginator.page(page)
	except PageNotAnInteger:
		result = paginator.page(1)
	except EmptyPage:
		result = paginator.page(paginator.num_pages)
	# serial_number_offset = (result.number - 1) * paginator.per_page
	# for index, item in enumerate(result, start=1):
	# 		item.serial_number = index + serial_number_offset				
	total_items = paginator.count
    # Calculate serial number within the result list
	# serial_number = paginator.count * (paginator.number - 1) if paginator.count > 0 else 0

	# for row in result:
	# 	# Increment serial number within the loop
	# 	serial_number += 1
	# 	row['serial_number'] = serial_number



	context = {
	    "region_list": regions,
	    "region_name": region_name,
	    "blf_name": blf_name,
	    "blf_username": blf_username,
	    "result": result,
	    "total_items": total_items,
	    "selected_blf_id": selected_blf_id,
	    "selected_region_id": selected_region_id,
		'entity_id':selected_blf_id,
		'region':selected_region_id,
		"blf_list" : blf_list,
	}    
	return CommonMixin.render(request, "reports/report_grower_register_list.html", context)






def report_grower_register_excel(request):
	try:
		selected_region_id = request.GET.get('selected_region_id', None)
		selected_blf_id = request.GET.get('selected_blf_id', None)
	
		result = []
		selected_blf=None
	
		# Fetch Aggregator details
		if selected_region_id is not None and selected_blf_id is not None:
			try:
				# Convert IDs to integers
				selected_region_id = int(selected_region_id)
				selected_blf_id = int(selected_blf_id)
				# Fetch the selected BLF and region
				selected_blf = BlfProfile.objects.get(pk=selected_blf_id)
				
				# Filter Aggregator details
				result = GrowerProfile.objects.filter(associated_entity=selected_blf)
			except (ValueError, BlfProfile.DoesNotExist, Region.DoesNotExist):
				# Handle exceptions or incorrect IDs here
				pass
				
		# Create a workbook and add a worksheet
		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = "Grower Register Report"
		# Define headers for the Excel file
		# Define headers for the Excel file
		headers = ["SL No.", "Grower Tracetea ID", "Grower Name", "Type", "Entity",  "Aggregator", "Father's name", "Date of Birth", "Age", "Gender", "Address", "Region", "State", "District", "Village/Town", "Post Office", "Pin Code", "Mobile No.", "Aadhar No.", "Voter ID", "Tea Board ID", "Total Male worker", "Total Female worker",\
			  "Estimated Production of Green Leaves per annum", "Estimated Production of Made Tea per annum", "Land Type", "Production Area"]
        # Add headers to the worksheet
		for col_num, header in enumerate(headers, 1):
			col_letter = get_column_letter(col_num)
			ws[f"{col_letter}1"] = header
		# Populate data into the worksheet
		row_num = 2  # Start from the second row
		for index, item in enumerate(result, start=1):
			ws[f"A{row_num}"] = index  # SL No.
			ws[f"B{row_num}"] = item.user.username or "NA"
			ws[f"C{row_num}"] = item.name or "NA"
			ws[f"D{row_num}"] = item.grower_type or "NA"
			ws[f"E{row_num}"] = ', '.join([entity.user.username for entity in item.associated_entity.all()]) if item.associated_entity.all() else "NA"
			# ws[f"F{row_num}"] = ', '.join([entity.user.username for entity in item.associated_entity.all()]) if item.associated_entity.all() else "NA"
			ws[f"F{row_num}"] = ', '.join([aggregator.user.username for aggregator in item.associated_aggregator.all()]) if item.associated_aggregator.all() else "NA"
			ws[f"G{row_num}"] = item.father_name or "NA"
			ws[f"H{row_num}"] = str(item.date_of_birth) or "NA"
			ws[f"I{row_num}"] = item.age or "NA"
			ws[f"J{row_num}"] = item.gender or "NA"
			ws[f"K{row_num}"] = item.address or "NA"
			ws[f"L{row_num}"] = item.region.region_name if item.region else "NA"
			ws[f"M{row_num}"] = item.state.name if item.state else "NA"
			ws[f"N{row_num}"] = item.district.name if item.district else "NA"
			ws[f"O{row_num}"] = item.village_or_town or "NA"
			ws[f"P{row_num}"] = item.postoffice or "NA"
			ws[f"Q{row_num}"] = item.pincode or "NA"
			ws[f"R{row_num}"] = item.mobile_number or "NA"
			ws[f"S{row_num}"] = item.aadhar_no or "NA"
			ws[f"T{row_num}"] = item.voter_id or "NA"
			ws[f"U{row_num}"] = item.tea_board_id or "NA"
			ws[f"V{row_num}"] = item.total_male_worker or "NA"
			ws[f"W{row_num}"] = item.total_female_worker or "NA"
			ws[f"X{row_num}"] = item.estimated_production_of_green_tea or "NA"
			ws[f"Y{row_num}"] = item.estimated_production_of_made_tea or "NA"
			 # Fetch land_type and production_area from Garden model
			garden_data = Gardens.objects.filter(grower=item).first()
	
			if garden_data:
				print(garden_data.production_area)
                # Handle the data type conversion issue
				# try:
				if garden_data.land_type:
					ws[f"Z{row_num}"] = garden_data.land_type.name
				else:
					ws[f"Z{row_num}"] = "NA"			
				if garden_data.production_area:
					ws[f"AA{row_num}"] = garden_data.production_area 
				else:
					ws[f"AA{row_num}"] = "NA"			
				# except Exception as e:
				# 	# Log or handle the specific exception if needed
				# 	ws[f"Z{row_num}"] = "NA"
				# 	ws[f"AA{row_num}"] = "NA"
			else:
				ws[f"Z{row_num}"] = "NA"
				ws[f"AA{row_num}"] = "NA"
	
			row_num += 1  # Move to the next row for the next data item
	
			# Create an HTTP response with the Excel file as an attachment
			response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
			blf_name_for_filename = selected_blf.entity_unit.replace(' ', '_') if selected_blf else 'grower_lists'
			filename = f"{blf_name_for_filename}_grower_lists.xlsx"
			response['Content-Disposition'] = f'attachment; filename="{filename}"'
		# Save the workbook content to the response
		wb.save(response)
		return response
	except (ValueError, BlfProfile.DoesNotExist, Region.DoesNotExist) as e:
		error_message = "Something went wrong."
		return render(request, 'error.html', {'error_message': error_message}) 
	except Exception as e:
		error_message = "Something went wrong."
		return render(request, 'error.html', {'error_message': error_message}) 





def entity_register_list_blf_report(request):
	regions = Region.objects.all()
	region_id = request.GET.get('region')
	entity_type = request.GET.get('entity_type')
	# Fetch all entities initially


	print("entity_type", entity_type)

	# If a region is selected, filter entities based on that region
	if region_id and entity_type == "blf":
		item_list = BlfProfile.cmobjects.filter(region_id=region_id).order_by('-id')
	elif region_id and entity_type == "estate":
		item_list = EstateProfile.cmobjects.filter(region_id=region_id).order_by('-id')
	else:
		item_list = ""

	# Paginate the result
	paginator = Paginator(item_list, 10)
	page = request.GET.get('page', 1)

	try:
		blf_list_page = paginator.page(page)
	except PageNotAnInteger:
		blf_list_page = paginator.page(1)
	except EmptyPage:
		blf_list_page = paginator.page(paginator.num_pages)

	# Calculate the serial number adjusted for pagination
	serial_number_offset = (blf_list_page.number - 1) * paginator.per_page
	for index, item in enumerate(blf_list_page, start=1):
		item.serial_number = index + serial_number_offset

	context = {
		"blf_list": blf_list_page,
		"total_items": paginator.count,
		"regions": regions,
		"selected_region_id": region_id,
		"entity_type" : entity_type,
	}

	return CommonMixin.render(request, "reports/reports_blf_register_list.html", context)
from rest_framework import status

class DownloadVistingCardPDF(APIView):
	authentication_classes = (TokenAuthentication,)
	permission_classes = (IsAuthenticated,)
	# permission_classes = (AllowAny,)
	def get(self, request, *args, **kwargs):
		user_pk = self.request.query_params.get('user_id', None)
		user_type = self.request.query_params.get('user_type', None)

		try:
			if not user_pk or not user_type:
				raise ValueError('Invalid parameters')

			if user_type.lower() == 'grower':
				details = GrowerProfile.objects.filter(user_id=user_pk).first()
			elif user_type.lower() == 'aggregator':
				details = AggregatorProfile.objects.filter(user_id=user_pk).first()
			else:
				return Response({
				'msg': 'Invalid user type',
				'status': status.HTTP_400_BAD_REQUEST,
				'request_status': 0
			}, status=status.HTTP_400_BAD_REQUEST)

			if not details:
				return Response({
				'msg': 'User details not found',
				'status': status.HTTP_404_NOT_FOUND,
				'request_status': 0
			}, status=status.HTTP_200_OK)

			user_type = details.profile_type.name
			user_id = details.user

			qr_code = {
				'userid': details.user.id,
				'name': details.name,
				'id': details.id,
				'username': details.user.username
			}

			qr_code_image_path = generate_qr_code(qr_code, user_id, user_type)
			media_url = settings.MEDIA_URL.rstrip('/')  # Remove trailing slash if present
			absolute_qr_code_image_url = self.request.build_absolute_uri(qr_code_image_path.replace(settings.MEDIA_DIR, media_url).replace("\\", "/"))

			if details.photo:
				details_photo_url = self.request.build_absolute_uri(details.photo.url)
			else:
				details_photo_url = None

			context = {
				'details': details,
				'user_type': user_type,
				'details_photo_url': details_photo_url,
				'absolute_qr_code_image_url': absolute_qr_code_image_url
			}

			template_path = 'visiting_download.html'
			html = get_template(template_path).render(context)

			response = HttpResponse(content_type='application/pdf')
			response['Content-Disposition'] = 'attachment; filename="visiting_card.pdf"'

			# Generate PDF from HTML content
			pisa_status = pisa.CreatePDF(html, dest=response)
			if pisa_status.err:
				return Response({
					'msg': 'pdf generation failed',
					'status': status.HTTP_400_BAD_REQUEST,
					'request_status': 0
				}, status=status.HTTP_400_BAD_REQUEST)

			return response

		except Exception as e:
			return Response({
				'msg': 'An error occurred while retrieving visiting cad',
				'status': status.HTTP_400_BAD_REQUEST,
				'request_status': 0
			}, status=status.HTTP_400_BAD_REQUEST)
class VisitingCardGenerateWebAPIViewUrl(APIView):
	authentication_classes = (TokenAuthentication,)
	permission_classes = (IsAuthenticated,)
	def get(self, request, *args, **kwargs):
		try:
			user_type = self.request.query_params.get('user_type', None)
			id = self.request.query_params.get('id', None)

			if user_type == 'grower':
				url = reverse('user_profile:visiting_card_web', args=[user_type, id])

				print(url)
				absolute_url = request.build_absolute_uri(url)
				response_data = {
					'url': absolute_url,
				}
			elif user_type == 'aggregator':
				url = reverse('user_profile:visiting_card_web', args=[user_type, id])

				absolute_url = request.build_absolute_uri(url)
				response_data = {
					'url': absolute_url,
				}
			else:
				return Response({
					'msg': 'Invalid user type',
					'status': status.HTTP_400_BAD_REQUEST,
					'request_status': 0
				}, status=status.HTTP_400_BAD_REQUEST)

			return Response({
				'results': {'Data': response_data},
				'msg': 'Successful',
				'status': status.HTTP_200_OK,
				'request_status': 1
			})

		except Exception as e:
			return Response({
				'msg': str(e),
				'status': status.HTTP_400_BAD_REQUEST,
				'request_status': 0
			}, status=status.HTTP_400_BAD_REQUEST)
class VistingCardWebView(APIView):
	permission_classes = [AllowAny]

	def get(self, request, user_type,id):
		"""For APP """
		try:
			if user_type =='grower':
				details=GrowerProfile.objects.filter(id=id).first()
				print(details)
				if details:
					user_type=details.profile_type.name
					# qr_code=GrowerProfile.objects.filter(user_id=user).first()	
					user_id=details.user
					qr_code = {
						'userid': details.user.id,
						'name': details.name,
						'id': details.id,
						'username': details.user.username
					}
					qr_code_image_path = generate_qr_code(qr_code,user_id ,user_type)

					media_url = settings.MEDIA_URL.rstrip('/')  # Remove trailing slash if present
					absolute_qr_code_image_url = request.build_absolute_uri(qr_code_image_path.replace(settings.MEDIA_DIR, media_url).replace("\\", "/"))
					if details is not None:
						if details.photo:
							details_photo_url = request.build_absolute_uri(details.photo.url)
						else:
							details_photo_url = None
					else:
						details_photo_url = None
					context ={
						'details':details,
						'user_type':user_type,
						'details_photo_url':details_photo_url,
						'absolute_qr_code_image_url':absolute_qr_code_image_url

					}
		
				return render(request,"visiting.html", context)	
			elif user_type == 'aggregator':
				details=AggregatorProfile.objects.filter(id=id).first()
				if details:
					user_type=details.profile_type.name
					# qr_code=GrowerProfile.objects.filter(user_id=user).first()	
					user_id=details.user
					# qr_code= "Aggregator ID : " + str(details.user) + "\n" + "Name : " + str(details.name)
					qr_code = {
						'userid': details.user.id,
						'name': details.name,
						'id': details.id,
						'username': details.user.username
					}
					qr_code_image_path = generate_qr_code(qr_code,user_id ,user_type)

					media_url = settings.MEDIA_URL.rstrip('/')  # Remove trailing slash if present
					absolute_qr_code_image_url = request.build_absolute_uri(qr_code_image_path.replace(settings.MEDIA_DIR, media_url).replace("\\", "/"))
					if details is not None:
						if details.user_file:
							details_photo_url = request.build_absolute_uri(details.user_file.url)
						else:
							details_photo_url = None
					else:
						details_photo_url = None
					context ={
						'details':details,
						'user_type':user_type,
						'details_photo_url':details_photo_url,
						'absolute_qr_code_image_url':absolute_qr_code_image_url
				}
				return render(request,"visiting.html", context)			
			else:
				messages.error(request, "Grower details not found")
				# Return an appropriate response in case details are not found
				return HttpResponse("Grower details not found")
		except (ValueError) as e:
			error_message = "Something went wrong."
			return render(request, 'error.html', {'error_message': error_message}) 
		except Exception as e:
			error_message = "Something went wrong."
			return render(request, 'error.html', {'error_message': error_message}) 
		
#######  NEW USER PROFILE API #######
# class UserProfileMainAPIView(APIView):
# 	authentication_classes = (TokenAuthentication,)
# 	permission_classes = (IsAuthenticated,)
# 	def get(self, request, *args, **kwargs):
# 		default_type = 'grower'
# 		user_type = self.request.query_params.get('user_type', None)
# 		if not user_type:
# 			raise APIException({'request_status': 0, 'msg': "User type required"})
# 		search = {}
# 		search = custom_filters(self.request, search, ['user_type'])
# 		# id = self.request.query_params.get('id', None)
# 		all = request.query_params.get('all', None)
# 		order_by = request.query_params.get('order_by', '-id')
# 		created_by__full_name__icontains = request.query_params.get('created_by__full_name__icontains', None)
# 		USER_TYPE_CONFIG = {
# 			"grower": {
# 				"model": GrowerProfile,
# 				"serializer": GrowerProfileSerializer,
# 				"select_related" : ['state', 'user', 'region', 'district', 'profile_type'],
# 				"prefetch_related" : ["associated_aggregator", "associated_entity"],
# 				"values": [
# 					'id','user__username', 'is_tcms_user', 'tcms_supplier_code',
# 					'name', 'age', 'gender', 'date_of_birth', 'grower_type',
# 					'region__region_name', 'state__name', 'district__name',
# 					'region', 'state', 'district', 'production_area',
# 					'email', 'mobile_number', 'aadhar_no', 'address', 'tea_board_id', 'associated_aggregator', 
# 					'associated_entity', 'poi_type', 'poi_id',
# 				],
# 			},
# 			"aggregator": {
# 				"model": AggregatorProfile,
# 				"serializer": AggregatorProfileSerializer,
# 				"select_related" : ['state', 'user', 'region', 'district', 'profile_type'],
# 				"prefetch_related" : [],
# 				"values": [],
# 			},
# 			"blf": {
# 				"model": BlfProfile,
# 				"serializer": BlfProfileSerializer,
# 				"select_related" : ['state', 'user', 'region', 'district', 'profile_type'],
# 				"prefetch_related" : [],
# 				"values": [],
# 			},
# 		}
# 		config = USER_TYPE_CONFIG.get(user_type)
# 		if not config:
# 			raise APIException({'request_status': 0, 'msg': "Invalid user_type"})
# 		model = config["model"]
# 		values = config["values"]   

# 		_list = model.cmobjects.select_related(*config["select_related"]).prefetch_related(*config["prefetch_related"]).values(*values).distinct()
# 		_list = _list.filter(*search).order_by(*str(order_by).split(","))
# 		if all == 'true':
# 			return Response({'results': list(_list)})
# 		page_size = int(request.query_params.get('page_size', settings.MIN_PAGE_SIZE))
# 		paginator = Paginator(_list, page_size)
# 		page_number = request.query_params.get('page', 1)
# 		page = paginator.get_page(page_number)
# 		return Response({
# 			'count': paginator.count,
# 			'next': page.next_page_number() if page.has_next() else None,
# 			'previous': page.previous_page_number() if page.has_previous() else None,
# 			'results': {
# 				'data': list(page.object_list),
# 			},
# 		})
	
class UserProfileMainAPIView(APIView):
	authentication_classes = (TokenAuthentication,)
	permission_classes = (IsAuthenticated,)

	def get(self, request, *args, **kwargs):
		user_type = self.request.query_params.get('user_type', None)
		if not user_type:
			raise APIException({'request_status': 0, 'msg': "User type required"})
		search ={}
		search['user__is_active'] = True
		search = custom_filters(self.request, search, ['user_type'])
		all = request.query_params.get('all', None)
		order_by = request.query_params.get('order_by', '-id')
		USER_TYPE_CONFIG = {
			"grower": {
				"model": GrowerProfile,
				"select_related": ['state', 'user', 'region', 'district', 'profile_type'],
				"prefetch_related": ["associated_aggregator", "associated_entity"],
				"values": [
					'id','user__username', 'is_tcms_user', 'tcms_supplier_code',
					'name', 'age', 'gender', 'date_of_birth', 'grower_type',
					'region__region_name', 'state__name', 'district__name',
					'region', 'state', 'district', 'production_area',
					'email', 'mobile_number', 'aadhar_no', 'address', 'tea_board_id',
					'associated_aggregator', 'associated_entity', 'poi_type', 'poi_id',
					# annotated fields
					'total_quantity_plucked', 'total_area_plucked',
				],
			},
			"aggregator": {
				"model": AggregatorProfile,
				"select_related": ['state', 'user', 'region', 'district', 'profile_type'],
				"prefetch_related": [],
				"values": [],
			},
			"blf": {
				"model": BlfProfile,
				"select_related": ['state', 'user', 'region', 'district', 'profile_type'],
				"prefetch_related": [],
				"values": [],
			},
		}
		
		config = USER_TYPE_CONFIG.get(user_type)
		if not config:
			raise APIException({'request_status': 0, 'msg': "Invalid user_type"})

		model = config["model"]
		values = config["values"]

		# Subqueries for totals
		plucking_qs = PluckingData.cmobjects.filter(grower=OuterRef('pk'))
		# Cast the CharField to IntegerField before summing
		total_quantity = (
			plucking_qs
			.annotate(q_int=Cast('quantity_plucked', IntegerField()))
			.values('grower')
			.annotate(total=Sum('q_int'))
			.values('total')[:1]
		)
		total_area = (
			plucking_qs
			.annotate(a_int=Cast('area_plucked', IntegerField()))
			.values('grower')
			.annotate(total=Sum('a_int'))
			.values('total')[:1]
		)
		_list = (
			model.cmobjects
			.select_related(*config["select_related"])
			.prefetch_related(*config["prefetch_related"])
			.annotate(
				total_quantity_plucked=Coalesce(Subquery(total_quantity), Value(0)),
				total_area_plucked=Coalesce(Subquery(total_area), Value(0)),
			)
			.values(*values).filter(*search).order_by(*str(order_by).split(",")).distinct('id')  
		)
		_list = _list.filter(*search).order_by(*str(order_by).split(",")).distinct()

		# _list = _list.filter(*search).order_by(*str(order_by).split(","))
		if all == 'true':
			return Response({'results': list(_list)})

		page_size = int(request.query_params.get('page_size', settings.MIN_PAGE_SIZE))
		paginator = Paginator(_list, page_size)
		page_number = request.query_params.get('page', 1)
		page = paginator.get_page(page_number)

		return Response({
			'count': paginator.count,
			'next': page.next_page_number() if page.has_next() else None,
			'previous': page.previous_page_number() if page.has_previous() else None,
			'results': {
				'data': list(page.object_list),
			},
		})





class UserProfileAPIView(APIView):
	authentication_classes = (TokenAuthentication,)
	permission_classes = (IsAuthenticated,)
	USER_TYPE_CONFIG = {
		"grower": {
			"model": GrowerProfile,
			"serializer" : GrowerProfileSerializer,
		},
		"aggregator": {
			"model": AggregatorProfile,
			"serializer" : AggregatorProfileSerializer,
		},
		"blf": {
			"model": BlfProfile,
			"serializer" : BlfProfileSerializer,
		},
	}
	def get_config(self, user_type):
		config = self.USER_TYPE_CONFIG.get(user_type)
		if not config:
			raise ValidationError({"msg": "Invalid user_type"})
		return config
	def get(self, request):
		user_type = self.request.query_params.get('user_type', None)
		if not user_type:
			raise APIException({'request_status': 0, 'msg': "User type required"})
		id = self.request.query_params.get('id', None)
		all = self.request.query_params.get('all', None)
		order_by = self.request.query_params.get('order_by', '-id')
		search = {}
		search['user__is_active'] = True
		search = custom_filters(self.request, search, ['user_type'])
		config = self.get_config(user_type)
		model = config["model"]
		serializer_class = config["serializer"]
		if id:
			list_data = model.cmobjects.filter(id=id).first()
			serializer = serializer_class(list_data)
			return Response(serializer.data)
		
		list_data = model.cmobjects.filter(*search).order_by(*str(order_by).split(","))
		if all == 'true':
			serializer = serializer_class(list_data, many=True)
			return Response({'results': serializer.data})
		
		page_size = int(request.query_params.get('page_size', settings.MIN_PAGE_SIZE))
		paginator = Paginator(list_data, page_size)
		page_number = self.request.query_params.get('page', 1)
		page = paginator.get_page(page_number)
		serializer = serializer_class(page, many=True)
		return Response({
			'count': paginator.count,
			'next': page.next_page_number() if page.has_next() else None,
			'previous': page.previous_page_number() if page.has_previous() else None,
			'results': serializer.data,
		})
	def post(self, request):
		request.data['created_by'] = request.user.id
		user_type = self.request.query_params.get('user_type', None)
		config = self.get_config(user_type)
		model = config["model"]
		serializer_class = config["serializer"]

		serializer = serializer_class(data=request.data, context={'request': request})
		if serializer.is_valid():
			try:
				serializer.save()
			except Exception as e:
				error_message = str(e.args[0]) if e.args else str(e)
				raise APIException({'request_status': 0, 'msg': error_message})
			return Response(
				{'results': {'Data': serializer.data,},
					'msg': 'Successfully created',
					'status': status.HTTP_201_CREATED,
					"request_status": 1})
		else:
			error_message = next(iter(serializer.errors.values()))[0]
			raise APIException({
				"request_status": 0,
				"msg": error_message
			})
		raise APIException({'request_status': 0, 'msg': serializer.errors})
	def put(self, request):
		method = request.query_params.get('method')
		id = request.query_params.get('id')
		user_type = self.request.query_params.get('user_type', None)
		if not method:
			raise APIException({'request_status': 0, 'msg': "Method parameter is required"})
		if not id:
			raise APIException({'request_status': 0, 'msg': "ID parameter is required"})
		config = self.get_config(user_type)
		model = config["model"]
		serializer_class = config["serializer"]
		with transaction.atomic():
			if method.lower() == 'edit':
				supply = model.cmobjects.filter(pk=id).first()
				if not supply:
					raise APIException({'request_status': 0, 'msg': "Users not found"})

				serializer = serializer_class(
					supply, data=request.data, context={'request': request}
				)
				if serializer.is_valid():
					# try:
					serializer.save()
					# except Exception as e:
					# 	error_message = str(e.args[0]) if e.args else str(e)
					# 	raise APIException({'request_status': 0, 'msg': error_message})
					return Response({
						'results': {'Data': serializer.data},
						'msg': "Successfully updated",
						'status': status.HTTP_202_ACCEPTED,
						'request_status': 1
					})
				else:
					raise APIException({'request_status': 0, 'msg': serializer.errors})
				
			elif method.lower() == 'delete':
				supply = model.cmobjects.filter(pk=id).first()
				if not supply:
					raise APIException({'request_status': 0, 'msg': "User not found"})
				supply.is_deleted = True
				supply.updated_by_id = request.user.id
				supply.save()
				return Response({
					'results': {'data': model.cmobjects.filter(pk=id).values()},
					'msg': 'Successfully deleted',
					'request_status': 1
				})
			else:
				raise APIException({'request_status': 0, 'msg': "Invalid method parameter"})

class ValidateTcmsUsers2(APIView):
	authentication_classes = (TokenAuthentication,)
	permission_classes = (IsAuthenticated,)

	def post(self, request, *args, **kwargs):
		results = []
		updated_count = 0
		api_url = os.getenv("TCMS_ENTITY_URL")
		if not api_url:
			return Response({"error": "TCMS_ENTITY_URL not configured"}, status=500)

		try:
			response = requests.get(api_url, verify=certifi.where(), timeout=30)
			response.raise_for_status()
			json_data = response.json()
		except Exception as e:
			return Response({"error": f"Failed to fetch TCMS data: {str(e)}"}, status=502)

# @login_required
def generate_excel_blf(results):
	output_dir = os.path.join(settings.MEDIA_ROOT, "excel", "validate", "tcms_data")
	os.makedirs(output_dir, exist_ok=True) 
	output_file = os.path.join(output_dir, f"entity_validate_data.xlsx")
	df = pd.DataFrame(results)
	df.to_excel(output_file, index=False)

def sanitize_float(value):
    """Safely convert a value to float, returning None for NaN/inf/invalid."""
    try:
        if value is None:
            return None
        if isinstance(value, (float, np.floating)):
            if math.isnan(value) or math.isinf(value):
                return None
            return float(value)
        return float(value)
    except Exception:
        return None



class ValidateTcmsBlfUsers(APIView):
    authentication_classes = (TokenAuthentication,)
    permission_classes = (IsAuthenticated,)

    def post(self, request, *args, **kwargs):

        results = []
        updated_count = 0
        matched_users_ids = []
        unmatched_users_ids = []

        api_url = os.getenv("TCMS_ENTITY_URL")

        if not api_url:
            return Response(
                {"error": "TCMS_ENTITY_URL not configured"},
                status=500
            )

        # --------------------------------------------------
        # STEP 1: Fetch TCMS Data
        # --------------------------------------------------
        try:
            response = requests.get(
                api_url,
                verify=certifi.where(),
                timeout=30
            )
            response.raise_for_status()
            json_data = response.json()
        except Exception as e:
            return Response(
                {"error": f"Failed to fetch TCMS data: {str(e)}"},
                status=502
            )

        api_entities = []

        for entity in json_data.get("data", []):
            entity_details = entity.get("entity_details") or {}

            api_entities.append({
                "unit_name": str(entity.get("name", "") or "").strip(),
                "user_name": str(entity_details.get("user_name", "") or "").strip(),
                "state": str(entity.get("state", "") or "").strip(),
                "district": str(entity.get("district", "") or "").strip(),
            })

        if not api_entities:
            return Response({
                "count": 0,
                "updated_count": 0,
                "results": [],
                "message": "No TCMS data found"
            })

        api_df = pd.DataFrame(api_entities)

        # --------------------------------------------------
        # STEP 2: Fetch Local Profiles
        # --------------------------------------------------
        profiles = BlfProfile.cmobjects.all().values(
            "id",
            "entity_unit",
            "user__username",
            "state__name",
            "district__name",
        )

        profiles_df = pd.DataFrame(list(profiles))

        if profiles_df.empty:
            return Response({
                "count": 0,
                "updated_count": 0,
                "results": [],
                "message": "No BLF profiles found"
            })

        # Clean profile fields
        profiles_df["entity_unit"] = profiles_df["entity_unit"].astype(str).str.strip()
        profiles_df["user__username"] = profiles_df["user__username"].astype(str).str.strip()
        profiles_df["state__name"] = profiles_df["state__name"].astype(str).str.strip()
        profiles_df["district__name"] = profiles_df["district__name"].astype(str).str.strip()

        # --------------------------------------------------
        # STEP 3: FUZZY MATCHING (No ID Matching)
        # --------------------------------------------------
        fuzzy_matches = fuzzymatcher.link_table(
            api_df,
            profiles_df,
            left_on=[
                "unit_name",
                "user_name",
                "state",
                "district",
            ],
            right_on=[
                "entity_unit",
                "user__username",
                "state__name",
                "district__name",
            ],
        )

        fuzzy_matches = (
            fuzzy_matches
            .sort_values("match_score", ascending=False)
            .drop_duplicates("unit_name")
        )

        fuzzy_matches["match_score"] = fuzzy_matches["match_score"].fillna(0.0)

        # --------------------------------------------------
        # STEP 4: UPDATE MATCHED USERS
        # --------------------------------------------------
        with transaction.atomic():

            for _, row in fuzzy_matches.iterrows():

                profile_id = row.get("id")
                raw_score = row.get("match_score", 0.0)

                if pd.isna(raw_score):
                    raw_score = 0.0

                match_score = round(float(raw_score) * 100, 2)

                # 🔥 IMPORTANT: Proper threshold
                # if profile_id and match_score >= 70:
					
                #     # update_count = BlfProfile.cmobjects.filter(
                #     #     id=profile_id
                #     # ).update(
                #     #     is_tcms_user=True
                #     # )

                #     # if update_count:
                #     #     updated_count += 1
                #     #     matched_users_ids.append(profile_id)
                #     #     user_update = "true"
                #     # else:
                #     #     user_update = "false"
                # else:
                #     user_update = "false"
                #     if profile_id:
                #         unmatched_users_ids.append(profile_id)

                results.append({
                    "profile_id": profile_id,
                    "tracetea_entity_unit": row.get("entity_unit", ""),
                    "tcms_entity_unit": row.get("unit_name", ""),
                    # "match_score": match_score,
                    # "user_update": user_update
                })

        # --------------------------------------------------
        # FINAL RESPONSE
        # --------------------------------------------------
        return Response({
            "count": len(results),
            "updated_count": updated_count,
            "results": results,
            "matched_users_ids": list(set(matched_users_ids)),
            "unmatched_users_ids": list(set(unmatched_users_ids)),
        })







class ValidateTcmsBlfUsers2(APIView):
	authentication_classes = (TokenAuthentication,)
	permission_classes = (IsAuthenticated,)
	def post(self, request, *args, **kwargs):
		results = []
		updated_count = 0
		api_url = os.getenv("TCMS_ENTITY_URL")
		if not api_url:
			return Response({"error": "TCMS_ENTITY_URL not configured"}, status=500)
		try:
			response = requests.get(api_url, verify=certifi.where(), timeout=30)
			response.raise_for_status()
			json_data = response.json()
		except Exception as e:
			return Response({"error": f"Failed to fetch TCMS data: {str(e)}"}, status=502)
		api_entities = []
		for entity in json_data.get("data", []):
			entity_details = entity.get("entity_details") or {}
			api_entities.append({
				"unit_name": entity.get("name", ""),
				"entity_name" : entity_details.get("name", ""),
				"user_name": entity_details.get("user_name", ""),
				"state": entity.get("state", ""),
				"district": entity.get("district", ""),
				"region_name": entity_details.get("region_name", ""),
				"whats_app_no" : entity_details.get("whats_app_no", ""),	
				"head_contact_name": entity.get("head_contact_name", ""),
				"head_contact_phone": entity.get("head_contact_phone", ""),
				"head_contact_email": entity.get("head_contact_email", ""),
				"tcmo_no" : entity.get("tcmo_no", ""),
				"unit_id": entity.get("unit_id", ""),
				"password": entity_details.get("actual_password", ""),
			})
		api_df = pd.DataFrame(api_entities)
		profiles = BlfProfile.cmobjects.all().values(
			"id", "entity_unit", "user__username", "region__region_name",
			"state__name", "district__name", "ho_contact_person",
			"ho_contact_number", "ho_contact_email", "profile_type",
		)
		profiles_df = pd.DataFrame(list(profiles))
		matched = fuzzymatcher.link_table(
			api_df,
			profiles_df,
			left_on=[
				"unit_name", "user_name", "state", "district",
				"unit_id", 
			],
			right_on=[
				"entity_unit", "user__username", "state__name", "district__name", "id",
			],
		)
		matched = matched.sort_values("match_score", ascending=False).drop_duplicates("unit_name")
		matched["match_score"] = matched["match_score"].fillna(0.0)
		matched_users_ids = []
		unmatched_users_ids = []
		for _, row in matched.iterrows():
			raw_score = row.get("match_score", 0.0)
			# print("raw_score ####", raw_score)
			if pd.isna(raw_score) or raw_score in ["nan", None]:
				raw_score = 0.0
			match_score = round(float(raw_score) * 100, 2)
			if str(row.get("unit_name", "")).strip().lower() == str(row.get("entity_unit", "")).strip().lower():
				match_score = 150.0
			elif (
				str(row.get("unit_name", "")).strip().lower() == str(row.get("entity_unit", "")).strip().lower()
				and str(row.get("state", "")).strip().lower() == str(row.get("state__name", "")).strip().lower()
				and str(row.get("district", "")).strip().lower() == str(row.get("district__name", "")).strip().lower()
			):
				match_score = 120.0
			user_update = "false"
			if match_score >= 0:
				profile_id = row.get("id", None)  
				update_data = BlfProfile.cmobjects.filter(id=profile_id).update(
					is_tcms_user=True,
					# tcms_unit_id=row.get("unit_id", ""),
					# tcmo_no=row.get("tcmo_no", ""),
				)
				if update_data:
					updated_count += 1
					user_update = "true"
					matched_users_ids.append(row.get("id"))
			else:
				unmatched_users_ids.append(row.get("id"))

			results.append({
				"tracetea_username": row.get("user__username", ""),
				"tcms_username": row.get("user_name", ""),
				"tracetea_entity_unit": row.get("entity_unit", ""),
				"tcms_entity": row.get("unit_name", ""),
				"tracetea_state": row.get("state__name", ""),
				"tcms_state": row.get("state", ""),
				"tracetea_district": row.get("district__name", ""),
				"tcms_district": row.get("district", ""),
				"match_score": sanitize_float(match_score),
				"user_update" : user_update
			})
			# generate_excel_blf(results)
		return Response({
			"count": len(results),
			"updated_count": updated_count,
			"results": results,
			"unmatched_users_ids" : set(unmatched_users_ids),
			"matched_users_ids" : set(unmatched_users_ids),
		})

class ValidateTcmsSupplierUsers(APIView):
	authentication_classes = (TokenAuthentication,)
	permission_classes = (IsAuthenticated,)
	
	def post(self, request, *args, **kwargs):
		results = []
		updated_count = 0
		supplier_mode = self.request.query_params.get('supplier_mode', "LEAD FARMER")
		blf_unit_ids_param = self.request.query_params.get("blf_unit_ids", None)
		if blf_unit_ids_param:
			unit_ids = [int(x.strip()) for x in blf_unit_ids_param.split(",") if x.strip().isdigit()]
		else:
			# tcms_unit_id = list(BlfProfile.cmobjects.values_list("tcms_unit_id", flat=True))
			# unit_ids = [int(x) for x in tcms_unit_id if x is not None]
			unit_ids = [1474,1258]

		api_url = os.getenv("TCMS_SUPPLIER_URL")
		if not api_url:
			return Response({"error": "TCMS_SUPPLIER_URL not configured"}, status=500)
		api_suppliers = []

		for unit_id in unit_ids:
			payload = {
				"unit_id": unit_id,
				"filters": [
					{"key": "supply_mode", "value": supplier_mode},
					# {"key": "land_type", "value": "Own"}, DIRECT,
				],
			}
			try:
				response = requests.post(
					api_url,
					json=payload,
					verify=certifi.where(),
					timeout=30,
				)
				response.raise_for_status()
				json_data = response.json()
			except Exception as e:
				return Response({"error": f"Failed to fetch TCMS data: {str(e)}"}, status=502)

			for supplier in json_data.get("data", []):
				api_suppliers.append({
					"unit_id": unit_id,
					"supplier_code": supplier.get("supplier_code", ""),
					"supplier_name": supplier.get("supplier_name", ""),
					"supply_mode": supplier.get("supply_mode", ""),
					"supplier_type": supplier.get("supplier_type", ""),
					"region": supplier.get("region", ""),
					"state": supplier.get("state", ""),
					"district": supplier.get("district", ""),
					"address": supplier.get("address", ""),
					"sex": supplier.get("sex", ""),
					"land_type": supplier.get("land_type", ""),
					"annual_production": supplier.get("annual_production", ""),
					"phone_number": supplier.get("phone_number", ""),
					"male_worker" : supplier.get("male_worker", ""),
					"female_worker" : supplier.get("female_worker", ""),
				})

		api_df = pd.DataFrame(api_suppliers)
		supply_mode = "LEAD FARMER"

		if supply_mode == "LEAD FARMER":
			print("AGG DATA HERE ################")
			model_name = AggregatorProfile
			profiles = model_name.cmobjects.all().values(
				"id", "name", "user__username", "region__region_name",
				"state__name", "district__name", "aggregator_type",
				"address", "mobile_number",
			)
			left_on_data = [
				"supplier_name", "region", "state", "district",
				"address", "phone_number", "supplier_code", 
			]
			right_on_data = [
				"name", "region__region_name", "state__name", "district__name",
				"address", "mobile_number", "aggregator_type", "id", "user__username",
			]
		else:
			model_name = GrowerProfile
			profiles = model_name.cmobjects.all().values(
				"id", "name", "user__username", "region__region_name",
				"state__name", "district__name", "age",
				"gender", "grower_type", "village_or_town", "address",
				"mobile_number", "estimated_production_of_made_tea", "estimated_production_of_green_tea",
				"production_area", "total_female_worker", "total_male_worker",
			)
			left_on_data = [
				"supplier_name", "region", "state", "district",
				"address", "phone_number", "sex", "male_worker", 
				"female_worker", "supplier_code", 
			]
			right_on_data = [
				"name", "region__region_name", "state__name", "district__name",
				"address", "mobile_number", "gender", "total_male_worker", "total_female_worker", 
				"id",
			]
		profiles_df = pd.DataFrame(list(profiles))
		matched = fuzzymatcher.link_table(
			api_df,
			profiles_df,
			left_on=left_on_data,
			right_on=right_on_data
		)
		matched = matched.sort_values("match_score", ascending=False).drop_duplicates("supplier_name")
		for _, row in matched.iterrows():
			raw_score = row.get("match_score", 0.0)
			if pd.isna(raw_score):
				raw_score = 0.0
			match_score = round(float(raw_score) * 100, 2)
			user_update = "false"
			if match_score >= 100:
				profile_id = row.get("id")
				model_name.cmobjects.filter(id=profile_id).update(
				    is_tcms_user=True,
				    tcms_unit_id=row.get("unit_id", ""),
				)
				user_update = "true"
				updated_count += 1
			results.append({
				"tracetea_username": row.get("user__username", ""),
				"tcms_username": row.get("user_name", ""),
				"tracetea_supplier_name": row.get("name", ""),
				"tcms_supplier_name": row.get("supplier_name", ""),
				"tracetea_region": row.get("region__region_name", ""),
				"tcms_region": row.get("region", ""),
				"tracetea_state": row.get("state__name", ""),
				"tcms_state": row.get("state", ""),
				"tracetea_district": row.get("district__name", ""),
				"tcms_district": row.get("district", ""),
				"tracetea_phone_number": row.get("mobile_number", ""),
				"tcms_phone_number": row.get("phone_number", ""),
				"tracetea_male_worker": row.get("mobile_number", ""),
				"tcms_male_worker": row.get("male_worker", ""),
				"tracetea_female_worker": row.get("female_worker", ""),
				"tcms_female_worker": row.get("female_worker", ""),
				"match_score": sanitize_float(match_score),
				"user_update": user_update
			})
		return Response({
			"count": len(results),
			"updated_count": updated_count,
			"results": results,
		})



@permission_required_admin
def users_profile_list(request, usertype_slug):
	user_type = usertype_slug
	if not user_type:
		raise APIException({'request_status': 0, 'msg': "User type required"})
	search = {}
	search = custom_web_filters(request, search, [])
	all = request.GET.get('all', None)
	order_by = request.GET.get('order_by', '-id')

	USER_TYPE_CONFIG = {
		"grower": {
			"model": GrowerProfile,
			"serializer": GrowerProfileSerializer,
			"select_related" : ['state', 'user', 'region', 'district', 'profile_type'],
			"prefetch_related" : ["associated_aggregator", "associated_entity"],
			"edit_url" : 'user_profile:grower_profile_edit',
			"values": [
				'id','user__username', 'is_tcms_user', 'tcms_supplier_code',
				'name', 'age', 'gender', 'date_of_birth', 'grower_type',
				'region__region_name', 'state__name', 'district__name',
				'region', 'state', 'district',
				'email', 'mobile_number', 'aadhar_no', 'address', 'tea_board_id', 'associated_aggregator', 
				'associated_entity', 'poi_type', 'poi_id',
			],
		},
		"aggregator": {
			"model": AggregatorProfile,
			"serializer": AggregatorProfileSerializer,
			"select_related" : ['state', 'user', 'region', 'district', 'profile_type'],
			"prefetch_related" : [],
			"values": [],
			"edit_url" : 'user_profile:aggregator_profile_edit',
		},
		"blf": {
			"model": BlfProfile,
			"serializer": BlfProfileSerializer,
			"select_related" : ['state', 'user', 'region', 'district', 'profile_type'],
			"prefetch_related" : [],
			"values": [],
			"edit_url" : 'user_profile:blf_profile_edit'
		},
	}
	config = USER_TYPE_CONFIG.get(user_type)
	if not config:
		raise APIException({'request_status': 0, 'msg': "Invalid user_type"})
	model = config["model"]
	values = config["values"] 
	edit_url = config["edit_url"] 

	_list = model.cmobjects.select_related(*config["select_related"]).prefetch_related(*config["prefetch_related"]).values(*values).distinct()
	_list = _list.filter(*search).order_by(*str(order_by).split(","))
	
	tot_count = _list.count()
	page = request.GET.get('page', 1)
	paginator = Paginator(_list, settings.MIN_PAGE_SIZE)
	try:
		_list = paginator.page(page)
	except PageNotAnInteger:
		_list = paginator.page(1)
	except EmptyPage:
		_list = paginator.page(paginator.num_pages)
	context = {
		'tot_count' : tot_count,
		'user_type' : usertype_slug,
		'edit_url' : edit_url,
		'count': paginator.count,
		'profile_list' : _list,
	}
	return render(request, 'profile/user_profile_lists.html', context)










class GrowerProfileALLUpdateAPIView(APIView):
    permission_classes = (AllowAny,)

    def put(self, request):
        with transaction.atomic():
            triggered_count = 0
            for grower in GrowerProfile.objects.all():
                post_save.send(sender=GrowerProfile, instance=grower, created=False)
                triggered_count += 1

        return Response({
            'results': {'Data': triggered_count},
            'msg': "Signals manually triggered",
            'status': status.HTTP_202_ACCEPTED,
            'request_status': 1
        })




				
